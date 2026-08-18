"""Movimentações da folha: quem mudou de centro de custo, salário ou função.

Compara cada competência importada com a anterior em que a pessoa aparece.
Nada é cadastrado aqui — a movimentação é deduzida dos extratos já importados.
"""

from __future__ import annotations

import re
from io import BytesIO

from datetime import datetime

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import engine, get_db
from app.template_config import templates
from app.utils import empresa_curta, redirect_with_message

router = APIRouter(tags=["folha_movimentacoes"])

TIPOS = {
    "centro_custo": "Centro de custo",
    "cargo": "Função",
    "salario": "Salário",
}

# Motivos sugeridos por tipo de alteração. "Outros" abre o campo de texto.
MOTIVOS = {
    "salario": [
        "Promoção", "Mérito / desempenho", "Reajuste coletivo / sindicato",
        "Dissídio / data-base", "Equiparação salarial", "Ajuste de mercado",
        "Retenção de profissional", "Aumento por tempo de serviço",
        "Alteração de jornada/carga horária", "Correção cadastral",
        "Reestruturação salarial", "Negociação individual", "Outros",
    ],
    "cargo": [
        "Promoção", "Progressão de carreira", "Rebaixamento / reenquadramento",
        "Mudança lateral", "Substituição de função", "Retorno à função anterior",
        "Adequação às atividades exercidas", "Reestruturação organizacional",
        "Alteração de nomenclatura do cargo", "Mudança de carreira/trilha",
        "Correção cadastral", "Outros",
    ],
    "centro_custo": [
        "Transferência interna", "Reestruturação organizacional",
        "Necessidade operacional", "Movimentação a pedido do colaborador",
        "Movimentação a pedido da gestão", "Promoção",
        "Mudança de projeto/contrato", "Mudança de unidade/filial",
        "Centralização de atividades", "Descentralização de atividades",
        "Correção cadastral", "Outros",
    ],
}


def _garantir_tabela_motivos():
    """Guarda o motivo de cada movimentação.

    A movimentação em si é deduzida dos extratos e não existe como registro,
    então a chave é o que a identifica: pessoa, empresa, competência e tipo.
    """
    try:
        with engine.begin() as conn:
            serial = ("INTEGER PRIMARY KEY AUTOINCREMENT"
                      if conn.dialect.name == "sqlite" else "SERIAL PRIMARY KEY")
            conn.execute(text(f"""
                CREATE TABLE IF NOT EXISTS folha_movimentacao_motivos (
                    id {serial},
                    matricula VARCHAR(40) NOT NULL,
                    empresa VARCHAR(160) NOT NULL,
                    competencia VARCHAR(10) NOT NULL,
                    tipo VARCHAR(30) NOT NULL,
                    motivo VARCHAR(120),
                    observacao TEXT,
                    atualizado_em TIMESTAMP
                )
            """))
            conn.execute(text(
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_folha_mov_motivo "
                "ON folha_movimentacao_motivos (matricula, empresa, competencia, tipo)"
            ))
    except Exception as exc:
        print(f"AVISO - tabela de motivos de movimentação: {exc}")


_garantir_tabela_motivos()


def _chave_motivo(m: dict) -> str:
    return f"{m['matricula']}|{m['empresa']}|{m['competencia']}|{m['tipo']}"


def _motivos_gravados(db: Session) -> dict[str, dict]:
    try:
        return {
            f"{r['matricula']}|{r['empresa']}|{r['competencia']}|{r['tipo']}": dict(r)
            for r in db.execute(text(
                "SELECT matricula, empresa, competencia, tipo, motivo, observacao "
                "FROM folha_movimentacao_motivos")).mappings()
        }
    except Exception as exc:
        print(f"AVISO - não consegui ler os motivos: {exc}")
        return {}


def _ordem_competencia(competencia: str) -> tuple:
    """'03/2026' → (2026, 3). Competência é texto e não ordena sozinha."""
    partes = str(competencia or "").split("/")
    if len(partes) == 2 and partes[0].isdigit() and partes[1].isdigit():
        return (int(partes[1]), int(partes[0]))
    return (0, 0)


def _e_mensal(registro) -> bool:
    """A folha mensal é a referência da posição; adiantamento e 13º não são."""
    tipo = str(registro.get("tipo_calculo") or "").lower()
    return "mensal" in tipo or not tipo


def _nomes_centros(db: Session) -> dict[str, str]:
    try:
        return {
            (r["codigo"] or "").strip(): r["nome"]
            for r in db.execute(text(
                "SELECT codigo, nome FROM folha_centros_custo")).mappings()
        }
    except Exception:
        return {}


def _rotulo_cc(mapa: dict[str, str], codigo) -> str:
    cod = str(codigo or "").strip()
    if not cod:
        return "—"
    nome = mapa.get(cod)
    return f"{cod} · {nome}" if nome else cod


def _movimentacoes(db: Session) -> list[dict]:
    """Toda mudança de CC, função ou salário entre competências seguidas."""
    linhas = db.execute(text("""
        SELECT a.empresa_nome AS empresa, a.tipo_calculo, f.competencia,
               f.matricula, f.nome, f.centro_custo, f.cargo, f.salario
        FROM folha_funcionarios f
        JOIN folha_arquivos a ON a.id_arquivo = f.id_arquivo
        WHERE COALESCE(f.matricula, '') <> ''
    """)).mappings().all()

    mapa_cc = _nomes_centros(db)

    # a chave é matrícula + empresa: a mesma matrícula pode existir nas duas
    #
    # Uma competência pode ter mais de um cálculo (mensal, adiantamento, 13º) e
    # a pessoa aparece em cada um. Comparar esses registros entre si inventaria
    # movimentação: o salário do adiantamento contra o da mensal viraria um
    # "aumento" que não houve. Fica um registro por competência, preferindo a
    # folha mensal, que é a que descreve a posição da pessoa no mês.
    por_competencia: dict[tuple, dict] = {}
    for r in linhas:
        chave = (str(r["matricula"]).strip(), r["empresa"], r["competencia"])
        atual = por_competencia.get(chave)
        if atual is None or (not _e_mensal(atual) and _e_mensal(r)):
            por_competencia[chave] = dict(r)

    por_pessoa: dict[tuple, list] = {}
    for (matricula, empresa, _), registro in por_competencia.items():
        por_pessoa.setdefault((matricula, empresa), []).append(registro)

    movimentos: list[dict] = []
    for (matricula, empresa), registros in por_pessoa.items():
        registros.sort(key=lambda x: _ordem_competencia(x["competencia"]))
        for anterior, atual in zip(registros, registros[1:]):
            base = {
                "matricula": matricula,
                "nome": atual["nome"],
                "empresa": empresa,
                "empresa_curta": empresa_curta(empresa),
                "competencia": atual["competencia"],
                "competencia_anterior": anterior["competencia"],
                "ordem": _ordem_competencia(atual["competencia"]),
            }

            cc_de, cc_para = (str(anterior["centro_custo"] or "").strip(),
                              str(atual["centro_custo"] or "").strip())
            if cc_de != cc_para:
                movimentos.append({**base, "tipo": "centro_custo",
                                   "de": _rotulo_cc(mapa_cc, cc_de),
                                   "para": _rotulo_cc(mapa_cc, cc_para),
                                   "variacao": None})

            cargo_de = str(anterior["cargo"] or "").strip()
            cargo_para = str(atual["cargo"] or "").strip()
            if cargo_de != cargo_para:
                movimentos.append({**base, "tipo": "cargo",
                                   "de": cargo_de or "—", "para": cargo_para or "—",
                                   "variacao": None})

            sal_de = float(anterior["salario"] or 0)
            sal_para = float(atual["salario"] or 0)
            # centavos de arredondamento não são movimentação
            if abs(sal_para - sal_de) >= 0.01:
                movimentos.append({
                    **base, "tipo": "salario",
                    "de": f"{sal_de:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    "para": f"{sal_para:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    "variacao": ((sal_para - sal_de) / sal_de * 100) if sal_de else None,
                })

    movimentos.sort(key=lambda m: (m["ordem"], m["nome"], m["tipo"]), reverse=True)
    return movimentos


def _filtrar(movimentos: list[dict], competencia: str, empresa: str,
             tipo: str, busca: str, justificado: str = "") -> list[dict]:
    resultado = movimentos
    if competencia:
        resultado = [m for m in resultado if m["competencia"] == competencia]
    if empresa:
        resultado = [m for m in resultado if m["empresa"] == empresa]
    if tipo:
        resultado = [m for m in resultado if m["tipo"] == tipo]
    if busca:
        alvo = busca.strip().lower()
        resultado = [m for m in resultado
                     if alvo in m["nome"].lower() or alvo in m["matricula"].lower()]
    if justificado == "sim":
        resultado = [m for m in resultado if m.get("motivo")]
    elif justificado == "nao":
        resultado = [m for m in resultado if not m.get("motivo")]
    return resultado


@router.get("/folha/movimentacoes")
def index(request: Request, competencia: str = "", empresa: str = "",
          tipo: str = "", q: str = "", justificado: str = "nao",
          db: Session = Depends(get_db)):
    try:
        todos = _movimentacoes(db)
    except Exception as exc:
        print(f"AVISO - não consegui apurar movimentações: {exc}")
        todos = []

    gravados = _motivos_gravados(db)
    for m in todos:
        salvo = gravados.get(_chave_motivo(m), {})
        m["chave"] = _chave_motivo(m)
        m["motivo"] = salvo.get("motivo") or ""
        m["observacao"] = salvo.get("observacao") or ""

    competencias = sorted({m["competencia"] for m in todos},
                          key=_ordem_competencia, reverse=True)
    empresas = sorted({m["empresa"] for m in todos})

    # os contadores olham o recorte sem o filtro de justificado: senão o de
    # pendentes some justamente quando se escolhe "justificadas"
    do_recorte = _filtrar(todos, competencia, empresa, tipo, q)
    movimentos = _filtrar(todos, competencia, empresa, tipo, q, justificado)

    resumo = {chave: sum(1 for m in movimentos if m["tipo"] == chave) for chave in TIPOS}
    sem_motivo = sum(1 for m in do_recorte if not m["motivo"])
    com_motivo = sum(1 for m in do_recorte if m["motivo"])

    return templates.TemplateResponse("folha/movimentacoes.html", {
        "request": request,
        "movimentos": movimentos,
        "total": len(movimentos),
        "total_geral": len(todos),
        "resumo": resumo,
        "sem_motivo": sem_motivo,
        "com_motivo": com_motivo,
        "justificado_sel": justificado,
        "tipos": TIPOS,
        "motivos": MOTIVOS,
        "competencias": competencias,
        "empresas": empresas,
        "competencia_sel": competencia,
        "empresa_sel": empresa,
        "tipo_sel": tipo,
        "q": q,
        "tem_filtro": bool(competencia or empresa or tipo or q or justificado),
    })


@router.post("/folha/movimentacoes/motivo")
def salvar_motivo(
    request: Request,
    matricula: str = Form(...),
    empresa: str = Form(...),
    competencia: str = Form(...),
    tipo: str = Form(...),
    motivo: str = Form(""),
    observacao: str = Form(""),
    db: Session = Depends(get_db),
):
    motivo = motivo.strip()
    # a observação só faz sentido junto de "Outros"; trocar o motivo a descarta
    observacao = observacao.strip() if motivo == "Outros" else ""

    dados = {"m": matricula.strip(), "e": empresa.strip(),
             "c": competencia.strip(), "t": tipo.strip(),
             "mo": motivo or None, "ob": observacao or None,
             "ts": datetime.utcnow()}

    existe = db.execute(text(
        "SELECT id FROM folha_movimentacao_motivos "
        "WHERE matricula = :m AND empresa = :e AND competencia = :c AND tipo = :t"
    ), dados).first()

    if existe:
        db.execute(text(
            "UPDATE folha_movimentacao_motivos SET motivo = :mo, observacao = :ob, "
            "atualizado_em = :ts WHERE id = :id"
        ), {**dados, "id": existe[0]})
    else:
        db.execute(text(
            "INSERT INTO folha_movimentacao_motivos "
            "(matricula, empresa, competencia, tipo, motivo, observacao, atualizado_em) "
            "VALUES (:m, :e, :c, :t, :mo, :ob, :ts)"
        ), dados)
    db.commit()

    if request.headers.get("x-fetch") == "1":
        return JSONResponse({"ok": True})
    return redirect_with_message("/folha/movimentacoes", success="Motivo salvo.")


@router.get("/folha/movimentacoes/exportar")
def exportar(request: Request, competencia: str = "", empresa: str = "",
             tipo: str = "", q: str = "", justificado: str = "",
             db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    todos = _movimentacoes(db)
    gravados = _motivos_gravados(db)
    for m in todos:
        salvo = gravados.get(_chave_motivo(m), {})
        m["motivo"] = salvo.get("motivo") or ""
        m["observacao"] = salvo.get("observacao") or ""
    movimentos = _filtrar(todos, competencia, empresa, tipo, q, justificado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações"
    colunas = ["Matrícula", "Nome", "Empresa", "Competência", "Competência anterior",
               "Tipo", "De", "Para", "Variação %", "Motivo", "Observação"]
    ws.append(colunas)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for m in movimentos:
        ws.append([m["matricula"], m["nome"], m["empresa_curta"], m["competencia"],
                   m["competencia_anterior"], TIPOS.get(m["tipo"], m["tipo"]),
                   m["de"], m["para"],
                   round(m["variacao"], 2) if m["variacao"] is not None else None,
                   m.get("motivo") or "", m.get("observacao") or ""])

    for i, titulo in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(titulo) + 6)
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    sufixo = re.sub(r"[^0-9]", "", competencia) or "todas"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="movimentacoes_folha_{sufixo}.xlsx"'},
    )
