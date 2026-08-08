from pathlib import Path
import os
import json
import unicodedata
from io import BytesIO
from datetime import datetime

from catworld import CatworldClient
from catworld.exceptions import (
    AuthenticationError,
    ConnectionError as CatworldConnectionError,
    PermissionDeniedError,
    QueryTimeoutError,
    ValidationError,
)
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi import APIRouter, Depends, Form, Request, UploadFile, File
from sqlalchemy import inspect
from sqlalchemy import (
    Column,
    DateTime,
    Float,
    Integer,
    MetaData,
    String,
    Table,
    Text,
    delete,
    insert,
    select,
    text,
    update,
)
from sqlalchemy.orm import Session

from app.auth import tem_acesso_modulo, is_admin as _is_admin
from app.database import engine, get_db
from app.models import DimAutonomo
from app.template_config import templates
from app.utils import flash_from_request, redirect_with_message

router = APIRouter(tags=["dho"])

BASE = Path(__file__).resolve().parents[2]

metadata_dho = MetaData()

dho_departamentos = Table(
    "dho_departamentos",
    metadata_dho,
    Column("id_departamento", Integer, primary_key=True, autoincrement=True),
    Column("nome_departamento", String(160), nullable=False),
    Column("status", String(40), default="Ativo"),
    Column("criado_em", DateTime, default=datetime.utcnow),
    Column("atualizado_em", DateTime, default=datetime.utcnow),
)

dho_cargos = Table(
    "dho_cargos",
    metadata_dho,
    Column("id_cargo", Integer, primary_key=True, autoincrement=True),
    Column("id_departamento", Integer, nullable=True),
    Column("nome_cargo", String(160), nullable=False),
    Column("status", String(40), default="Ativo"),
    Column("criado_em", DateTime, default=datetime.utcnow),
    Column("atualizado_em", DateTime, default=datetime.utcnow),
)

dho_vagas = Table(
    "dho_vagas",
    metadata_dho,
    Column("id_vaga", Integer, primary_key=True, autoincrement=True),
    Column("id_departamento", Integer, nullable=True),
    Column("id_cargo", Integer, nullable=True),
    Column("tipo_vaga", String(80), default="Nova vaga"),
    Column("motivo_substituicao", String(80)),
    Column("sexo", String(40), default="Indiferente"),
    Column("qtd_vagas", Integer, default=1),
    Column("status", String(60), default="Aberta"),
    Column("responsavel", String(160)),
    Column("data_abertura", String(20)),
    Column("data_conclusao", String(20)),
    Column("tipo_recrutamento", String(40), default="Externo"),
    Column("observacoes", Text),
    Column("criado_em", DateTime, default=datetime.utcnow),
    Column("atualizado_em", DateTime, default=datetime.utcnow),
)

dho_treinamentos = Table(
    "dho_treinamentos",
    metadata_dho,
    Column("id_treinamento", Integer, primary_key=True, autoincrement=True),
    Column("nome_treinamento", String(180), nullable=False),
    Column("tipo_treinamento", String(120), nullable=False),
    Column("carga_horaria_padrao", Float),
    Column("valor_treinamento", Float),
    Column("status", String(60), default="Ativo"),
    Column("observacoes", Text),
    Column("criado_em", DateTime, default=datetime.utcnow),
    Column("atualizado_em", DateTime, default=datetime.utcnow),
)

dho_treinamento_aplicacoes = Table(
    "dho_treinamento_aplicacoes",
    metadata_dho,
    Column("id_aplicacao", Integer, primary_key=True, autoincrement=True),
    Column("id_treinamento", Integer, nullable=False),
    Column("tipo_pessoa", String(60), default="Autônomo"),
    Column("id_autonomo", Integer, nullable=True),
    Column("pessoa_nome", String(180)),
    Column("matricula", String(80)),
    Column("funcao", String(180)),
    Column("centro_custo", String(180)),
    Column("data_treinamento", String(20)),
    Column("carga_horaria", Float),
    Column("status", String(60), default="Realizado"),
    Column("observacoes", Text),
    Column("criado_em", DateTime, default=datetime.utcnow),
    Column("atualizado_em", DateTime, default=datetime.utcnow),
)


def _separar_valores_multiplos(valor):
    texto = str(valor or "").replace("\r\n", "\n").replace("\r", "\n")
    return [parte.strip() for parte in texto.replace("\n", ";").split(";") if parte.strip()]


def _expandir_aplicacao_por_pessoa(dados):
    nomes = _separar_valores_multiplos(dados.get("pessoa_nome"))
    if not nomes:
        return [dict(dados)]

    campos_individuais = ("matricula", "funcao", "centro_custo")
    valores_individuais = {
        campo: _separar_valores_multiplos(dados.get(campo))
        for campo in campos_individuais
    }

    aplicacoes = []
    for indice, nome in enumerate(nomes):
        item = dict(dados)
        item["pessoa_nome"] = nome
        for campo in campos_individuais:
            partes = valores_individuais[campo]
            if len(partes) == len(nomes):
                item[campo] = partes[indice]
        aplicacoes.append(item)
    return aplicacoes


def _normalizar_aplicacoes_multiplas(conn, tabela=dho_treinamento_aplicacoes):
    """Converte registros antigos com varios nomes em uma linha por pessoa."""
    normalizados = 0
    registros = conn.execute(select(tabela)).mappings().all()

    for registro in registros:
        aplicacoes = _expandir_aplicacao_por_pessoa(dict(registro))
        if len(aplicacoes) <= 1:
            continue

        id_aplicacao = registro["id_aplicacao"]
        primeira = dict(aplicacoes[0])
        primeira.pop("id_aplicacao", None)
        conn.execute(
            update(tabela)
            .where(tabela.c.id_aplicacao == id_aplicacao)
            .values(**primeira)
        )

        for aplicacao in aplicacoes[1:]:
            nova = dict(aplicacao)
            nova.pop("id_aplicacao", None)
            conn.execute(insert(tabela).values(**nova))
            normalizados += 1

    return normalizados


dho_empregados = Table(
    "dho_empregados",
    metadata_dho,
    Column("id_empregado", Integer, primary_key=True, autoincrement=True),
    Column("matricula", String(80), nullable=True),
    Column("nome", String(200), nullable=False),
    Column("email", String(200), nullable=True),
    Column("id_departamento", Integer, nullable=True),
    Column("id_cargo", Integer, nullable=True),
    Column("data_admissao", String(20), nullable=True),
    Column("data_desligamento", String(20), nullable=True),
    Column("status", String(50), default="Ativo"),
    Column("observacoes", Text, nullable=True),
    Column("foto_url", Text, nullable=True),
    # preenchido à mão: não vem da fonte oficial
    Column("numero_fornecedor", String(60), nullable=True),
    Column("criado_em", DateTime, default=datetime.utcnow),
    Column("atualizado_em", DateTime, default=datetime.utcnow),
)

metadata_dho.create_all(engine)


def garantir_schema_dho_treinamentos():
    """
    Corrige banco antigo do DHO quando o código novo espera colunas novas.

    Problema típico:
    - A tabela dho_treinamentos já existia antes.
    - create_all não altera tabela existente.
    - Então colunas como tipo_treinamento ou carga_horaria_padrao podem não existir.
    """
    try:
        with engine.begin() as conn:
            dialect = conn.dialect.name

            if dialect == "sqlite":
                tabelas = [
                    r[0]
                    for r in conn.execute(
                        text("SELECT name FROM sqlite_master WHERE type='table'")
                    ).fetchall()
                ]

                def cols(table):
                    return [
                        r[1]
                        for r in conn.execute(text(f"PRAGMA table_info({table})")).fetchall()
                    ]

                if "dho_treinamentos" in tabelas:
                    c = cols("dho_treinamentos")

                    add_cols = {
                        "tipo_treinamento": "TEXT DEFAULT 'Autônomo'",
                        "carga_horaria_padrao": "REAL",
                        "valor_treinamento": "REAL",
                        "status": "TEXT DEFAULT 'Ativo'",
                        "observacoes": "TEXT",
                        "criado_em": "DATETIME",
                        "atualizado_em": "DATETIME",
                    }

                    for col, sql_type in add_cols.items():
                        if col not in c:
                            conn.execute(
                                text(f"ALTER TABLE dho_treinamentos ADD COLUMN {col} {sql_type}")
                            )

                    conn.execute(text("""
                        UPDATE dho_treinamentos
                        SET tipo_treinamento = 'Autônomo'
                        WHERE tipo_treinamento IS NULL OR tipo_treinamento = ''
                    """))

                    conn.execute(text("""
                        UPDATE dho_treinamentos
                        SET status = 'Ativo'
                        WHERE status IS NULL OR status = ''
                    """))

                if "dho_treinamento_aplicacoes" in tabelas:
                    c = cols("dho_treinamento_aplicacoes")

                    add_cols = {
                        "id_treinamento": "INTEGER",
                        "tipo_pessoa": "TEXT DEFAULT 'Autônomo'",
                        "id_autonomo": "INTEGER",
                        "pessoa_nome": "TEXT",
                        "matricula": "TEXT",
                        "funcao": "TEXT",
                        "centro_custo": "TEXT",
                        "data_treinamento": "TEXT",
                        "carga_horaria": "REAL",
                        "status": "TEXT DEFAULT 'Realizado'",
                        "observacoes": "TEXT",
                        "criado_em": "DATETIME",
                        "atualizado_em": "DATETIME",
                    }

                    for col, sql_type in add_cols.items():
                        if col not in c:
                            conn.execute(
                                text(f"ALTER TABLE dho_treinamento_aplicacoes ADD COLUMN {col} {sql_type}")
                            )

            elif dialect == "postgresql":
                conn.execute(text("""
                    ALTER TABLE IF EXISTS dho_empregados
                    ADD COLUMN IF NOT EXISTS foto_url TEXT,
                    ADD COLUMN IF NOT EXISTS numero_fornecedor VARCHAR(60)
                """))

                conn.execute(text("""
                    ALTER TABLE IF EXISTS dho_treinamentos
                    ADD COLUMN IF NOT EXISTS tipo_treinamento VARCHAR(120) DEFAULT 'Autônomo',
                    ADD COLUMN IF NOT EXISTS carga_horaria_padrao DOUBLE PRECISION,
                    ADD COLUMN IF NOT EXISTS valor_treinamento DOUBLE PRECISION,
                    ADD COLUMN IF NOT EXISTS status VARCHAR(60) DEFAULT 'Ativo',
                    ADD COLUMN IF NOT EXISTS observacoes TEXT,
                    ADD COLUMN IF NOT EXISTS criado_em TIMESTAMP,
                    ADD COLUMN IF NOT EXISTS atualizado_em TIMESTAMP
                """))

                conn.execute(text("""
                    UPDATE dho_treinamentos
                    SET tipo_treinamento = 'Autônomo'
                    WHERE tipo_treinamento IS NULL OR tipo_treinamento = ''
                """))

                conn.execute(text("""
                    UPDATE dho_treinamentos
                    SET status = 'Ativo'
                    WHERE status IS NULL OR status = ''
                """))

                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS dho_treinamento_aplicacoes (
                        id_aplicacao SERIAL PRIMARY KEY,
                        id_treinamento INTEGER,
                        tipo_pessoa VARCHAR(60) DEFAULT 'Autônomo',
                        id_autonomo INTEGER,
                        pessoa_nome VARCHAR(180),
                        data_treinamento VARCHAR(20),
                        carga_horaria DOUBLE PRECISION,
                        status VARCHAR(60) DEFAULT 'Realizado',
                        observacoes TEXT,
                        criado_em TIMESTAMP,
                        atualizado_em TIMESTAMP
                    )
                """))

                conn.execute(text("""
                    ALTER TABLE IF EXISTS dho_treinamento_aplicacoes
                    ADD COLUMN IF NOT EXISTS id_treinamento INTEGER,
                    ADD COLUMN IF NOT EXISTS tipo_pessoa VARCHAR(60) DEFAULT 'Autônomo',
                    ADD COLUMN IF NOT EXISTS id_autonomo INTEGER,
                    ADD COLUMN IF NOT EXISTS pessoa_nome VARCHAR(180),
                    ADD COLUMN IF NOT EXISTS matricula VARCHAR(80),
                    ADD COLUMN IF NOT EXISTS funcao VARCHAR(180),
                    ADD COLUMN IF NOT EXISTS centro_custo VARCHAR(180),
                    ADD COLUMN IF NOT EXISTS data_treinamento VARCHAR(20),
                    ADD COLUMN IF NOT EXISTS carga_horaria DOUBLE PRECISION,
                    ADD COLUMN IF NOT EXISTS status VARCHAR(60) DEFAULT 'Realizado',
                    ADD COLUMN IF NOT EXISTS observacoes TEXT,
                    ADD COLUMN IF NOT EXISTS criado_em TIMESTAMP,
                    ADD COLUMN IF NOT EXISTS atualizado_em TIMESTAMP
                """))

            elif dialect == "mssql":
                def _col_exists(table, col):
                    return conn.execute(text(
                        "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                        f"WHERE TABLE_NAME='{table}' AND COLUMN_NAME='{col}'"
                    )).scalar() > 0

                for col, sql_type in {
                    "tipo_treinamento": "NVARCHAR(120) DEFAULT 'Autônomo'",
                    "carga_horaria_padrao": "FLOAT",
                    "valor_treinamento": "FLOAT",
                    "status": "NVARCHAR(60) DEFAULT 'Ativo'",
                    "observacoes": "NVARCHAR(MAX)",
                    "criado_em": "DATETIME",
                    "atualizado_em": "DATETIME",
                }.items():
                    if not _col_exists("dho_treinamentos", col):
                        conn.execute(text(f"ALTER TABLE dho_treinamentos ADD {col} {sql_type}"))

                for col, sql_type in {
                    "id_treinamento": "INT",
                    "tipo_pessoa": "NVARCHAR(60)",
                    "id_autonomo": "INT",
                    "pessoa_nome": "NVARCHAR(180)",
                    "matricula": "NVARCHAR(80)",
                    "funcao": "NVARCHAR(180)",
                    "centro_custo": "NVARCHAR(180)",
                    "data_treinamento": "NVARCHAR(20)",
                    "carga_horaria": "FLOAT",
                    "status": "NVARCHAR(60)",
                    "observacoes": "NVARCHAR(MAX)",
                    "criado_em": "DATETIME",
                    "atualizado_em": "DATETIME",
                }.items():
                    if not _col_exists("dho_treinamento_aplicacoes", col):
                        conn.execute(text(f"ALTER TABLE dho_treinamento_aplicacoes ADD {col} {sql_type}"))

    except Exception as exc:
        print(f"AVISO - erro ao ajustar schema de treinamentos DHO: {exc}")


garantir_schema_dho_treinamentos()

try:
    with engine.begin() as conn:
        total_normalizados = _normalizar_aplicacoes_multiplas(conn)
        if total_normalizados:
            print(
                f"OK - {total_normalizados} aplicacao(oes) de treinamento "
                "separada(s) em uma linha por colaborador."
            )
except Exception as exc:
    print(f"AVISO - nao consegui separar aplicacoes de treinamento antigas: {exc}")


TIPOS_TREINAMENTO = [
    "Onboarding obrigatório",
    "Ação de RH",
    "Autônomo",
    "Reciclagem de autônomos",
]

TIPOS_VAGA = ["Nova vaga", "Substituição"]
MOTIVOS_SUBSTITUICAO = ["", "Voluntário", "Involuntário"]
SEXOS = ["Indiferente", "Masculino", "Feminino"]
STATUS_VAGA = ["Aberta", "Em andamento", "Concluída", "Cancelada"]
TIPOS_RECRUTAMENTO = ["Interno", "Externo"]
STATUS_TREINAMENTO = ["Ativo", "Inativo"]
STATUS_APLICACAO = ["Realizado", "Pendente", "Cancelado"]


def _texto_comparavel(valor) -> str:
    texto = str(valor or "").strip()
    if not texto:
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return " ".join(texto.lower().split())


def _normalizar_valor_enumerado(valor, opcoes):
    chave = _texto_comparavel(valor)
    if not chave:
        return None

    for opcao in opcoes:
        if _texto_comparavel(opcao) == chave:
            return opcao
    return None


def garantir_schema():
    """
    create_all cria tabelas novas, mas não altera tabela existente.
    Esse bloco garante colunas novas no PostgreSQL e SQLite local.
    """
    try:
        with engine.begin() as conn:
            dialect = conn.dialect.name

            if dialect == "postgresql":
                conn.execute(text("""
                    ALTER TABLE IF EXISTS dho_empregados
                    ADD COLUMN IF NOT EXISTS foto_url TEXT
                """))

                conn.execute(text("""
                    ALTER TABLE IF EXISTS dho_vagas
                    ADD COLUMN IF NOT EXISTS id_departamento INTEGER,
                    ADD COLUMN IF NOT EXISTS id_cargo INTEGER,
                    ADD COLUMN IF NOT EXISTS tipo_vaga VARCHAR(80) DEFAULT 'Nova vaga',
                    ADD COLUMN IF NOT EXISTS motivo_substituicao VARCHAR(80),
                    ADD COLUMN IF NOT EXISTS sexo VARCHAR(40) DEFAULT 'Indiferente',
                    ADD COLUMN IF NOT EXISTS qtd_vagas INTEGER DEFAULT 1,
                    ADD COLUMN IF NOT EXISTS responsavel VARCHAR(160),
                    ADD COLUMN IF NOT EXISTS data_abertura VARCHAR(20),
                    ADD COLUMN IF NOT EXISTS data_conclusao VARCHAR(20),
                    ADD COLUMN IF NOT EXISTS tipo_recrutamento VARCHAR(40) DEFAULT 'Externo',
                    ADD COLUMN IF NOT EXISTS observacoes TEXT
                """))

                conn.execute(text("""
                    ALTER TABLE IF EXISTS dho_treinamentos
                    ADD COLUMN IF NOT EXISTS tipo_treinamento VARCHAR(120),
                    ADD COLUMN IF NOT EXISTS carga_horaria_padrao DOUBLE PRECISION,
                    ADD COLUMN IF NOT EXISTS valor_treinamento DOUBLE PRECISION,
                    ADD COLUMN IF NOT EXISTS status VARCHAR(60) DEFAULT 'Ativo',
                    ADD COLUMN IF NOT EXISTS observacoes TEXT
                """))

            elif dialect == "mssql":
                def _col_exists2(table, col):
                    return conn.execute(text(
                        "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                        f"WHERE TABLE_NAME='{table}' AND COLUMN_NAME='{col}'"
                    )).scalar() > 0

                for col, sql_type in {
                    "id_departamento": "INT",
                    "id_cargo": "INT",
                    "tipo_vaga": "NVARCHAR(80)",
                    "motivo_substituicao": "NVARCHAR(80)",
                    "sexo": "NVARCHAR(40)",
                    "qtd_vagas": "INT DEFAULT 1",
                    "responsavel": "NVARCHAR(160)",
                    "data_abertura": "NVARCHAR(20)",
                    "data_conclusao": "NVARCHAR(20)",
                    "tipo_recrutamento": "NVARCHAR(40)",
                    "observacoes": "NVARCHAR(MAX)",
                }.items():
                    if not _col_exists2("dho_vagas", col):
                        conn.execute(text(f"ALTER TABLE dho_vagas ADD {col} {sql_type}"))

                for col, sql_type in {
                    "tipo_treinamento": "NVARCHAR(120)",
                    "carga_horaria_padrao": "FLOAT",
                    "valor_treinamento": "FLOAT",
                    "status": "NVARCHAR(60)",
                    "observacoes": "NVARCHAR(MAX)",
                }.items():
                    if not _col_exists2("dho_treinamentos", col):
                        conn.execute(text(f"ALTER TABLE dho_treinamentos ADD {col} {sql_type}"))

            elif dialect == "sqlite":
                def cols(table):
                    return [r[1] for r in conn.execute(text(f"PRAGMA table_info({table})")).fetchall()]

                tabelas_existentes = [r[0] for r in conn.execute(text("SELECT name FROM sqlite_master WHERE type='table'")).fetchall()]

                if "dho_empregados" in tabelas_existentes:
                    c_emp = cols("dho_empregados")
                    if "foto_url" not in c_emp:
                        conn.execute(text("ALTER TABLE dho_empregados ADD COLUMN foto_url TEXT"))
                    if "numero_fornecedor" not in c_emp:
                        conn.execute(text(
                            "ALTER TABLE dho_empregados ADD COLUMN numero_fornecedor VARCHAR(60)"))

                if "dho_vagas" in tabelas_existentes:
                    c = cols("dho_vagas")
                    add_cols = {
                        "id_departamento": "INTEGER",
                        "id_cargo": "INTEGER",
                        "tipo_vaga": "TEXT DEFAULT 'Nova vaga'",
                        "motivo_substituicao": "TEXT",
                        "sexo": "TEXT DEFAULT 'Indiferente'",
                        "qtd_vagas": "INTEGER DEFAULT 1",
                        "responsavel": "TEXT",
                        "data_abertura": "TEXT",
                        "data_conclusao": "TEXT",
                        "tipo_recrutamento": "TEXT DEFAULT 'Externo'",
                        "observacoes": "TEXT",
                    }
                    for col, sql_type in add_cols.items():
                        if col not in c:
                            conn.execute(text(f"ALTER TABLE dho_vagas ADD COLUMN {col} {sql_type}"))

    except Exception as exc:
        print(f"AVISO - não consegui ajustar schema DHO automaticamente: {exc}")


garantir_schema()


def fmt_num(value):
    try:
        v = float(value or 0)
        if v.is_integer():
            return str(int(v))
        return str(v).replace(".", ",")
    except Exception:
        return str(value or "")


def fmt_data(value):
    """Converte yyyy-mm-dd para dd/mm/yyyy; retorna o valor original em outros formatos."""
    s = str(value or "").strip()
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
    return s or "-"


def to_float(value):
    value = str(value or "").strip().replace(",", ".")
    if not value:
        return None
    try:
        return float(value)
    except Exception:
        return None


def to_int_or_none(value):
    value = str(value or "").strip()
    if not value:
        return None
    return int(value)


def count_table(db: Session, table):
    try:
        return db.execute(select(text("count(*)")).select_from(table)).scalar() or 0
    except Exception:
        return 0


def get_departamentos(db: Session):
    try:
        sincronizar_estrutura_dataworld(db)
    except Exception as exc:
        print(f"AVISO - não consegui sincronizar departamentos Catworld: {exc}")

    return db.execute(
        select(dho_departamentos).order_by(dho_departamentos.c.nome_departamento)
    ).mappings().all()


def get_cargos(db: Session):
    try:
        sincronizar_estrutura_dataworld(db)
    except Exception as exc:
        print(f"AVISO - não consegui sincronizar cargos Catworld: {exc}")

    rows = db.execute(
        select(
            dho_cargos,
            dho_departamentos.c.nome_departamento,
        )
        .select_from(dho_cargos.outerjoin(dho_departamentos, dho_departamentos.c.id_departamento == dho_cargos.c.id_departamento))
        .order_by(dho_cargos.c.nome_cargo)
    ).mappings().all()

    return rows


def get_treinamentos(db: Session):
    try:
        return db.execute(
            select(dho_treinamentos).where(dho_treinamentos.c.status == "Ativo").order_by(dho_treinamentos.c.nome_treinamento)
        ).mappings().all()
    except Exception as exc:
        print(f"AVISO - não consegui listar treinamentos ativos DHO: {exc}")
        return []


def get_autonomos(db: Session):
    try:
        return db.query(DimAutonomo).order_by(DimAutonomo.nome_autonomo).all()
    except Exception:
        return []


@router.get("/dho")
def dho_home(request: Request, db: Session = Depends(get_db)):
    if not tem_acesso_modulo(request, "dho"):
        return RedirectResponse("/?sem_acesso=dho", status_code=303)

    qtd_vagas = count_table(db, dho_vagas)
    qtd_treinamentos = count_table(db, dho_treinamentos)
    qtd_aplicacoes = count_table(db, dho_treinamento_aplicacoes)
    qtd_departamentos = count_table(db, dho_departamentos)
    qtd_cargos = count_table(db, dho_cargos)

    vagas_abertas = db.execute(
        select(text("count(*)")).select_from(dho_vagas).where(dho_vagas.c.status == "Aberta")
    ).scalar() or 0

    treinamentos_ativos = db.execute(
        select(text("count(*)")).select_from(dho_treinamentos).where(dho_treinamentos.c.status == "Ativo")
    ).scalar() or 0

    horas_realizadas = db.execute(
        select(text("COALESCE(SUM(carga_horaria), 0)")).select_from(dho_treinamento_aplicacoes)
    ).scalar() or 0

    try:
        ultimas_vagas = db.execute(
            select(
                dho_vagas,
                dho_departamentos.c.nome_departamento,
                dho_cargos.c.nome_cargo,
            )
            .select_from(
                dho_vagas
                .outerjoin(dho_departamentos, dho_departamentos.c.id_departamento == dho_vagas.c.id_departamento)
                .outerjoin(dho_cargos, dho_cargos.c.id_cargo == dho_vagas.c.id_cargo)
            )
            .order_by(dho_vagas.c.id_vaga.desc())
            .limit(6)
        ).mappings().all()
    except Exception as exc:
        print(f"AVISO - não consegui listar últimas vagas DHO: {exc}")
        ultimas_vagas = []

    try:
        ultimas_aplicacoes = db.execute(
            select(
                dho_treinamento_aplicacoes,
                dho_treinamentos.c.nome_treinamento,
            )
            .select_from(
                dho_treinamento_aplicacoes.join(
                    dho_treinamentos,
                    dho_treinamentos.c.id_treinamento == dho_treinamento_aplicacoes.c.id_treinamento,
                )
            )
            .order_by(dho_treinamento_aplicacoes.c.id_aplicacao.desc())
            .limit(6)
        ).mappings().all()
    except Exception as exc:
        print(f"AVISO - não consegui listar últimas aplicações DHO: {exc}")
        ultimas_aplicacoes = []

    return templates.TemplateResponse(
        "dho/index.html",
        {
            "request": request,
            "cards": {
                "qtd_vagas": qtd_vagas,
                "vagas_abertas": vagas_abertas,
                "qtd_treinamentos": qtd_treinamentos,
                "treinamentos_ativos": treinamentos_ativos,
                "qtd_aplicacoes": qtd_aplicacoes,
                "horas_realizadas": fmt_num(horas_realizadas),
                "qtd_departamentos": qtd_departamentos,
                "qtd_cargos": qtd_cargos,
            },
            "ultimas_vagas": ultimas_vagas,
            "ultimas_aplicacoes": ultimas_aplicacoes,
            "fmt_num": fmt_num,
            **flash_from_request(request),
        },
    )


@router.get("/dho/estrutura")
def estrutura(request: Request, db: Session = Depends(get_db)):
    erro_dataworld = ""
    departamentos = get_departamentos(db)
    cargos = get_cargos(db)

    return templates.TemplateResponse(
        "dho/estrutura.html",
        {
            "request": request,
            "departamentos": departamentos,
            "cargos": cargos,
            "erro_dataworld": erro_dataworld,
            **flash_from_request(request),
        },
    )


@router.post("/dho/estrutura/sincronizar")
def sincronizar_estrutura(request: Request, db: Session = Depends(get_db)):
    if not _is_admin(request):
        return RedirectResponse("/?sem_acesso=dho", status_code=303)

    _CACHE_PESSOAS_DW["dados"] = None
    _CACHE_ESTRUTURA_DW["sincronizada"] = False

    try:
        sincronizar_estrutura_dataworld(db, force=True)
    except Exception as exc:
        return redirect_with_message(
            "/dho/estrutura",
            error=f"Não consegui sincronizar a fonte oficial: {exc}",
        )

    return redirect_with_message(
        "/dho/estrutura",
        success="Departamentos e cargos sincronizados pela tabela rel_colab_77.",
    )


@router.post("/dho/departamentos")
def salvar_departamento(
    id_departamento: str = Form(""),
    nome_departamento: str = Form(...),
    status: str = Form("Ativo"),
    db: Session = Depends(get_db),
):
    dados = {
        "nome_departamento": nome_departamento,
        "status": status,
        "atualizado_em": datetime.utcnow(),
    }

    id_departamento = str(id_departamento or "").strip()

    if id_departamento:
        db.execute(update(dho_departamentos).where(dho_departamentos.c.id_departamento == int(id_departamento)).values(**dados))
        msg = "Departamento atualizado."
    else:
        dados["criado_em"] = datetime.utcnow()
        db.execute(insert(dho_departamentos).values(**dados))
        msg = "Departamento cadastrado."

    db.commit()
    return redirect_with_message("/dho/estrutura", success=msg)


@router.post("/dho/cargos")
def salvar_cargo(
    id_cargo: str = Form(""),
    id_departamento: str = Form(""),
    nome_cargo: str = Form(...),
    status: str = Form("Ativo"),
    db: Session = Depends(get_db),
):
    dados = {
        "id_departamento": to_int_or_none(id_departamento),
        "nome_cargo": nome_cargo,
        "status": status,
        "atualizado_em": datetime.utcnow(),
    }

    id_cargo = str(id_cargo or "").strip()

    if id_cargo:
        db.execute(update(dho_cargos).where(dho_cargos.c.id_cargo == int(id_cargo)).values(**dados))
        msg = "Cargo atualizado."
    else:
        dados["criado_em"] = datetime.utcnow()
        db.execute(insert(dho_cargos).values(**dados))
        msg = "Cargo cadastrado."

    db.commit()
    return redirect_with_message("/dho/estrutura", success=msg)


@router.post("/dho/departamentos/{id_departamento}/excluir")
def excluir_departamento(id_departamento: int, db: Session = Depends(get_db)):
    db.execute(delete(dho_departamentos).where(dho_departamentos.c.id_departamento == id_departamento))
    db.commit()
    return redirect_with_message("/dho/estrutura", success="Departamento excluído.")


@router.post("/dho/cargos/{id_cargo}/excluir")
def excluir_cargo(id_cargo: int, db: Session = Depends(get_db)):
    db.execute(delete(dho_cargos).where(dho_cargos.c.id_cargo == id_cargo))
    db.commit()
    return redirect_with_message("/dho/estrutura", success="Cargo excluído.")


@router.get("/dho/vagas")
def vagas(request: Request, q: str = "", db: Session = Depends(get_db)):
    query = (
        select(
            dho_vagas,
            dho_departamentos.c.nome_departamento,
            dho_cargos.c.nome_cargo,
        )
        .select_from(
            dho_vagas
            .outerjoin(dho_departamentos, dho_departamentos.c.id_departamento == dho_vagas.c.id_departamento)
            .outerjoin(dho_cargos, dho_cargos.c.id_cargo == dho_vagas.c.id_cargo)
        )
    )

    if q:
        like = f"%{q}%"
        query = query.where(
            dho_departamentos.c.nome_departamento.like(like)
            | dho_cargos.c.nome_cargo.like(like)
            | dho_vagas.c.status.like(like)
            | dho_vagas.c.tipo_vaga.like(like)
        )

    try:
        items = db.execute(query.order_by(dho_vagas.c.id_vaga.desc())).mappings().all()
    except Exception as exc:
        print(f"AVISO - não consegui listar vagas DHO: {exc}")
        items = []

    return templates.TemplateResponse(
        "dho/vagas.html",
        {
            "request": request,
            "items": items,
            "q": q,
            "departamentos": get_departamentos(db),
            "cargos": get_cargos(db),
            "tipos_vaga": TIPOS_VAGA,
            "motivos_substituicao": MOTIVOS_SUBSTITUICAO,
            "sexos": SEXOS,
            "status_vaga": STATUS_VAGA,
            "tipos_recrutamento": TIPOS_RECRUTAMENTO,
            **flash_from_request(request),
        },
    )


@router.post("/dho/vagas")
def salvar_vaga(
    id_vaga: str = Form(""),
    id_departamento: str = Form(""),
    id_cargo: str = Form(""),
    tipo_vaga: str = Form("Nova vaga"),
    motivo_substituicao: str = Form(""),
    sexo: str = Form("Indiferente"),
    qtd_vagas: int = Form(1),
    status: str = Form("Aberta"),
    responsavel: str = Form(""),
    data_abertura: str = Form(""),
    data_conclusao: str = Form(""),
    tipo_recrutamento: str = Form("Externo"),
    observacoes: str = Form(""),
    db: Session = Depends(get_db),
):
    dados = {
        "id_departamento": to_int_or_none(id_departamento),
        "id_cargo": to_int_or_none(id_cargo),
        "tipo_vaga": tipo_vaga,
        "motivo_substituicao": motivo_substituicao if tipo_vaga == "Substituição" else "",
        "sexo": sexo,
        "qtd_vagas": qtd_vagas,
        "status": status,
        "responsavel": responsavel,
        "data_abertura": data_abertura,
        "data_conclusao": data_conclusao,
        "tipo_recrutamento": tipo_recrutamento if tipo_recrutamento in TIPOS_RECRUTAMENTO else "Externo",
        "observacoes": observacoes,
        "atualizado_em": datetime.utcnow(),
    }

    id_vaga = str(id_vaga or "").strip()

    if id_vaga:
        db.execute(update(dho_vagas).where(dho_vagas.c.id_vaga == int(id_vaga)).values(**dados))
        msg = "Vaga atualizada com sucesso."
    else:
        dados["criado_em"] = datetime.utcnow()
        db.execute(insert(dho_vagas).values(**dados))
        msg = "Vaga cadastrada com sucesso."

    db.commit()
    return redirect_with_message("/dho/vagas", success=msg)


@router.post("/dho/vagas/{id_vaga}/excluir")
def excluir_vaga(id_vaga: int, db: Session = Depends(get_db)):
    db.execute(delete(dho_vagas).where(dho_vagas.c.id_vaga == id_vaga))
    db.commit()
    return redirect_with_message("/dho/vagas", success="Vaga excluída.")


@router.get("/dho/treinamentos")
def treinamentos(request: Request, q: str = "", db: Session = Depends(get_db)):
    query = select(dho_treinamentos)

    if q:
        like = f"%{q}%"
        query = query.where(
            dho_treinamentos.c.nome_treinamento.like(like)
            | dho_treinamentos.c.tipo_treinamento.like(like)
            | dho_treinamentos.c.status.like(like)
        )

    try:
        items = db.execute(query.order_by(dho_treinamentos.c.nome_treinamento)).mappings().all()
    except Exception as exc:
        print(f"AVISO - não consegui listar treinamentos DHO: {exc}")
        items = []

    return templates.TemplateResponse(
        "dho/treinamentos.html",
        {
            "request": request,
            "items": items,
            "q": q,
            "tipos_treinamento": TIPOS_TREINAMENTO,
            "status_treinamento": STATUS_TREINAMENTO,
            "fmt_num": fmt_num,
            **flash_from_request(request),
        },
    )


@router.post("/dho/treinamentos")
def salvar_treinamento(
    id_treinamento: str = Form(""),
    nome_treinamento: str = Form(...),
    tipo_treinamento: str = Form(...),
    carga_horaria_padrao: str = Form(""),
    valor_treinamento: str = Form(""),
    status: str = Form("Ativo"),
    observacoes: str = Form(""),
    db: Session = Depends(get_db),
):
    if tipo_treinamento not in TIPOS_TREINAMENTO:
        tipo_treinamento = "Autônomo"

    dados = {
        "nome_treinamento": nome_treinamento,
        "tipo_treinamento": tipo_treinamento,
        "carga_horaria_padrao": to_float(carga_horaria_padrao),
        "valor_treinamento": to_float(valor_treinamento),
        "status": status,
        "observacoes": observacoes,
        "atualizado_em": datetime.utcnow(),
    }

    id_treinamento = str(id_treinamento or "").strip()

    if id_treinamento:
        db.execute(update(dho_treinamentos).where(dho_treinamentos.c.id_treinamento == int(id_treinamento)).values(**dados))
        msg = "Treinamento atualizado."
    else:
        dados["criado_em"] = datetime.utcnow()
        db.execute(insert(dho_treinamentos).values(**dados))
        msg = "Treinamento cadastrado."

    db.commit()
    return redirect_with_message("/dho/treinamentos", success=msg)


@router.post("/dho/treinamentos/{id_treinamento}/excluir")
def excluir_treinamento(id_treinamento: int, db: Session = Depends(get_db)):
    db.execute(delete(dho_treinamentos).where(dho_treinamentos.c.id_treinamento == id_treinamento))
    db.commit()
    return redirect_with_message("/dho/treinamentos", success="Treinamento excluído.")





CATWORLD_QUERY_ERRORS = (
    AuthenticationError,
    PermissionDeniedError,
    ValidationError,
    QueryTimeoutError,
    CatworldConnectionError,
)


def _catworld_client():
    base_url = os.getenv("CATWORLD_URL")
    token = os.getenv("CATWORLD_TOKEN")

    if not base_url or not token:
        raise RuntimeError("CATWORLD_URL e CATWORLD_TOKEN precisam estar configurados.")

    return CatworldClient(base_url=base_url, token=token)


def _catworld_dataset_id():
    dataset_id = os.getenv("CATWORLD_DATASET_ID")

    if not dataset_id:
        raise RuntimeError("CATWORLD_DATASET_ID precisa estar configurado.")

    return dataset_id


def _consultar_catworld(sql: str):
    client = _catworld_client()
    try:
        resultado = client.query(sql, dataset_id=_catworld_dataset_id())
        return resultado.rows or []
    finally:
        client.close()


_CACHE_PESSOAS_DW = {
    "dados": None,
}

_CACHE_ESTRUTURA_DW = {
    "sincronizada": False,
}


def carregar_pessoas_interno(db: Session):
    """Retorna empregados da tabela interna como lista de dicts compatíveis com o template."""
    try:
        rows = db.execute(
            select(
                dho_empregados,
                dho_departamentos.c.nome_departamento,
                dho_cargos.c.nome_cargo,
            )
            .select_from(
                dho_empregados
                .outerjoin(dho_departamentos, dho_departamentos.c.id_departamento == dho_empregados.c.id_departamento)
                .outerjoin(dho_cargos, dho_cargos.c.id_cargo == dho_empregados.c.id_cargo)
            )
            .where(dho_empregados.c.status == "Ativo")
            .order_by(dho_empregados.c.nome)
        ).mappings().all()
        return [
            {
                "nome": r["nome"],
                "nome_exibicao": r["nome"],
                "matricula": r["matricula"] or "",
                "email": r["email"] or "",
                "cargo": r["nome_cargo"] or "",
                "departamento": r["nome_departamento"] or "",
            }
            for r in rows
        ]
    except Exception as exc:
        print(f"AVISO - erro ao carregar empregados internos: {exc}")
        return []


def carregar_pessoas_dataworld(raise_errors: bool = False):
    """
    Busca pessoas na tabela rel_colab_77 do dataset Catworld configurado.

    Campos usados:
    - nome
    - nome_completo
    - matricula
    - email
    - cargo
    - departamento

    Retorna lista já ordenada alfabeticamente.
    """
    if _CACHE_PESSOAS_DW.get("dados") is not None:
        return _CACHE_PESSOAS_DW["dados"]

    sql = """
        SELECT *
        FROM rel_colab_77
    """

    pessoas = []

    try:
        rows = _consultar_catworld(sql)

        for row in rows:
            nome = str(row.get("nome") or "").strip()
            nome_completo = str(row.get("nome_completo") or row.get("colaborador") or row.get("funcionario") or "").strip()
            matricula = str(row.get("matricula") or row.get("chapa") or "").strip()
            email = str(row.get("email") or row.get("email_corporativo") or "").strip()
            cargo = str(row.get("cargo_visivel") or row.get("cargo") or row.get("funcao") or "").strip()
            departamento = str(row.get("departamento") or row.get("centro_custo") or row.get("setor") or "").strip()

            nome_exibicao = nome_completo or nome

            if not nome_exibicao:
                continue

            detalhes = []
            if matricula:
                detalhes.append(f"Matrícula {matricula}")
            if cargo:
                detalhes.append(f"Função {cargo}")
            if departamento:
                detalhes.append(f"Centro de custo {departamento}")

            pessoas.append({
                "nome": nome,
                "nome_completo": nome_completo,
                "nome_exibicao": nome_exibicao,
                "matricula": matricula,
                "email": email,
                "cargo": cargo,
                "departamento": departamento,
                "label": nome_exibicao + ((" - " + " · ".join(detalhes)) if detalhes else ""),
            })

    except (RuntimeError, *CATWORLD_QUERY_ERRORS) as exc:
        if raise_errors:
            raise
        print(f"AVISO - Catworld indisponível: {exc}")

    # Deduplica por nome + matrícula/email
    dedup = {}

    for p in pessoas:
        chave = (
            str(p.get("nome_exibicao") or "").strip().lower(),
            str(p.get("matricula") or "").strip().lower(),
            str(p.get("email") or "").strip().lower(),
        )

        if chave not in dedup:
            dedup[chave] = p

    pessoas = sorted(
        dedup.values(),
        key=lambda x: str(x.get("nome_exibicao") or "").lower()
    )

    if pessoas:
        _CACHE_PESSOAS_DW["dados"] = pessoas

    return pessoas


def carregar_pessoas_aplicacao_treinamento(db: Session):
    pessoas = carregar_pessoas_interno(db)
    dedup = {
        (
            str(p.get("nome_exibicao") or p.get("nome") or "").strip().lower(),
            str(p.get("matricula") or "").strip().lower(),
            str(p.get("email") or "").strip().lower(),
        ): p
        for p in pessoas
    }

    return sorted(
        dedup.values(),
        key=lambda x: str(x.get("nome_exibicao") or "").lower(),
    )


def _chave_empregado_catworld(pessoa):
    matricula = str(pessoa.get("matricula") or "").strip().lower()
    email = str(pessoa.get("email") or "").strip().lower()
    nome = str(pessoa.get("nome_exibicao") or pessoa.get("nome_completo") or pessoa.get("nome") or "").strip().lower()

    if matricula:
        return ("matricula", matricula)
    if email:
        return ("email", email)
    if nome:
        return ("nome", nome)
    return None


def sincronizar_empregados_dataworld(db: Session):
    _CACHE_PESSOAS_DW["dados"] = None
    pessoas = carregar_pessoas_dataworld(raise_errors=True)

    sincronizar_estrutura_dataworld(db, force=True)

    departamentos_por_nome = {}
    for row in db.execute(select(dho_departamentos)).mappings().all():
        nome = str(row.get("nome_departamento") or "").strip().lower()
        if nome:
            departamentos_por_nome[nome] = row.get("id_departamento")

    cargos_por_nome_departamento = {}
    for row in db.execute(select(dho_cargos)).mappings().all():
        nome_cargo = str(row.get("nome_cargo") or "").strip().lower()
        if nome_cargo:
            cargos_por_nome_departamento[(nome_cargo, row.get("id_departamento"))] = row.get("id_cargo")

    empregados_por_chave = {}
    for row in db.execute(select(dho_empregados)).mappings().all():
        chaves = []
        matricula = str(row.get("matricula") or "").strip().lower()
        email = str(row.get("email") or "").strip().lower()
        nome = str(row.get("nome") or "").strip().lower()
        if matricula:
            chaves.append(("matricula", matricula))
        if email:
            chaves.append(("email", email))
        if nome:
            chaves.append(("nome", nome))
        for chave in chaves:
            empregados_por_chave.setdefault(chave, row)

    vistos = set()
    criados = 0
    atualizados = 0
    agora = datetime.utcnow()

    for pessoa in pessoas:
        chave = _chave_empregado_catworld(pessoa)
        if not chave:
            continue

        nome = str(pessoa.get("nome_exibicao") or pessoa.get("nome_completo") or pessoa.get("nome") or "").strip()
        if not nome:
            continue

        id_departamento = departamentos_por_nome.get(str(pessoa.get("departamento") or "").strip().lower())
        id_cargo = cargos_por_nome_departamento.get((
            str(pessoa.get("cargo") or "").strip().lower(),
            id_departamento,
        ))

        dados = {
            "matricula": str(pessoa.get("matricula") or "").strip() or None,
            "nome": nome,
            "email": str(pessoa.get("email") or "").strip() or None,
            "id_departamento": id_departamento,
            "id_cargo": id_cargo,
            "status": "Ativo",
            "observacoes": "Sincronizado pelo Catworld rel_colab_77.",
            "atualizado_em": agora,
        }

        existente = empregados_por_chave.get(chave)
        if existente:
            db.execute(
                update(dho_empregados)
                .where(dho_empregados.c.id_empregado == existente["id_empregado"])
                .values(**dados)
            )
            vistos.add(existente["id_empregado"])
            atualizados += 1
        else:
            dados["criado_em"] = agora
            result = db.execute(insert(dho_empregados).values(**dados))
            if result.inserted_primary_key:
                vistos.add(result.inserted_primary_key[0])
            criados += 1

    inativados = 0
    if vistos:
        inativados = db.execute(
            update(dho_empregados)
            .where(dho_empregados.c.id_empregado.notin_(vistos))
            .where(dho_empregados.c.status == "Ativo")
            .values(status="Inativo", atualizado_em=agora)
        ).rowcount or 0

    db.commit()
    return {
        "total_catworld": len(pessoas),
        "criados": criados,
        "atualizados": atualizados,
        "inativados": inativados,
    }


def carregar_estrutura_dataworld():
    pessoas = carregar_pessoas_dataworld()
    departamentos = sorted({
        str(p.get("departamento") or "").strip()
        for p in pessoas
        if str(p.get("departamento") or "").strip()
    }, key=lambda x: x.lower())

    cargos = sorted({
        (
            str(p.get("cargo") or "").strip(),
            str(p.get("departamento") or "").strip(),
        )
        for p in pessoas
        if str(p.get("cargo") or "").strip()
    }, key=lambda x: (x[0].lower(), x[1].lower()))

    return departamentos, cargos


def sincronizar_estrutura_dataworld(db: Session, force: bool = False):
    if _CACHE_ESTRUTURA_DW.get("sincronizada") and not force:
        return

    departamentos_dw, cargos_dw = carregar_estrutura_dataworld()

    departamentos_por_nome = {}
    for row in db.execute(select(dho_departamentos)).mappings().all():
        nome = str(row.get("nome_departamento") or "").strip().lower()
        if nome:
            departamentos_por_nome[nome] = row.get("id_departamento")

    for nome_departamento in departamentos_dw:
        chave = nome_departamento.strip().lower()
        if not chave:
            continue

        if chave not in departamentos_por_nome:
            result = db.execute(
                insert(dho_departamentos).values(
                    nome_departamento=nome_departamento,
                    status="Ativo",
                    criado_em=datetime.utcnow(),
                    atualizado_em=datetime.utcnow(),
                )
            )
            departamentos_por_nome[chave] = result.inserted_primary_key[0] if result.inserted_primary_key else None

    cargos_existentes = set()
    for row in db.execute(select(dho_cargos)).mappings().all():
        cargos_existentes.add((
            str(row.get("nome_cargo") or "").strip().lower(),
            row.get("id_departamento"),
        ))

    for nome_cargo, nome_departamento in cargos_dw:
        id_departamento = departamentos_por_nome.get(nome_departamento.strip().lower())
        chave = (nome_cargo.strip().lower(), id_departamento)
        if not nome_cargo.strip() or chave in cargos_existentes:
            continue

        db.execute(
            insert(dho_cargos).values(
                id_departamento=id_departamento,
                nome_cargo=nome_cargo,
                status="Ativo",
                criado_em=datetime.utcnow(),
                atualizado_em=datetime.utcnow(),
            )
        )
        cargos_existentes.add(chave)

    db.commit()
    _CACHE_ESTRUTURA_DW["sincronizada"] = True



@router.get("/dho/aplicacoes-treinamento")
def aplicacoes_treinamento(request: Request, q: str = "", db: Session = Depends(get_db)):
    query = (
        select(
            dho_treinamento_aplicacoes,
            dho_treinamentos.c.nome_treinamento,
            dho_treinamentos.c.tipo_treinamento,
        )
        .select_from(
            dho_treinamento_aplicacoes.join(
                dho_treinamentos,
                dho_treinamentos.c.id_treinamento == dho_treinamento_aplicacoes.c.id_treinamento,
            )
        )
    )

    if q:
        like = f"%{q}%"
        query = query.where(
            dho_treinamento_aplicacoes.c.pessoa_nome.like(like)
            | dho_treinamentos.c.nome_treinamento.like(like)
            | dho_treinamentos.c.tipo_treinamento.like(like)
        )

    try:
        items = db.execute(query.order_by(dho_treinamento_aplicacoes.c.id_aplicacao.desc())).mappings().all()
        items = list(items)
    except Exception as exc:
        print(f"AVISO - não consegui listar aplicações DHO: {exc}")
        items = []

    # Enriquece registros quando matricula/funcao/centro_custo estão vazios
    pessoas_dw = carregar_pessoas_interno(db) or list(carregar_pessoas_dataworld())
    dw_por_nome = {
        str(p.get("nome_exibicao") or p.get("nome") or "").strip().lower(): p
        for p in pessoas_dw
    }
    items_enriquecidos = []
    for row in items:
        item = dict(row)
        if not item.get("matricula") or not item.get("funcao") or not item.get("centro_custo"):
            nome_key = str(item.get("pessoa_nome") or "").strip().lower()
            pessoa_dw = dw_por_nome.get(nome_key)
            if pessoa_dw:
                item.setdefault("matricula", None)
                item.setdefault("funcao", None)
                item.setdefault("centro_custo", None)
                if not item["matricula"]:
                    item["matricula"] = pessoa_dw.get("matricula") or ""
                if not item["funcao"]:
                    item["funcao"] = pessoa_dw.get("cargo") or ""
                if not item["centro_custo"]:
                    item["centro_custo"] = pessoa_dw.get("departamento") or ""
        items_enriquecidos.append(item)
    items = items_enriquecidos

    return templates.TemplateResponse(
        "dho/aplicacoes_treinamento.html",
        {
            "request": request,
            "items": items,
            "q": q,
            "treinamentos": get_treinamentos(db),
            "autonomos": get_autonomos(db),
            "pessoas_dataworld": carregar_pessoas_aplicacao_treinamento(db),
            "status_aplicacao": STATUS_APLICACAO,
            "fmt_num": fmt_num,
            "fmt_data": fmt_data,
            **flash_from_request(request),
        },
    )


@router.post("/dho/aplicacoes-treinamento")
def salvar_aplicacao_treinamento(
    id_aplicacao: str = Form(""),
    id_treinamento: int = Form(...),
    tipo_pessoa: str = Form("Autônomo"),
    id_autonomo: str = Form(""),
    pessoa_nome: str = Form(""),
    matricula: str = Form(""),
    funcao: str = Form(""),
    centro_custo: str = Form(""),
    data_treinamento: str = Form(""),
    carga_horaria: str = Form(""),
    status: str = Form("Realizado"),
    observacoes: str = Form(""),
    db: Session = Depends(get_db),
):
    id_autonomo_int = to_int_or_none(id_autonomo)

    if id_autonomo_int and not pessoa_nome:
        aut = db.get(DimAutonomo, id_autonomo_int)
        if aut:
            pessoa_nome = aut.nome_autonomo

    dados = {
        "id_treinamento": id_treinamento,
        "tipo_pessoa": tipo_pessoa,
        "id_autonomo": id_autonomo_int,
        "pessoa_nome": pessoa_nome,
        "matricula": matricula,
        "funcao": funcao,
        "centro_custo": centro_custo,
        "data_treinamento": data_treinamento,
        "carga_horaria": to_float(carga_horaria),
        "status": status,
        "observacoes": observacoes,
        "atualizado_em": datetime.utcnow(),
    }

    id_aplicacao = str(id_aplicacao or "").strip()

    aplicacoes = _expandir_aplicacao_por_pessoa(dados)

    if id_aplicacao:
        db.execute(
            update(dho_treinamento_aplicacoes)
            .where(dho_treinamento_aplicacoes.c.id_aplicacao == int(id_aplicacao))
            .values(**aplicacoes[0])
        )
        for aplicacao in aplicacoes[1:]:
            aplicacao["criado_em"] = datetime.utcnow()
            db.execute(insert(dho_treinamento_aplicacoes).values(**aplicacao))
        msg = "Aplicação atualizada."
    else:
        for aplicacao in aplicacoes:
            aplicacao["criado_em"] = datetime.utcnow()
            db.execute(insert(dho_treinamento_aplicacoes).values(**aplicacao))
        msg = (
            f"Treinamento aplicado para {len(aplicacoes)} colaboradores."
            if len(aplicacoes) > 1
            else "Treinamento aplicado."
        )

    db.commit()
    return redirect_with_message("/dho/aplicacoes-treinamento", success=msg)


@router.post("/dho/aplicacoes-treinamento/{id_aplicacao}/excluir")
def excluir_aplicacao_treinamento(id_aplicacao: int, db: Session = Depends(get_db)):
    db.execute(delete(dho_treinamento_aplicacoes).where(dho_treinamento_aplicacoes.c.id_aplicacao == id_aplicacao))
    db.commit()
    return redirect_with_message("/dho/aplicacoes-treinamento", success="Aplicação excluída.")



def _norm_col(nome):
    return str(nome or "").strip().lower().replace(" ", "_").replace("-", "_")


def _valor_linha(row, *nomes):
    for n in nomes:
        if n in row and row[n] not in (None, ""):
            return str(row[n]).strip()
    return ""


def _get_or_create_departamento(db: Session, nome_departamento: str):
    nome_departamento = str(nome_departamento or "").strip()

    if not nome_departamento:
        return None

    existente = db.execute(
        select(dho_departamentos.c.id_departamento)
        .where(dho_departamentos.c.nome_departamento == nome_departamento)
    ).scalar()

    if existente:
        return existente

    result = db.execute(
        insert(dho_departamentos).values(
            nome_departamento=nome_departamento,
            status="Ativo",
            criado_em=datetime.utcnow(),
            atualizado_em=datetime.utcnow(),
        )
    )

    db.flush()

    try:
        return result.inserted_primary_key[0]
    except Exception:
        return db.execute(
            select(dho_departamentos.c.id_departamento)
            .where(dho_departamentos.c.nome_departamento == nome_departamento)
        ).scalar()


@router.get("/dho/importacoes")
def dho_importacoes(request: Request):
    return templates.TemplateResponse(
        "dho/importacoes.html",
        {
            "request": request,
            **flash_from_request(request),
        },
    )


@router.post("/dho/importacoes")
async def importar_dho(
    request: Request,
    tipo_importacao: str = Form(...),
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not _is_admin(request):
        return RedirectResponse("/?sem_acesso=dho", status_code=303)
    try:
        from openpyxl import load_workbook

        conteudo = await arquivo.read()
        wb = load_workbook(BytesIO(conteudo), data_only=True)
        ws = wb.active

        linhas = list(ws.iter_rows(values_only=True))

        if not linhas:
            return redirect_with_message("/dho/importacoes", error="Arquivo vazio.")

        headers = [_norm_col(h) for h in linhas[0]]
        total = 0

        for vals in linhas[1:]:
            row = {headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))}

            if tipo_importacao == "departamentos":
                nome = _valor_linha(row, "nome_departamento", "departamento", "area", "área")
                status = _valor_linha(row, "status") or "Ativo"

                if not nome:
                    continue

                existente = db.execute(
                    select(dho_departamentos.c.id_departamento)
                    .where(dho_departamentos.c.nome_departamento == nome)
                ).scalar()

                if existente:
                    db.execute(
                        update(dho_departamentos)
                        .where(dho_departamentos.c.id_departamento == existente)
                        .values(status=status, atualizado_em=datetime.utcnow())
                    )
                else:
                    db.execute(
                        insert(dho_departamentos).values(
                            nome_departamento=nome,
                            status=status,
                            criado_em=datetime.utcnow(),
                            atualizado_em=datetime.utcnow(),
                        )
                    )

                total += 1

            elif tipo_importacao == "cargos":
                nome_cargo = _valor_linha(row, "nome_cargo", "cargo")
                status = _valor_linha(row, "status") or "Ativo"

                id_departamento = _valor_linha(row, "id_departamento")
                departamento_nome = _valor_linha(row, "nome_departamento", "departamento", "area", "área")

                if not nome_cargo:
                    continue

                id_dep = None

                if id_departamento:
                    try:
                        id_dep = int(float(id_departamento))
                    except Exception:
                        id_dep = None

                if not id_dep and departamento_nome:
                    id_dep = _get_or_create_departamento(db, departamento_nome)

                existente = db.execute(
                    select(dho_cargos.c.id_cargo)
                    .where(dho_cargos.c.nome_cargo == nome_cargo)
                ).scalar()

                if existente:
                    db.execute(
                        update(dho_cargos)
                        .where(dho_cargos.c.id_cargo == existente)
                        .values(
                            id_departamento=id_dep,
                            status=status,
                            atualizado_em=datetime.utcnow(),
                        )
                    )
                else:
                    db.execute(
                        insert(dho_cargos).values(
                            id_departamento=id_dep,
                            nome_cargo=nome_cargo,
                            status=status,
                            criado_em=datetime.utcnow(),
                            atualizado_em=datetime.utcnow(),
                        )
                    )

                total += 1

            elif tipo_importacao in ("cursos", "treinamentos", "cadastro_curso", "cadastro_treinamento"):
                nome = _valor_linha(row, "nome_treinamento", "nome_curso", "curso", "treinamento")
                tipo = _valor_linha(row, "tipo_treinamento", "tipo_curso", "tipo") or "Autônomo"
                carga = _valor_linha(row, "carga_horaria_padrao", "carga_horaria", "horas")
                valor = _valor_linha(row, "valor_treinamento", "valor", "custo")
                status = _valor_linha(row, "status") or "Ativo"
                observacoes = _valor_linha(row, "observacoes", "observação", "obs")

                if not nome:
                    continue

                try:
                    carga_float = float(str(carga).replace(",", ".")) if str(carga or "").strip() else None
                except Exception:
                    carga_float = None

                if "TIPOS_TREINAMENTO" in globals() and tipo not in TIPOS_TREINAMENTO:
                    tipo = "Autônomo"

                existente = db.execute(
                    select(dho_treinamentos.c.id_treinamento)
                    .where(dho_treinamentos.c.nome_treinamento == nome)
                ).scalar()

                dados_curso = {
                    "nome_treinamento": nome,
                    "tipo_treinamento": tipo,
                    "carga_horaria_padrao": carga_float,
                    "valor_treinamento": to_float(valor),
                    "status": status,
                    "observacoes": observacoes,
                    "atualizado_em": datetime.utcnow(),
                }

                if existente:
                    db.execute(
                        update(dho_treinamentos)
                        .where(dho_treinamentos.c.id_treinamento == existente)
                        .values(**dados_curso)
                    )
                else:
                    dados_curso["criado_em"] = datetime.utcnow()
                    db.execute(insert(dho_treinamentos).values(**dados_curso))

                total += 1

            elif tipo_importacao == "vagas":
                departamento_nome = _valor_linha(row, "nome_departamento", "departamento", "area", "área")
                nome_cargo = _valor_linha(row, "nome_cargo", "cargo", "funcao", "função")
                tipo_vaga = _normalizar_valor_enumerado(
                    _valor_linha(row, "tipo_vaga", "tipo"),
                    TIPOS_VAGA,
                ) or "Nova vaga"
                motivo = _valor_linha(row, "motivo_substituicao", "motivo")
                sexo = _normalizar_valor_enumerado(
                    _valor_linha(row, "sexo"),
                    SEXOS,
                ) or "Indiferente"
                qtd = _valor_linha(row, "qtd_vagas", "quantidade", "qtd") or 1
                status = _normalizar_valor_enumerado(
                    _valor_linha(row, "status"),
                    STATUS_VAGA,
                ) or "Aberta"
                tipo_recrutamento = _normalizar_valor_enumerado(
                    _valor_linha(row, "tipo_recrutamento", "recrutamento"),
                    TIPOS_RECRUTAMENTO,
                ) or "Externo"
                responsavel = _valor_linha(row, "responsavel", "responsável")
                data_abertura = _valor_linha(row, "data_abertura", "abertura")
                data_conclusao = _valor_linha(row, "data_conclusao", "conclusao", "conclusão")
                observacoes = _valor_linha(row, "observacoes", "observação", "obs")

                if not departamento_nome and not nome_cargo:
                    continue

                id_dep = _get_or_create_departamento(db, departamento_nome) if departamento_nome else None
                id_cargo = None
                if nome_cargo:
                    existente_cargo = db.execute(
                        select(dho_cargos.c.id_cargo)
                        .where(dho_cargos.c.nome_cargo == nome_cargo)
                    ).scalar()
                    if existente_cargo:
                        id_cargo = existente_cargo
                    else:
                        result = db.execute(insert(dho_cargos).values(
                            id_departamento=id_dep,
                            nome_cargo=nome_cargo,
                            status="Ativo",
                            criado_em=datetime.utcnow(),
                            atualizado_em=datetime.utcnow(),
                        ))
                        id_cargo = result.inserted_primary_key[0] if result.inserted_primary_key else None

                db.execute(insert(dho_vagas).values(
                    id_departamento=id_dep,
                    id_cargo=id_cargo,
                    tipo_vaga=tipo_vaga,
                    motivo_substituicao=motivo if tipo_vaga == "Substituição" else "",
                    sexo=sexo,
                    qtd_vagas=int(float(qtd or 1)),
                    status=status,
                    tipo_recrutamento=tipo_recrutamento,
                    responsavel=responsavel,
                    data_abertura=data_abertura,
                    data_conclusao=data_conclusao,
                    observacoes=observacoes,
                    criado_em=datetime.utcnow(),
                    atualizado_em=datetime.utcnow(),
                ))

                total += 1

            else:
                return redirect_with_message("/dho/importacoes", error="Tipo de importação inválido.")

        db.commit()

        return redirect_with_message(
            "/dho/importacoes",
            success=f"Importação concluída. {total} registro(s) processado(s).",
        )

    except Exception as exc:
        db.rollback()
        return redirect_with_message(
            "/dho/importacoes",
            error=f"Erro ao importar arquivo: {exc}",
        )



@router.get("/dho/importacoes/modelo/{tipo_modelo}")
def baixar_modelo_importacao_dho(tipo_modelo: str):
    """
    Gera modelo Excel para importação DHO.
    Tipos:
    - departamentos
    - cargos
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    tipo_modelo = str(tipo_modelo or "").strip().lower()

    wb = Workbook()
    ws = wb.active

    if tipo_modelo == "departamentos":
        ws.title = "Departamentos"
        headers = ["nome_departamento", "status"]
        exemplos = [
            ["Operações", "Ativo"],
            ["DHO", "Ativo"],
            ["Comercial", "Ativo"],
        ]
        filename = "modelo_importacao_dho_departamentos.xlsx"

    elif tipo_modelo == "cargos":
        ws.title = "Cargos"
        headers = ["nome_cargo", "nome_departamento", "status"]
        exemplos = [
            ["Mecânico", "Operações", "Ativo"],
            ["Coordenador de Pista", "Operações", "Ativo"],
            ["Analista de DHO", "DHO", "Ativo"],
        ]
        filename = "modelo_importacao_dho_cargos.xlsx"

    elif tipo_modelo in ("cursos", "treinamentos", "cadastro_curso", "cadastro_treinamento"):
        ws.title = "Treinamentos"
        headers = ["nome_treinamento", "tipo_treinamento", "carga_horaria_padrao", "valor_treinamento", "status", "observacoes"]
        exemplos = [
            ["Integração de novos profissionais", "Onboarding obrigatório", 2, 0, "Ativo", ""],
            ["Ação de RH - comunicação e conduta", "Ação de RH", 1.5, 250, "Ativo", ""],
            ["Reciclagem operacional de autônomos", "Reciclagem de autônomos", 3, 500, "Ativo", ""],
        ]
        filename = "modelo_importacao_dho_treinamentos.xlsx"

    elif tipo_modelo == "vagas":
        ws.title = "Vagas"
        headers = [
            "nome_departamento",
            "nome_cargo",
            "tipo_vaga",
            "motivo_substituicao",
            "sexo",
            "qtd_vagas",
            "tipo_recrutamento",
            "status",
            "responsavel",
            "data_abertura",
            "data_conclusao",
            "observacoes",
        ]
        exemplos = [
            ["Operações", "Mecânico", "Nova vaga", "", "Indiferente", 1, "Externo", "Aberta", "DHO", "01/06/2026", "", ""],
            ["DHO", "Analista DHO", "Substituição", "Voluntário", "Indiferente", 1, "Interno", "Em andamento", "DHO", "01/06/2026", "", ""],
        ]
        filename = "modelo_importacao_dho_vagas.xlsx"

    else:
        ws.title = "Modelo"
        headers = ["erro"]
        exemplos = [["Tipo de modelo inválido. Use vagas, treinamentos ou aplicações."]]
        filename = "modelo_importacao_dho.xlsx"

    ws.append(headers)

    for row in exemplos:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = max(22, len(str(cell.value)) + 4)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    headers_response = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_response,
    )



def _normalizar_coluna_importacao(nome):
    return (
        str(nome or "")
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("ç", "c")
        .replace("ã", "a")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


def _valor_importacao(row, *nomes):
    for n in nomes:
        chave = _normalizar_coluna_importacao(n)
        if chave in row and row[chave] not in (None, ""):
            return str(row[chave]).strip()
    return ""


def _float_importacao(valor):
    valor = str(valor or "").strip().replace(",", ".")
    if not valor:
        return None
    try:
        return float(valor)
    except Exception:
        return None


def _normalizar_chave_pessoa(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "").strip())
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return " ".join(texto.lower().split())


def _empregados_importacao_por_nome(db: Session):
    rows = db.execute(
        select(
            dho_empregados,
            dho_departamentos.c.nome_departamento,
            dho_cargos.c.nome_cargo,
        )
        .select_from(
            dho_empregados
            .outerjoin(dho_departamentos, dho_departamentos.c.id_departamento == dho_empregados.c.id_departamento)
            .outerjoin(dho_cargos, dho_cargos.c.id_cargo == dho_empregados.c.id_cargo)
        )
        .where(dho_empregados.c.status == "Ativo")
        .order_by(dho_empregados.c.nome)
    ).mappings().all()

    return {
        _normalizar_chave_pessoa(row["nome"]): row
        for row in rows
        if str(row.get("nome") or "").strip()
    }


def _treinamentos_importacao_por_nome(db: Session):
    rows = db.execute(
        select(dho_treinamentos)
        .where(dho_treinamentos.c.status == "Ativo")
        .order_by(dho_treinamentos.c.nome_treinamento)
    ).mappings().all()

    return {
        _normalizar_chave_pessoa(row["nome_treinamento"]): row
        for row in rows
        if str(row.get("nome_treinamento") or "").strip()
    }


def _garantir_schema_aplicacao_treinamento_importacao():
    """
    Garante as colunas usadas na importação de aplicações de treinamento.
    """
    try:
        with engine.begin() as conn:
            dialect = conn.dialect.name

            if dialect == "sqlite":
                tabelas = [
                    r[0]
                    for r in conn.execute(
                        text("SELECT name FROM sqlite_master WHERE type='table'")
                    ).fetchall()
                ]

                if "dho_treinamento_aplicacoes" not in tabelas:
                    metadata_dho.create_all(engine)
                    return

                cols = [
                    r[1]
                    for r in conn.execute(
                        text("PRAGMA table_info(dho_treinamento_aplicacoes)")
                    ).fetchall()
                ]

                add_cols = {
                    "matricula": "TEXT",
                    "funcao": "TEXT",
                    "centro_custo": "TEXT",
                    "observacoes": "TEXT",
                    "status": "TEXT DEFAULT 'Realizado'",
                }

                for col, sql_type in add_cols.items():
                    if col not in cols:
                        conn.execute(
                            text(f"ALTER TABLE dho_treinamento_aplicacoes ADD COLUMN {col} {sql_type}")
                        )

            elif dialect == "postgresql":
                conn.execute(text("""
                    ALTER TABLE IF EXISTS dho_treinamento_aplicacoes
                    ADD COLUMN IF NOT EXISTS matricula VARCHAR(80),
                    ADD COLUMN IF NOT EXISTS funcao VARCHAR(180),
                    ADD COLUMN IF NOT EXISTS centro_custo VARCHAR(180),
                    ADD COLUMN IF NOT EXISTS observacoes TEXT,
                    ADD COLUMN IF NOT EXISTS status VARCHAR(60) DEFAULT 'Realizado'
                """))

    except Exception as exc:
        print(f"AVISO - erro ao garantir schema de aplicações importadas: {exc}")


def _get_or_create_treinamento_importacao(db: Session, nome_treinamento: str):
    nome_treinamento = str(nome_treinamento or "").strip()

    if not nome_treinamento:
        return None

    existente = db.execute(
        select(dho_treinamentos.c.id_treinamento)
        .where(dho_treinamentos.c.nome_treinamento == nome_treinamento)
    ).scalar()

    if existente:
        return existente

    result = db.execute(
        insert(dho_treinamentos).values(
            nome_treinamento=nome_treinamento,
            tipo_treinamento="Autônomo",
            carga_horaria_padrao=None,
            status="Ativo",
            observacoes="Criado automaticamente pela importação de aplicação de treinamento.",
            criado_em=datetime.utcnow(),
            atualizado_em=datetime.utcnow(),
        )
    )

    db.flush()

    try:
        return result.inserted_primary_key[0]
    except Exception:
        return db.execute(
            select(dho_treinamentos.c.id_treinamento)
            .where(dho_treinamentos.c.nome_treinamento == nome_treinamento)
        ).scalar()


@router.get("/dho/importacoes/modelo-aplicacoes-treinamento")
def baixar_modelo_aplicacoes_treinamento_dho(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Aplicacoes Treinamento"
    ws_pessoas = wb.create_sheet("Pessoas")
    ws_treinamentos = wb.create_sheet("Treinamentos")
    ws_listas = wb.create_sheet("Listas")

    headers = [
        "nome_treinamento",
        "pessoa_nome",
        "matricula",
        "funcao",
        "centro_custo",
        "data_treinamento",
        "carga_horaria",
        "status",
        "observacoes",
    ]

    ws.append(headers)

    pessoas = db.execute(
        select(
            dho_empregados,
            dho_departamentos.c.nome_departamento,
            dho_cargos.c.nome_cargo,
        )
        .select_from(
            dho_empregados
            .outerjoin(dho_departamentos, dho_departamentos.c.id_departamento == dho_empregados.c.id_departamento)
            .outerjoin(dho_cargos, dho_cargos.c.id_cargo == dho_empregados.c.id_cargo)
        )
        .where(dho_empregados.c.status == "Ativo")
        .order_by(dho_empregados.c.nome)
    ).mappings().all()

    treinamentos = db.execute(
        select(dho_treinamentos)
        .where(dho_treinamentos.c.status == "Ativo")
        .order_by(dho_treinamentos.c.nome_treinamento)
    ).mappings().all()

    ws_pessoas.append(["pessoa_nome", "matricula", "funcao", "centro_custo", "email"])
    for pessoa in pessoas:
        ws_pessoas.append([
            pessoa["nome"],
            pessoa["matricula"] or "",
            pessoa["nome_cargo"] or "",
            pessoa["nome_departamento"] or "",
            pessoa["email"] or "",
        ])

    ws_treinamentos.append(["nome_treinamento", "carga_horaria_padrao"])
    for treinamento in treinamentos:
        ws_treinamentos.append([
            treinamento["nome_treinamento"],
            treinamento["carga_horaria_padrao"] or "",
        ])

    ws_listas.append(["status"])
    for status_item in STATUS_APLICACAO:
        ws_listas.append([status_item])

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = max(22, len(str(cell.value)) + 4)

    for aba in (ws_pessoas, ws_treinamentos, ws_listas):
        for col_idx, cell in enumerate(aba[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            aba.column_dimensions[get_column_letter(col_idx)].width = max(22, len(str(cell.value)) + 4)

    max_rows = 1000
    last_pessoa = max(ws_pessoas.max_row, 2)
    last_treinamento = max(ws_treinamentos.max_row, 2)
    last_status = max(ws_listas.max_row, 2)

    dv_treinamento = DataValidation(
        type="list",
        formula1=f"=Treinamentos!$A$2:$A${last_treinamento}",
        allow_blank=False,
    )
    dv_pessoa = DataValidation(
        type="list",
        formula1=f"=Pessoas!$A$2:$A${last_pessoa}",
        allow_blank=False,
    )
    dv_status = DataValidation(
        type="list",
        formula1=f"=Listas!$A$2:$A${last_status}",
        allow_blank=False,
    )
    ws.add_data_validation(dv_treinamento)
    ws.add_data_validation(dv_pessoa)
    ws.add_data_validation(dv_status)
    dv_treinamento.add(f"A2:A{max_rows}")
    dv_pessoa.add(f"B2:B{max_rows}")
    dv_status.add(f"H2:H{max_rows}")

    for row_idx in range(2, max_rows + 1):
        ws.cell(row_idx, 3).value = f'=IFERROR(VLOOKUP($B{row_idx},Pessoas!$A:$E,2,FALSE),"")'
        ws.cell(row_idx, 4).value = f'=IFERROR(VLOOKUP($B{row_idx},Pessoas!$A:$E,3,FALSE),"")'
        ws.cell(row_idx, 5).value = f'=IFERROR(VLOOKUP($B{row_idx},Pessoas!$A:$E,4,FALSE),"")'
        ws.cell(row_idx, 7).value = f'=IFERROR(VLOOKUP($A{row_idx},Treinamentos!$A:$B,2,FALSE),"")'
        ws.cell(row_idx, 8).value = "Realizado"

    ws_pessoas.sheet_state = "hidden"
    ws_treinamentos.sheet_state = "hidden"
    ws_listas.sheet_state = "hidden"
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="modelo_importacao_dho_aplicacoes_treinamento.xlsx"'
        },
    )


@router.post("/dho/importacoes/aplicacoes-treinamento")
async def importar_aplicacoes_treinamento_dho(
    request: Request,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not _is_admin(request):
        return RedirectResponse("/?sem_acesso=dho", status_code=303)
    try:
        from openpyxl import load_workbook

        _garantir_schema_aplicacao_treinamento_importacao()

        conteudo = await arquivo.read()
        wb = load_workbook(BytesIO(conteudo), data_only=True)
        ws = wb.active

        linhas = list(ws.iter_rows(values_only=True))

        if not linhas:
            return redirect_with_message("/dho/importacoes", error="Arquivo vazio.")

        headers = [_normalizar_coluna_importacao(h) for h in linhas[0]]
        total = 0
        ignorados = 0
        empregados_por_nome = _empregados_importacao_por_nome(db)
        treinamentos_por_nome = _treinamentos_importacao_por_nome(db)
        pessoas_nao_encontradas = []
        treinamentos_nao_encontrados = []
        status_invalidos = []

        for vals in linhas[1:]:
            row = {
                headers[i]: vals[i] if i < len(vals) else None
                for i in range(len(headers))
            }

            nome_treinamento = _valor_importacao(row, "nome_treinamento", "curso", "treinamento")
            pessoa_nome = _valor_importacao(row, "pessoa_nome", "nome", "nome_completo", "pessoa")
            data_treinamento = _valor_importacao(row, "data_treinamento", "data", "data_realizacao")
            carga_horaria = _valor_importacao(row, "carga_horaria", "horas", "carga")
            status = _valor_importacao(row, "status") or "Realizado"
            observacoes = _valor_importacao(row, "observacoes", "observacao", "obs")

            if not nome_treinamento or not pessoa_nome:
                ignorados += 1
                continue

            treinamento = treinamentos_por_nome.get(_normalizar_chave_pessoa(nome_treinamento))
            if not treinamento:
                ignorados += 1
                treinamentos_nao_encontrados.append(nome_treinamento)
                continue

            empregado = empregados_por_nome.get(_normalizar_chave_pessoa(pessoa_nome))
            if not empregado:
                ignorados += 1
                pessoas_nao_encontradas.append(pessoa_nome)
                continue

            if status not in STATUS_APLICACAO:
                ignorados += 1
                status_invalidos.append(status)
                continue

            carga_horaria_importada = _float_importacao(carga_horaria)
            if carga_horaria_importada is None:
                carga_horaria_importada = treinamento["carga_horaria_padrao"]

            dados_aplicacao = {
                "id_treinamento": treinamento["id_treinamento"],
                "tipo_pessoa": "Colaborador",
                "id_autonomo": None,
                "pessoa_nome": empregado["nome"],
                "matricula": empregado["matricula"] or "",
                "funcao": empregado["nome_cargo"] or "",
                "centro_custo": empregado["nome_departamento"] or "",
                "data_treinamento": data_treinamento,
                "carga_horaria": carga_horaria_importada,
                "status": status,
                "observacoes": observacoes,
                "criado_em": datetime.utcnow(),
                "atualizado_em": datetime.utcnow(),
            }

            aplicacoes = _expandir_aplicacao_por_pessoa(dados_aplicacao)
            for aplicacao in aplicacoes:
                db.execute(insert(dho_treinamento_aplicacoes).values(**aplicacao))
                total += 1

        db.commit()

        detalhe_ignorados = ""
        if pessoas_nao_encontradas:
            nomes = ", ".join(sorted(set(pessoas_nao_encontradas))[:10])
            detalhe_ignorados = f" Pessoas não encontradas em Empregados: {nomes}."
        if treinamentos_nao_encontrados:
            nomes = ", ".join(sorted(set(treinamentos_nao_encontrados))[:10])
            detalhe_ignorados += f" Treinamentos não encontrados: {nomes}."
        if status_invalidos:
            nomes = ", ".join(sorted(set(status_invalidos))[:10])
            detalhe_ignorados += f" Status inválidos: {nomes}."

        return redirect_with_message(
            "/dho/importacoes",
            success=f"Importação concluída. {total} aplicação(ões) importada(s). {ignorados} linha(s) ignorada(s).{detalhe_ignorados}",
        )

    except Exception as exc:
        db.rollback()
        return redirect_with_message(
            "/dho/importacoes",
            error=f"Erro ao importar aplicações de treinamento: {exc}",
        )


# ── EMPREGADOS ─────────────────────────────────────────────────────────────────

@router.get("/dho/empregados")
def empregados_list(request: Request, db: Session = Depends(get_db)):
    current_user = getattr(request.state, "current_user", None)
    if not current_user:
        return RedirectResponse("/auth/login", status_code=303)
    rows = db.execute(
        select(
            dho_empregados,
            dho_departamentos.c.nome_departamento,
            dho_cargos.c.nome_cargo,
        )
        .select_from(
            dho_empregados
            .outerjoin(dho_departamentos, dho_departamentos.c.id_departamento == dho_empregados.c.id_departamento)
            .outerjoin(dho_cargos, dho_cargos.c.id_cargo == dho_empregados.c.id_cargo)
        )
        .order_by(dho_empregados.c.nome)
    ).mappings().all()

    return templates.TemplateResponse(
        "dho/empregados.html",
        {
            "request": request,
            "current_user": current_user,
            "empregados": rows,
            "total_empregados": len(rows),
            **flash_from_request(request),
        },
    )


@router.post("/dho/empregados/{id_empregado}/fornecedor")
def empregados_fornecedor(
    id_empregado: int,
    request: Request,
    numero_fornecedor: str = Form(""),
    db: Session = Depends(get_db),
):
    """Salva o número de fornecedor.

    O resto do cadastro vem da fonte oficial e é somente leitura, mas esse
    número é interno e não existe lá — por isso é editável aqui.
    """
    current_user = getattr(request.state, "current_user", None)
    if not current_user:
        return RedirectResponse("/auth/login", status_code=303)

    db.execute(
        dho_empregados.update()
        .where(dho_empregados.c.id_empregado == id_empregado)
        .values(numero_fornecedor=(numero_fornecedor or "").strip() or None,
                atualizado_em=datetime.utcnow())
    )
    db.commit()

    if request.headers.get("x-fetch") == "1":
        return JSONResponse({"ok": True})
    return redirect_with_message("/dho/empregados", success="Número de fornecedor salvo.")


@router.post("/dho/empregados/sincronizar")
def empregados_sincronizar(request: Request, db: Session = Depends(get_db)):
    current_user = getattr(request.state, "current_user", None)
    if not current_user:
        return RedirectResponse("/auth/login", status_code=303)

    try:
        resultado = sincronizar_empregados_dataworld(db)
    except Exception as exc:
        db.rollback()
        return redirect_with_message(
            "/dho/empregados",
            error=f"Não consegui sincronizar empregados pela fonte oficial: {exc}",
        )

    return redirect_with_message(
        "/dho/empregados",
        success=(
            "Empregados sincronizados pela fonte oficial. "
            f"{resultado['total_catworld']} registro(s) lido(s), "
            f"{resultado['criados']} criado(s), "
            f"{resultado['atualizados']} atualizado(s), "
            f"{resultado['inativados']} inativado(s)."
        ),
    )


@router.post("/dho/empregados/importar-fotos")
async def empregados_importar_fotos(
    request: Request,
    json_file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Carrega fotos em lote a partir de JSON com {matricula, foto_base64}."""
    current_user = getattr(request.state, "current_user", None)
    if not current_user:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/auth/login", status_code=303)
    try:
        conteudo = await json_file.read()
        fotos = json.loads(conteudo)
        if not isinstance(fotos, list):
            raise ValueError("JSON deve ser uma lista")
        atualizados = 0
        for item in fotos:
            mat = str(item.get("matricula") or "").strip()
            foto = item.get("foto_base64") or item.get("imagem_base64") or ""
            if not mat or not foto:
                continue
            r = db.execute(
                update(dho_empregados)
                .where(dho_empregados.c.matricula == mat)
                .values(foto_url=foto)
            )
            atualizados += r.rowcount or 0
        db.commit()
        return redirect_with_message(
            "/dho/empregados",
            success=f"{atualizados} foto(s) importada(s) com sucesso.",
        )
    except Exception as exc:
        db.rollback()
        return redirect_with_message(
            "/dho/empregados",
            error=f"Erro ao importar fotos: {exc}",
        )


@router.post("/dho/empregados")
async def empregados_criar(
    request: Request,
    matricula: str = Form(""),
    nome: str = Form(...),
    email: str = Form(""),
    id_departamento: str = Form(""),
    id_cargo: str = Form(""),
    data_admissao: str = Form(""),
    data_desligamento: str = Form(""),
    status: str = Form("Ativo"),
    observacoes: str = Form(""),
    db: Session = Depends(get_db),
):
    current_user = getattr(request.state, "current_user", None)
    if not current_user:
        return RedirectResponse("/auth/login", status_code=303)
    return redirect_with_message(
        "/dho/empregados",
        error="Empregados são somente leitura e vêm exclusivamente da fonte oficial.",
    )


@router.post("/dho/empregados/{id_empregado}/editar")
async def empregados_editar(
    id_empregado: int,
    request: Request,
    matricula: str = Form(""),
    nome: str = Form(...),
    email: str = Form(""),
    id_departamento: str = Form(""),
    id_cargo: str = Form(""),
    data_admissao: str = Form(""),
    data_desligamento: str = Form(""),
    status: str = Form("Ativo"),
    observacoes: str = Form(""),
    db: Session = Depends(get_db),
):
    current_user = getattr(request.state, "current_user", None)
    if not current_user:
        return RedirectResponse("/auth/login", status_code=303)
    return redirect_with_message(
        "/dho/empregados",
        error="Empregados são somente leitura e vêm exclusivamente da fonte oficial.",
    )


@router.post("/dho/empregados/{id_empregado}/excluir")
async def empregados_excluir(
    id_empregado: int,
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = getattr(request.state, "current_user", None)
    if not current_user:
        return RedirectResponse("/auth/login", status_code=303)
    return redirect_with_message(
        "/dho/empregados",
        error="Empregados são somente leitura e vêm exclusivamente da fonte oficial.",
    )
