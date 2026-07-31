"""Módulo Autônomo Sede — lançamento de custos de serviço prestado à sede."""

from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import get_db, engine
from app.models import DimAutonomo
from app.template_config import templates
from app.utils import redirect_with_message

router = APIRouter(tags=["autonomo_sede"])

MOTIVOS = [
    "Serviço de manutenção",
    "Serviço administrativo",
    "Consultoria",
    "Suporte técnico",
    "Serviço de logística",
    "Serviço de comunicação",
    "Produção de conteúdo",
    "Serviço de TI",
    "Outro",
]

SETORES = [
    "Operações",
    "Marketing",
    "Comercial",
    "RH",
    "Financeiro",
    "Comunicação",
    "Logística",
    "Tecnologia",
    "Outros",
]

# cache dos departamentos vindos do Feedz (setores)
_CACHE_SETORES: dict = {"lista": None}


def _setores_feedz() -> list[str]:
    """Lista de setores = departamentos do Feedz (rel_colab_77).

    Busca uma vez e mantém em cache; se a API do Feedz falhar, usa a lista
    fixa SETORES como fallback (sem cachear, para tentar de novo depois).
    """
    if _CACHE_SETORES["lista"]:
        return _CACHE_SETORES["lista"]
    try:
        from app.routers.indicadores import _fetch_feedz
        rows = _fetch_feedz("rel_colab_77")
        deps = sorted({
            str(r.get("departamento") or r.get("setor") or r.get("area") or "").strip().upper()
            for r in rows
        } - {"", "NÃO INFORMADO"})
        if deps:
            _CACHE_SETORES["lista"] = deps
            return deps
    except Exception as exc:
        print(f"AVISO - não consegui buscar departamentos do Feedz: {exc}")
    return SETORES

# ─── migração automática ────────────────────────────────────────────
def _garantir_tabela():
    with engine.begin() as conn:
        dialect = conn.dialect.name
        if dialect == "sqlite":
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS autonomo_sede_lancamentos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    id_autonomo VARCHAR(40),
                    nome_autonomo VARCHAR(200),
                    motivo VARCHAR(200),
                    valor FLOAT DEFAULT 0,
                    competencia VARCHAR(20),
                    observacao TEXT,
                    setor VARCHAR(60),
                    criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))
        else:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS autonomo_sede_lancamentos (
                    id SERIAL PRIMARY KEY,
                    id_autonomo VARCHAR(40),
                    nome_autonomo VARCHAR(200),
                    motivo VARCHAR(200),
                    valor FLOAT DEFAULT 0,
                    competencia VARCHAR(20),
                    observacao TEXT,
                    setor VARCHAR(60),
                    criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

_garantir_tabela()


def _migrar_tabela():
    """Adiciona colunas novas em tabelas criadas antes de alterações."""
    colunas = {"setor": "VARCHAR(60)"}
    with engine.connect() as conn:
        dialect = conn.dialect.name
        for col, tipo in colunas.items():
            if dialect == "postgresql":
                conn.execute(text(f"ALTER TABLE autonomo_sede_lancamentos ADD COLUMN IF NOT EXISTS {col} {tipo}"))
                conn.commit()
            else:
                try:
                    conn.execute(text(f"SELECT {col} FROM autonomo_sede_lancamentos LIMIT 1"))
                except Exception:
                    conn.rollback()
                    conn.execute(text(f"ALTER TABLE autonomo_sede_lancamentos ADD COLUMN {col} {tipo}"))
                    conn.commit()

_migrar_tabela()


# expressão SQL do setor com fallback
_SETOR_EXPR = "COALESCE(NULLIF(TRIM(setor), ''), 'Não informado')"
# expressão para identificar um autônomo (id, com fallback no nome)
_AUT_EXPR = "COALESCE(NULLIF(TRIM(id_autonomo), ''), nome_autonomo)"


def _parse_valor(s: str) -> float:
    if not s:
        return 0.0
    s = s.strip().replace("R$", "").replace(" ", "")
    # aceita vírgula como decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ─── listagem + indicadores ──────────────────────────────────────────
@router.get("/autonomo-sede")
def sede_index(request: Request, competencia: str = "", comp_de: str = "",
               comp_ate: str = "", setor_sel: str = "", db: Session = Depends(get_db)):
    where = "WHERE 1=1"
    params: dict = {}
    if competencia:
        where += " AND competencia = :comp"
        params["comp"] = competencia
    else:
        if comp_de:
            where += " AND competencia >= :cde"
            params["cde"] = comp_de
        if comp_ate:
            where += " AND competencia <= :cate"
            params["cate"] = comp_ate
    if setor_sel:
        where += f" AND {_SETOR_EXPR} = :setor"
        params["setor"] = setor_sel

    lancamentos = db.execute(
        text(f"""
            SELECT id, id_autonomo, nome_autonomo, motivo, valor, competencia,
                   observacao, setor, criado_em
            FROM autonomo_sede_lancamentos
            {where}
            ORDER BY competencia DESC, criado_em DESC
        """), params
    ).mappings().all()

    competencias = db.execute(
        text("SELECT DISTINCT competencia FROM autonomo_sede_lancamentos WHERE competencia IS NOT NULL AND competencia <> '' ORDER BY competencia DESC")
    ).scalars().all()

    total = sum(float(r["valor"] or 0) for r in lancamentos)

    # ── INDICADORES (respeitam os mesmos filtros) ──
    qtd_autonomos = db.execute(
        text(f"SELECT COUNT(DISTINCT {_AUT_EXPR}) FROM autonomo_sede_lancamentos {where}"), params
    ).scalar() or 0

    por_setor_rows = db.execute(
        text(f"""
            SELECT {_SETOR_EXPR} AS setor,
                   COUNT(DISTINCT {_AUT_EXPR}) AS qtd,
                   COALESCE(SUM(valor), 0) AS custo
            FROM autonomo_sede_lancamentos
            {where}
            GROUP BY {_SETOR_EXPR}
        """), params
    ).mappings().all()

    por_setor = [dict(r) for r in por_setor_rows]
    por_setor_qtd = sorted(por_setor, key=lambda d: (-d["qtd"], d["setor"]))
    por_setor_custo = sorted(por_setor, key=lambda d: (-float(d["custo"] or 0), d["setor"]))
    top5_qtd = por_setor_qtd[:5]
    top5_custo = por_setor_custo[:5]

    indicadores = {
        "qtd_autonomos": qtd_autonomos,
        "total": total,
        "registros": len(lancamentos),
        "por_setor_qtd": por_setor_qtd,
        "por_setor_custo": por_setor_custo,
        "top5_qtd": top5_qtd,
        "top5_custo": top5_custo,
    }

    return templates.TemplateResponse("autonomo_sede/index.html", {
        "request": request,
        "lancamentos": lancamentos,
        "competencias": competencias,
        "filtro_competencia": competencia,
        "comp_de": comp_de,
        "comp_ate": comp_ate,
        "setor_sel": setor_sel,
        "total": total,
        "motivos": MOTIVOS,
        "setores": _setores_feedz(),
        "indicadores": indicadores,
        **_flash(request),
    })


# ─── salvar (novo ou edição) ─────────────────────────────────────────
@router.post("/autonomo-sede/salvar")
def sede_salvar(
    request: Request,
    id: str = Form(""),
    id_autonomo: str = Form(""),
    nome_autonomo: str = Form(""),
    motivo: str = Form(""),
    valor: str = Form("0"),
    competencia: str = Form(""),
    observacao: str = Form(""),
    setor: str = Form(""),
    db: Session = Depends(get_db),
):
    dados = {
        "id_a": id_autonomo.strip(),
        "nome": nome_autonomo.strip(),
        "motivo": motivo.strip(),
        "valor": _parse_valor(valor),
        "comp": competencia.strip(),
        "obs": observacao.strip(),
        "setor": setor.strip(),
    }
    id = (id or "").strip()
    if id:
        dados["id"] = int(id)
        db.execute(text("""
            UPDATE autonomo_sede_lancamentos
            SET id_autonomo=:id_a, nome_autonomo=:nome, motivo=:motivo, valor=:valor,
                competencia=:comp, observacao=:obs, setor=:setor
            WHERE id=:id
        """), dados)
        msg = "Lançamento atualizado."
    else:
        db.execute(text("""
            INSERT INTO autonomo_sede_lancamentos (id_autonomo, nome_autonomo, motivo, valor, competencia, observacao, setor)
            VALUES (:id_a, :nome, :motivo, :valor, :comp, :obs, :setor)
        """), dados)
        msg = "Lançamento registrado."
    db.commit()
    return redirect_with_message("/autonomo-sede", success=msg)


# ─── upload CSV/Excel ────────────────────────────────────────────────
@router.post("/autonomo-sede/upload")
async def sede_upload(
    request: Request,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    nome = (arquivo.filename or "").lower()
    conteudo = await arquivo.read()
    linhas_ok = 0
    erros = 0

    try:
        if nome.endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(conteudo.decode("utf-8-sig")))
            for row in reader:
                ok = _inserir_linha(db, row)
                if ok:
                    linhas_ok += 1
                else:
                    erros += 1

        elif nome.endswith(".xlsx") or nome.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
            ws = wb.active
            # detecta a linha de cabeçalho: linha com ≥3 células preenchidas
            # e pelo menos uma delas contém "id" (evita pegar células mescladas de título)
            header_row = 1
            for r in ws.iter_rows(min_row=1, max_row=10):
                nao_vazias = [c for c in r if c.value is not None and str(c.value).strip()]
                if len(nao_vazias) >= 3 and any("id" in str(c.value).lower() for c in nao_vazias):
                    header_row = r[0].row
                    break
            headers = [str(c.value or "").strip().lower() for c in ws[header_row]]
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if all(v is None for v in row):
                    continue
                row_dict = {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)}
                ok = _inserir_linha(db, row_dict)
                if ok:
                    linhas_ok += 1
                else:
                    erros += 1
        else:
            return redirect_with_message("/autonomo-sede", error="Formato não suportado. Use .csv ou .xlsx")

        db.commit()
        msg = f"{linhas_ok} lançamento(s) importado(s)."
        if erros:
            msg += f" {erros} linha(s) ignorada(s) (sem ID ou valor)."
        return redirect_with_message("/autonomo-sede", success=msg)

    except Exception as exc:
        return redirect_with_message("/autonomo-sede", error=f"Erro ao processar arquivo: {exc}")


def _inserir_linha(db: Session, row: dict) -> bool:
    """Normaliza os nomes das colunas e insere uma linha. Retorna True se inseriu."""
    def _get(*keys):
        for k in keys:
            for rk in row:
                if rk.lower().replace(" ", "_") == k.lower().replace(" ", "_"):
                    return str(row[rk]).strip()
            for rk in row:
                if k.lower() in rk.lower():
                    return str(row[rk]).strip()
        return ""

    id_a = _get("id_autonomo", "id", "código", "codigo")
    nome = _get("nome_autonomo", "nome")
    motivo = _get("motivo")
    valor_str = _get("valor")
    comp = _get("competencia", "competência")
    obs = _get("observacao", "observação")
    setor = _get("setor")

    # normaliza competência: "2026-05-01 00:00:00" → "2026-05"
    if comp and len(comp) >= 7 and comp[4] == "-":
        comp = comp[:7]

    valor = _parse_valor(valor_str)
    if not id_a and not nome:
        return False

    db.execute(text("""
        INSERT INTO autonomo_sede_lancamentos (id_autonomo, nome_autonomo, motivo, valor, competencia, observacao, setor)
        VALUES (:id_a, :nome, :motivo, :valor, :comp, :obs, :setor)
    """), {"id_a": id_a, "nome": nome, "motivo": motivo, "valor": valor, "comp": comp, "obs": obs, "setor": setor})
    return True


# ─── modelo Excel ────────────────────────────────────────────────────
@router.get("/autonomo-sede/modelo")
def sede_modelo(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    autonomos = (
        db.query(DimAutonomo)
        .filter(DimAutonomo.status_autonomo == "Ativo")
        .order_by(DimAutonomo.nome_autonomo)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Autônomo Sede"

    ws["A1"] = "Modelo — Autônomo Sede"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Preencha: Motivo, Valor, Competência, Setor e Observação. Não altere ID e Nome."
    ws["A2"].font = Font(italic=True, color="888888", size=9)
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")

    headers = ["ID Autônomo", "Nome", "Motivo", "Valor (R$)", "Competência", "Setor", "Observação"]
    red = PatternFill("solid", fgColor="DC2626")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = red
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    alt = PatternFill("solid", fgColor="F9FAFB")
    for i, a in enumerate(autonomos):
        row = 5 + i
        fill = alt if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, v in enumerate([a.id_autonomo, a.nome_autonomo, "", "", "", "", ""], 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=9)
            if col == 1:
                cell.alignment = Alignment(horizontal="center")

    for col, w in enumerate([14, 40, 28, 16, 14, 18, 40], 1):
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo_autonomo_sede.xlsx"'},
    )


# ─── excluir ─────────────────────────────────────────────────────────
@router.post("/autonomo-sede/{id}/excluir")
def sede_excluir(id: int, request: Request, db: Session = Depends(get_db)):
    db.execute(text("DELETE FROM autonomo_sede_lancamentos WHERE id = :id"), {"id": id})
    db.commit()
    return redirect_with_message("/autonomo-sede", success="Lançamento excluído.")


def _flash(request: Request) -> dict:
    try:
        from app.utils import flash_from_request
        return flash_from_request(request)
    except Exception:
        return {}
