"""Planejamento de equipe por etapa — tudo em uma tela. v2"""

from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import get_db, engine
from app.models import DimAutonomo, DimEtapa
from app.template_config import templates
from app.utils import redirect_with_message

router = APIRouter(tags=["composicao_meta_equipe"])


def _garantir_tabela():
    with engine.begin() as conn:
        dialect = conn.dialect.name
        if dialect == "sqlite":
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS planejamento_equipe_etapa (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    id_etapa INTEGER,
                    id_autonomo INTEGER,
                    nome_autonomo VARCHAR(200),
                    data_inclusao VARCHAR(20),
                    status_disponibilidade VARCHAR(20) DEFAULT 'Não escalado',
                    criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))
        else:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS planejamento_equipe_etapa (
                    id SERIAL PRIMARY KEY,
                    id_etapa INTEGER,
                    id_autonomo INTEGER,
                    nome_autonomo VARCHAR(200),
                    data_inclusao VARCHAR(20),
                    status_disponibilidade VARCHAR(20) DEFAULT 'Não escalado',
                    criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

_garantir_tabela()


def _migrar_tabela():
    """Adiciona colunas que podem estar faltando em tabelas criadas antes de alterações."""
    colunas = {
        "data_inclusao": "VARCHAR(20)",
        "status_disponibilidade": "VARCHAR(20) DEFAULT 'Não escalado'",
    }
    with engine.connect() as conn:
        dialect = conn.dialect.name
        for col, tipo in colunas.items():
            if dialect == "postgresql":
                conn.execute(text(f"ALTER TABLE planejamento_equipe_etapa ADD COLUMN IF NOT EXISTS {col} {tipo}"))
                conn.commit()
            else:
                # SQLite: testa primeiro, pois não suporta ADD COLUMN IF NOT EXISTS
                try:
                    conn.execute(text(f"SELECT {col} FROM planejamento_equipe_etapa LIMIT 1"))
                except Exception:
                    conn.rollback()
                    conn.execute(text(f"ALTER TABLE planejamento_equipe_etapa ADD COLUMN {col} {tipo}"))
                    conn.commit()

_migrar_tabela()


# valores válidos de status de disponibilidade por etapa
STATUS_DISPONIBILIDADE = ["Escalado", "Não escalado", "Treinamento"]


def _funcao_expr() -> str:
    """Expressão SQL para a função do autônomo (cargo cadastrado, com fallback)."""
    return "COALESCE(ca.nome_cargo, da.tipo_autonomo, '—')"


# ─── página única ─────────────────────────────────────────────────────
@router.get("/composicao-meta-equipe")
def index(request: Request, etapa_sel: str = "", temp_sel: str = "",
          funcao_sel: str = "", status_sel: str = "", db: Session = Depends(get_db)):
    etapas = (
        db.query(DimEtapa)
        .order_by(DimEtapa.temporada.desc(), DimEtapa.data_inicio.asc())
        .all()
    )

    temporadas = sorted({e.temporada for e in etapas}, reverse=True)

    contagem = db.execute(
        text("SELECT id_etapa, COUNT(*) as qtd FROM planejamento_equipe_etapa GROUP BY id_etapa")
    ).mappings().all()
    etapas_com_plano = {r["id_etapa"]: r["qtd"] for r in contagem}

    func = _funcao_expr()
    join = """
        FROM planejamento_equipe_etapa p
        LEFT JOIN dim_etapas e ON e.id_etapa = p.id_etapa
        LEFT JOIN dim_autonomos da ON da.id_autonomo = p.id_autonomo
        LEFT JOIN dim_cargos_autonomos ca ON ca.id_cargo_autonomo = da.id_cargo_autonomo
    """
    status_expr = "COALESCE(NULLIF(TRIM(p.status_disponibilidade), ''), 'Não escalado')"

    # lista de funções disponíveis (para o filtro)
    funcoes = [r[0] for r in db.execute(
        text(f"SELECT DISTINCT {func} AS f {join} ORDER BY f")
    ).all() if r[0]]

    # ── PAINEL: agregados por status/função (respeita etapa/temporada/função, NÃO o status) ──
    panel_where = "WHERE 1=1"
    panel_params: dict = {}
    if etapa_sel:
        panel_where += " AND p.id_etapa = :eid"
        panel_params["eid"] = int(etapa_sel)
    if temp_sel:
        panel_where += " AND e.temporada = :temp"
        panel_params["temp"] = temp_sel
    if funcao_sel:
        panel_where += f" AND {func} = :func"
        panel_params["func"] = funcao_sel

    agg = db.execute(
        text(f"""
            SELECT {func} AS funcao, {status_expr} AS status, COUNT(*) AS qtd
            {join}
            {panel_where}
            GROUP BY {func}, {status_expr}
        """), panel_params
    ).mappings().all()

    escalados = sum(r["qtd"] for r in agg if r["status"] == "Escalado")
    reserva = sum(r["qtd"] for r in agg if r["status"] == "Não escalado")
    treinamento = sum(r["qtd"] for r in agg if r["status"] == "Treinamento")
    disponiveis = escalados + reserva  # treinamento NÃO conta como disponível

    # distribuição por função (ordenada por total desc)
    por_funcao_map: dict = {}
    for r in agg:
        f = r["funcao"] or "—"
        d = por_funcao_map.setdefault(f, {"funcao": f, "escalados": 0, "reserva": 0, "treinamento": 0})
        if r["status"] == "Escalado":
            d["escalados"] += r["qtd"]
        elif r["status"] == "Treinamento":
            d["treinamento"] += r["qtd"]
        else:
            d["reserva"] += r["qtd"]
    for d in por_funcao_map.values():
        d["disponiveis"] = d["escalados"] + d["reserva"]
    por_funcao = sorted(por_funcao_map.values(), key=lambda d: (-(d["disponiveis"] + d["treinamento"]), d["funcao"]))

    painel = {
        "disponiveis": disponiveis,
        "escalados": escalados,
        "reserva": reserva,
        "treinamento": treinamento,
        "por_funcao": por_funcao,
    }

    # ── TABELA: respeita todos os filtros (inclusive status) ──
    where = "WHERE 1=1"
    params: dict = {}
    if etapa_sel:
        where += " AND p.id_etapa = :eid"
        params["eid"] = int(etapa_sel)
    if temp_sel:
        where += " AND e.temporada = :temp"
        params["temp"] = temp_sel
    if funcao_sel:
        where += f" AND {func} = :func"
        params["func"] = funcao_sel
    if status_sel:
        where += f" AND {status_expr} = :status"
        params["status"] = status_sel

    linhas = db.execute(
        text(f"""
            SELECT p.id, p.id_etapa, e.nome_etapa, p.id_autonomo,
                   p.nome_autonomo, p.data_inclusao, p.criado_em,
                   {status_expr} AS status_disponibilidade,
                   {func} AS funcao
            {join}
            {where}
            ORDER BY e.temporada DESC, e.data_inicio, p.nome_autonomo
        """), params
    ).mappings().all()

    return templates.TemplateResponse("composicao_meta_equipe/index.html", {
        "request": request,
        "etapas": etapas,
        "temporadas": temporadas,
        "etapas_com_plano": etapas_com_plano,
        "etapa_sel": etapa_sel,
        "temp_sel": temp_sel,
        "funcao_sel": funcao_sel,
        "status_sel": status_sel,
        "funcoes": funcoes,
        "status_opcoes": STATUS_DISPONIBILIDADE,
        "painel": painel,
        "linhas": linhas,
        **_flash(request),
    })


# ─── modelo Excel ─────────────────────────────────────────────────────
@router.get("/composicao-meta-equipe/modelo")
def exportar_modelo(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    autonomos = (
        db.query(DimAutonomo)
        .filter(DimAutonomo.status_autonomo == "Ativo")
        .order_by(DimAutonomo.nome_autonomo)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planejamento"

    ws["A1"] = "Modelo — Planejamento de Equipe"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Não altere as colunas ID e Nome. Salve e importe na tela de Composição Meta Equipe."
    ws["A2"].font = Font(italic=True, color="888888", size=9)
    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")

    headers = ["ID Autônomo", "Nome", "Data Inclusão"]
    red = PatternFill("solid", fgColor="DC2626")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = red
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    alt = PatternFill("solid", fgColor="F9FAFB")
    for i, a in enumerate(autonomos):
        row = 5 + i
        fill = alt if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        data_str = a.data_inclusao.strftime("%d/%m/%Y") if a.data_inclusao else ""
        for col, v in enumerate([a.id_autonomo, a.nome_autonomo, data_str], 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=9)
            if col == 1:
                cell.alignment = Alignment(horizontal="center")

    for col, w in enumerate([14, 42, 18], 1):
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo_planejamento_equipe.xlsx"'},
    )


# ─── exportar tabela filtrada ─────────────────────────────────────────
@router.get("/composicao-meta-equipe/exportar")
def exportar_tabela(etapa_sel: str = "", temp_sel: str = "", db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    params: dict = {}
    where = "WHERE 1=1"
    etapa_nome = "Todas as etapas"
    if etapa_sel:
        where += " AND p.id_etapa = :eid"
        params["eid"] = int(etapa_sel)
        e = db.get(DimEtapa, int(etapa_sel))
        etapa_nome = e.nome_etapa if e else etapa_sel
    if temp_sel:
        where += " AND e.temporada = :temp"
        params["temp"] = temp_sel
        etapa_nome = f"Temporada {temp_sel}"

    linhas = db.execute(
        text(f"""
            SELECT p.id_etapa, e.nome_etapa, p.id_autonomo, p.nome_autonomo, p.data_inclusao, p.criado_em
            FROM planejamento_equipe_etapa p
            LEFT JOIN dim_etapas e ON e.id_etapa = p.id_etapa
            {where}
            ORDER BY e.temporada, e.data_inicio, p.nome_autonomo
        """), params
    ).mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planejamento"

    ws["A1"] = f"Planejamento de Equipe — {etapa_nome}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:E1")

    headers = ["Etapa", "ID Autônomo", "Nome", "Data Inclusão", "Importado em"]
    red = PatternFill("solid", fgColor="DC2626")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = red
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    alt = PatternFill("solid", fgColor="F9FAFB")
    for i, r in enumerate(linhas):
        row = 4 + i
        fill = alt if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        values = [r["nome_etapa"], r["id_autonomo"], r["nome_autonomo"], r["data_inclusao"], str(r["criado_em"] or "")]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=9)

    for col, w in enumerate([32, 12, 40, 16, 20], 1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = etapa_nome.replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="planejamento_{safe}.xlsx"'},
    )


# ─── upload ───────────────────────────────────────────────────────────
@router.post("/composicao-meta-equipe/upload")
async def upload_planejamento(
    request: Request,
    id_etapa: str = Form(...),
    substituir: str = Form("nao"),
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    import openpyxl

    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
        ws = wb.active
    except Exception as exc:
        return redirect_with_message(f"/composicao-meta-equipe?etapa_sel={id_etapa}", error=f"Erro ao ler arquivo: {exc}")

    # detecta linha do cabeçalho
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            v = str(cell.value or "").strip().lower()
            if v in ("id autônomo", "id autonomo", "id_autonomo"):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        return redirect_with_message(f"/composicao-meta-equipe?etapa_sel={id_etapa}", error="Cabeçalho não encontrado. Use o modelo exportado.")

    headers = [str(ws.cell(row=header_row, column=c).value or "").strip().lower()
               for c in range(1, ws.max_column + 1)]

    def col_idx(keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h:
                    return i + 1
        return None

    c_id   = col_idx(["id"])
    c_nome = col_idx(["nome"])
    c_data = col_idx(["data", "admis", "inclus"])

    if substituir == "sim":
        db.execute(text("DELETE FROM planejamento_equipe_etapa WHERE id_etapa = :eid"), {"eid": int(id_etapa)})

    ok = skip = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        id_a  = row[c_id - 1]   if c_id   else None
        nome  = str(row[c_nome - 1] or "").strip() if c_nome else ""
        data  = str(row[c_data - 1] or "").strip() if c_data else ""

        if not id_a and not nome:
            skip += 1
            continue

        db.execute(text("""
            INSERT INTO planejamento_equipe_etapa (id_etapa, id_autonomo, nome_autonomo, data_inclusao, criado_em)
            VALUES (:eid, :id_a, :nome, :data, :now)
        """), {
            "eid":  int(id_etapa),
            "id_a": int(id_a) if id_a else None,
            "nome": nome,
            "data": data,
            "now":  datetime.utcnow(),
        })
        ok += 1

    db.commit()
    msg = f"{ok} autônomo(s) importado(s) com sucesso."
    if skip:
        msg += f" {skip} linha(s) em branco ignorada(s)."
    return redirect_with_message(f"/composicao-meta-equipe?etapa_sel={id_etapa}", success=msg)


# ─── atualizar status de disponibilidade ──────────────────────────────
@router.post("/composicao-meta-equipe/linha/{linha_id}/status")
def atualizar_status(linha_id: int, status: str = Form(...),
                     etapa_sel: str = Form(""), temp_sel: str = Form(""),
                     funcao_sel: str = Form(""), status_sel: str = Form(""),
                     db: Session = Depends(get_db)):
    if status not in STATUS_DISPONIBILIDADE:
        status = "Não escalado"
    db.execute(
        text("UPDATE planejamento_equipe_etapa SET status_disponibilidade = :s WHERE id = :id"),
        {"s": status, "id": linha_id},
    )
    db.commit()
    qs = f"etapa_sel={etapa_sel}&temp_sel={temp_sel}&funcao_sel={funcao_sel}&status_sel={status_sel}"
    return redirect_with_message(f"/composicao-meta-equipe?{qs}", success="Status atualizado.")


# ─── excluir linha ────────────────────────────────────────────────────
@router.post("/composicao-meta-equipe/linha/{linha_id}/excluir")
def excluir_linha(linha_id: int, etapa_sel: str = Form(""), temp_sel: str = Form(""),
                  funcao_sel: str = Form(""), status_sel: str = Form(""), db: Session = Depends(get_db)):
    db.execute(text("DELETE FROM planejamento_equipe_etapa WHERE id = :id"), {"id": linha_id})
    db.commit()
    qs = f"etapa_sel={etapa_sel}&temp_sel={temp_sel}&funcao_sel={funcao_sel}&status_sel={status_sel}"
    return redirect_with_message(f"/composicao-meta-equipe?{qs}", success="Linha removida.")


# ─── limpar etapa ─────────────────────────────────────────────────────
@router.post("/composicao-meta-equipe/limpar")
def limpar(id_etapa: str = Form(...), db: Session = Depends(get_db)):
    db.execute(text("DELETE FROM planejamento_equipe_etapa WHERE id_etapa = :eid"), {"eid": int(id_etapa)})
    db.commit()
    return redirect_with_message(f"/composicao-meta-equipe?etapa_sel={id_etapa}", success="Planejamento desta etapa limpo.")


def _flash(request: Request) -> dict:
    try:
        from app.utils import flash_from_request
        return flash_from_request(request)
    except Exception:
        return {}
