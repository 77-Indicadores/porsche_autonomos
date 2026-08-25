import csv
import os
import json
import shutil
import tempfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import RedirectResponse
from sqlalchemy import func, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session
from google_auth_oauthlib.flow import Flow

from app.auth import tem_acesso_modulo, is_admin as _is_admin
from app.database import get_db
from app.integrations.google_sheets.client import (
    GoogleSheetsAuthRequiredError,
    GoogleSheetsPermissionError,
    SCOPES,
)
from app.integrations.google_sheets.mapper import map_row
from app.integrations.google_sheets.sync import (
    GID_ABA_RESPOSTAS,
    SHEET_NAME,
    SPREADSHEET_ID,
    WORKSHEET_NAME,
    maintenance_tickets,
    sync_maintenance_tickets,
)
from app.template_config import templates
from app.utils import flash_from_request, redirect_with_message

router = APIRouter(tags=["facilities"])

BASE_DIR = Path(__file__).resolve().parents[2]
GOOGLE_TOKEN_DB_KEY = "google_sheets_token"
GOOGLE_OAUTH_CLIENT_DB_KEY = "google_oauth_client"
GOOGLE_OAUTH_STATE_DB_KEY = "google_oauth_state"
GOOGLE_OAUTH_CODE_VERIFIER_DB_KEY = "google_oauth_code_verifier"
GOOGLE_WEB_CLIENT_JSON = os.getenv("GOOGLE_WEB_CLIENT_JSON", "").strip()

DEFAULT_CREDENTIALS_DIR = os.getenv(
    "GOOGLE_SHEETS_CREDENTIALS_DIR",
    "C:/planilha_google",
)
DEFAULT_OAUTH_CLIENT_PATH = os.getenv(
    "GOOGLE_SHEETS_OAUTH_CLIENT_PATH",
    os.path.join(DEFAULT_CREDENTIALS_DIR, "oauth_client.json"),
)
DEFAULT_TOKEN_PATH = os.getenv(
    "GOOGLE_SHEETS_TOKEN_PATH",
    os.path.join(DEFAULT_CREDENTIALS_DIR, "token_google.json"),
)
DEFAULT_AUDIT_DIR = os.getenv("FACILITIES_AUDIT_DIR", DEFAULT_CREDENTIALS_DIR)
DEFAULT_CACHE_CSV_PATH = os.getenv(
    "FACILITIES_CACHE_CSV_PATH",
    os.path.join(DEFAULT_CREDENTIALS_DIR, "dados_reformas.csv"),
)
DEFAULT_COMPLEMENTOS_PATH = os.getenv(
    "FACILITIES_COMPLEMENTOS_PATH",
    os.path.join(tempfile.gettempdir(), "facilities_complementos.json"),
)
COMPLEMENTAR_XLSX_PATH = BASE_DIR.parent / "porsche_facilities" / "Porsche facilities complementar.xlsx"


def fmt_date(value):
    if not value:
        return "-"
    return value.strftime("%d/%m/%Y %H:%M")


def fmt_file_date(path: Path):
    if not path.exists():
        return "-"
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%d/%m/%Y %H:%M")


def fmt_money(value):
    if value is None or value == "":
        return "-"
    number = Decimal(value)
    formatted = f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def fmt_date_text(value):
    if not value:
        return "-"
    value = str(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return value


def parse_date_value(value):
    if not value:
        return None
    value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return None


def parse_money_value(value):
    if value is None:
        return Decimal("0")
    value = str(value).replace("R$", "").strip()
    if not value:
        return Decimal("0")
    value = value.replace(".", "").replace(",", ".")
    try:
        return Decimal(value)
    except Exception:
        return Decimal("0")


def status_label(value):
    labels = {
        "completed": "Concluído",
        "closed": "Fechado",
        "open": "Aberto",
        "in_progress": "Em andamento",
        "pending": "Pendente",
    }
    return labels.get(value or "", value or "-")


def priority_label(value):
    labels = {
        "low": "Baixa",
        "medium": "Média",
        "high": "Alta",
        "urgent": "Urgente",
    }
    return labels.get(value or "", value or "-")


def status_complemento(complemento):
    if complemento.get("data_finalizacao"):
        return "Finalizado"
    if complemento.get("data_inicio"):
        return "Em andamento"
    return "Pendente"


def tempo_atendimento(complemento):
    inicio = parse_date_value(complemento.get("data_inicio"))
    fim = parse_date_value(complemento.get("data_finalizacao"))
    if not inicio or not fim:
        return "-"
    dias = (fim.date() - inicio.date()).days
    return max(dias, 0)


def google_sync_status(has_oauth_client: bool, has_token: bool) -> dict[str, str]:
    if has_oauth_client and has_token:
        return {
            "label": "OK",
            "help": "Conexao com o Google pronta para sincronizar.",
        }
    if has_oauth_client:
        return {
            "label": "Pendente",
            "help": "Clique em 'Atualizar espelho' para conectar sua conta Google novamente.",
        }
    return {
        "label": "Pendente",
        "help": "Configuracao OAuth do Google pendente no ambiente do sistema.",
    }


def _get_config(db: Session, chave: str) -> str | None:
    from app.models import ConfigSistema
    obj = db.query(ConfigSistema).filter(ConfigSistema.chave == chave).first()
    return obj.valor if obj else None


def _set_config(db: Session, chave: str, valor: str):
    from app.models import ConfigSistema
    obj = db.query(ConfigSistema).filter(ConfigSistema.chave == chave).first()
    if obj:
        obj.valor = valor
    else:
        db.add(ConfigSistema(chave=chave, valor=valor))
    db.flush()


def _delete_config(db: Session, chave: str):
    from app.models import ConfigSistema
    obj = db.query(ConfigSistema).filter(ConfigSistema.chave == chave).first()
    if obj:
        db.delete(obj)
    db.flush()


def load_google_oauth_client_config() -> dict | None:
    raw = GOOGLE_WEB_CLIENT_JSON or ""
    if not raw:
        return None
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return None
    return payload if isinstance(payload, dict) else None


def persist_google_oauth_client_config(db: Session, client_config: dict) -> bool:
    if not client_config:
        return False
    _set_config(
        db,
        GOOGLE_OAUTH_CLIENT_DB_KEY,
        json.dumps(client_config, ensure_ascii=False),
    )
    db.commit()
    return True


def ensure_google_oauth_client_config(db: Session) -> bool:
    persisted = _get_config(db, GOOGLE_OAUTH_CLIENT_DB_KEY)
    if persisted:
        return True
    client_config = load_google_oauth_client_config()
    if not client_config:
        return False
    return persist_google_oauth_client_config(db, client_config)


def has_google_oauth_client(db: Session) -> bool:
    return os.path.exists(DEFAULT_OAUTH_CLIENT_PATH) or bool(_get_config(db, GOOGLE_OAUTH_CLIENT_DB_KEY)) or bool(load_google_oauth_client_config())


def has_google_token(db: Session) -> bool:
    return os.path.exists(DEFAULT_TOKEN_PATH) or bool(_get_config(db, GOOGLE_TOKEN_DB_KEY))


def get_google_oauth_client_config(db: Session) -> dict | None:
    persisted = _get_config(db, GOOGLE_OAUTH_CLIENT_DB_KEY)
    if persisted:
        try:
            payload = json.loads(persisted)
            if isinstance(payload, dict):
                return payload
        except json.JSONDecodeError:
            pass
    return load_google_oauth_client_config()


def get_google_oauth_redirect_uri(db: Session) -> str | None:
    client_config = get_google_oauth_client_config(db)
    if not client_config:
        return None
    web_config = client_config.get("web") or {}
    redirect_uris = web_config.get("redirect_uris") or []
    return redirect_uris[0] if redirect_uris else None


def build_google_oauth_flow(
    client_config: dict,
    state: str | None = None,
    code_verifier: str | None = None,
):
    kwargs = {"scopes": SCOPES}
    if state:
        kwargs["state"] = state
    if code_verifier:
        kwargs["code_verifier"] = code_verifier
    else:
        kwargs["autogenerate_code_verifier"] = True
    return Flow.from_client_config(client_config, **kwargs)


def ler_complementos(db: Session) -> dict:
    from app.models import FacilitiesComplemento
    dados: dict = {}
    try:
        rows = db.query(FacilitiesComplemento).all()
        for row in rows:
            dados[row.source_row] = {
                "area_servico": row.area_servico or "",
                "unidade_local": row.unidade_local or "",
                "tipo_atendimento": row.tipo_atendimento or "",
                "data_inicio": row.data_inicio or "",
                "data_finalizacao": row.data_finalizacao or "",
                "dentro_prazo": row.dentro_prazo or "",
                "retrabalho": row.retrabalho or "",
                "custo": row.custo or "",
                "observacao": row.observacao or "",
                "rastreamento": row.rastreamento or "",
                "atualizado_em": row.atualizado_em or "",
            }
    except Exception:
        pass
    # Migração única do JSON legado quando o banco ainda está vazio
    if not dados and os.path.exists(DEFAULT_COMPLEMENTOS_PATH):
        try:
            with open(DEFAULT_COMPLEMENTOS_PATH, "r", encoding="utf-8") as arquivo:
                dados = json.load(arquivo)
        except Exception:
            pass
    return dados


def _upsert_complemento(db: Session, chave: str, dados: dict):
    from app.models import FacilitiesComplemento
    obj = db.query(FacilitiesComplemento).filter(FacilitiesComplemento.source_row == chave).first()
    if obj is None:
        obj = FacilitiesComplemento(source_row=chave)
        db.add(obj)
    obj.area_servico = dados.get("area_servico", "")
    obj.unidade_local = dados.get("unidade_local", "")
    obj.tipo_atendimento = dados.get("tipo_atendimento", "")
    obj.data_inicio = dados.get("data_inicio", "")
    obj.data_finalizacao = dados.get("data_finalizacao", "")
    obj.dentro_prazo = dados.get("dentro_prazo", "")
    obj.retrabalho = dados.get("retrabalho", "")
    obj.custo = dados.get("custo", "")
    obj.observacao = dados.get("observacao", "")
    obj.rastreamento = dados.get("rastreamento", "")
    obj.atualizado_em = dados.get("atualizado_em", "")
    db.flush()


def _areas_do_feedz() -> list[str]:
    """Departamentos da base oficial (dho_departamentos, sincronizada do Feedz).

    A lista da planilha LISTA SUSPENSA fica como reserva: em produção o
    arquivo não existe e o seletor de área vinha vazio.
    """
    try:
        from sqlalchemy import text as _text
        from app.database import engine
        with engine.connect() as conn:
            rows = conn.execute(_text(
                "SELECT DISTINCT nome_departamento FROM dho_departamentos "
                "WHERE COALESCE(status, 'Ativo') = 'Ativo' "
                "AND COALESCE(nome_departamento, '') <> ''"
            )).fetchall()
        return sorted({str(r[0]).strip() for r in rows if str(r[0]).strip()})
    except Exception as exc:
        print(f"AVISO - áreas do Feedz para facilities: {exc}")
        return []


def lista_opcoes_complementares():
    opcoes = {
        "setores": [],
        "areas": [],
        "unidades": [],
        "tipos": [],
        "prazos": ["Sim", "Não"],
        "retrabalhos": ["Sim", "Não"],
    }

    # A área/serviço vem dos departamentos sincronizados do Feedz. A planilha
    # LISTA SUSPENSA segue alimentando as demais listas — e serve de reserva
    # para a área apenas quando a base ainda não foi sincronizada.
    areas_feedz = _areas_do_feedz()
    if areas_feedz:
        opcoes["areas"] = areas_feedz

    if not COMPLEMENTAR_XLSX_PATH.exists():
        return opcoes

    try:
        from openpyxl import load_workbook

        wb = load_workbook(COMPLEMENTAR_XLSX_PATH, data_only=True, read_only=True)
        ws = wb["LISTA SUSPENSA"]
        mapa = {
            "setores": 1,
            "areas": 2,
            "unidades": 3,
            "tipos": 4,
            "prazos": 5,
            "retrabalhos": 6,
        }
        for nome, coluna in mapa.items():
            if nome == "areas" and areas_feedz:
                continue  # o Feedz manda; a planilha não sobrescreve
            valores = []
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row=row, column=coluna).value
                if value is not None and str(value).strip():
                    valores.append(str(value).strip())
            if valores:
                opcoes[nome] = sorted(dict.fromkeys(valores))
    except Exception:
        pass

    return opcoes


def _restaurar_arquivo_do_banco(db: Session, db_key: str, caminho: str):
    """Escreve no disco um arquivo cujo conteúdo está salvo no banco."""
    if os.path.exists(caminho):
        return True
    conteudo = _get_config(db, db_key)
    if not conteudo:
        return False
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    with open(caminho, "w", encoding="utf-8") as fh:
        fh.write(conteudo)
    return True


def _persistir_arquivo_no_banco(db: Session, db_key: str, caminho: str):
    if os.path.exists(caminho):
        try:
            with open(caminho, "r", encoding="utf-8") as fh:
                _set_config(db, db_key, fh.read())
        except Exception:
            pass


def tentar_atualizar_espelho(db: Session):
    # Restaura oauth_client do banco se o arquivo não existir localmente
    if not _restaurar_arquivo_do_banco(db, GOOGLE_OAUTH_CLIENT_DB_KEY, DEFAULT_OAUTH_CLIENT_PATH):
        return None, (
            "Credenciais OAuth do Google não configuradas. "
            "Configure a credencial OAuth do Google no ambiente do sistema."
        )

    # Restaura token do banco se o arquivo não existir localmente
    if not _restaurar_arquivo_do_banco(db, GOOGLE_TOKEN_DB_KEY, DEFAULT_TOKEN_PATH):
        return None, (
            "Token Google não encontrado. "
            "Clique em 'Atualizar espelho' para autenticar novamente."
        )

    try:
        result = sync_maintenance_tickets(
            db=db,
            oauth_client_path=DEFAULT_OAUTH_CLIENT_PATH,
            token_path=DEFAULT_TOKEN_PATH,
            output_dir=DEFAULT_AUDIT_DIR,
        )
        # Persiste token renovado e oauth_client no banco
        _persistir_arquivo_no_banco(db, GOOGLE_TOKEN_DB_KEY, DEFAULT_TOKEN_PATH)
        _persistir_arquivo_no_banco(db, GOOGLE_OAUTH_CLIENT_DB_KEY, DEFAULT_OAUTH_CLIENT_PATH)
        try:
            db.commit()
        except Exception:
            db.rollback()
        return result, ""
    except GoogleSheetsPermissionError as exc:
        return None, str(exc)
    except Exception as exc:
        message = str(exc)
        if "oauth2.googleapis.com" in message or "sheets.googleapis.com" in message:
            return None, (
                "Não consegui conectar no Google Sheets agora. "
                "O Windows/ambiente bloqueou a saída para os serviços do Google "
                "(oauth2.googleapis.com / sheets.googleapis.com)."
            )
        return None, f"Não consegui atualizar o espelho do Google Sheets agora: {exc}"


def _restaurar_credenciais_google(db: Session):
    ensure_google_oauth_client_config(db)
    _restaurar_arquivo_do_banco(db, GOOGLE_OAUTH_CLIENT_DB_KEY, DEFAULT_OAUTH_CLIENT_PATH)
    _restaurar_arquivo_do_banco(db, GOOGLE_TOKEN_DB_KEY, DEFAULT_TOKEN_PATH)


def _persistir_credenciais_google(db: Session):
    _persistir_arquivo_no_banco(db, GOOGLE_TOKEN_DB_KEY, DEFAULT_TOKEN_PATH)
    _persistir_arquivo_no_banco(db, GOOGLE_OAUTH_CLIENT_DB_KEY, DEFAULT_OAUTH_CLIENT_PATH)
    try:
        db.commit()
    except Exception:
        db.rollback()


def carregar_csv_espelho():
    if not os.path.exists(DEFAULT_CACHE_CSV_PATH):
        return []

    chamados = []
    with open(DEFAULT_CACHE_CSV_PATH, "r", encoding="utf-8-sig", newline="") as arquivo:
        reader = csv.DictReader(arquivo)
        for index, row in enumerate(reader, start=2):
            item = map_row(row)
            item["id"] = index
            item["source"] = "csv_cache"
            item["source_row"] = index
            item["source_spreadsheet_id"] = SPREADSHEET_ID
            item["source_gid"] = GID_ABA_RESPOSTAS
            item["raw_payload"] = json.dumps(row, ensure_ascii=False)
            chamados.append(item)

    return chamados


def preparar_chamados(rows, complementos=None):
    complementos = complementos or {}
    chamados = []
    for row in rows:
        item = dict(row)
        chave = str(item.get("source_row") or item.get("id") or "")
        complemento = complementos.get(chave, {})
        item["created_at_fmt"] = fmt_date(item.get("created_at"))
        item["completed_at_fmt"] = fmt_date(item.get("completed_at"))
        item["amount_fmt"] = fmt_money(item.get("amount"))
        item["status_label"] = status_label(item.get("status"))
        item["priority_label"] = priority_label(item.get("priority"))
        item["complemento"] = complemento
        item["complemento_status"] = status_complemento(complemento)
        item["complemento_data_inicio_fmt"] = fmt_date_text(complemento.get("data_inicio"))
        item["complemento_data_finalizacao_fmt"] = fmt_date_text(complemento.get("data_finalizacao"))
        item["complemento_custo_fmt"] = fmt_money(parse_money_value(complemento.get("custo")))
        item["complemento_tempo"] = tempo_atendimento(complemento)
        item["tem_complemento"] = bool(complemento)
        chamados.append(item)

    return chamados


def carregar_chamados_para_tela(db: Session):
    rows = []
    source_notice = ""
    source_error = ""

    if os.path.exists(DEFAULT_CACHE_CSV_PATH):
        rows = carregar_csv_espelho()
        source_notice = f"Dados carregados do espelho local: {DEFAULT_CACHE_CSV_PATH}"
    else:
        try:
            rows = db.execute(
                select(maintenance_tickets).order_by(maintenance_tickets.c.source_row.desc())
            ).mappings().all()
        except SQLAlchemyError as exc:
            source_error = f"Não consegui ler o espelho local: {exc}"

    return preparar_chamados(rows, ler_complementos(db)), source_notice, source_error


@router.get("/facilities")
def index(request: Request, db: Session = Depends(get_db)):
    if not tem_acesso_modulo(request, "facilities"):
        return RedirectResponse("/?sem_acesso=facilities", status_code=303)

    ensure_google_oauth_client_config(db)

    sync_result = None
    sync_error = ""
    cache_notice = ""

    try:
        rows = db.execute(
            select(maintenance_tickets).order_by(maintenance_tickets.c.source_row.desc())
        ).mappings().all()
    except SQLAlchemyError:
        rows = []

    if not rows and os.path.exists(DEFAULT_CACHE_CSV_PATH):
        rows = carregar_csv_espelho()
        cache_notice = (
            "Mostrando dados do espelho local. "
            "Clique em 'Atualizar espelho' para sincronizar com o Google Sheets."
        )

    complementos = ler_complementos(db)
    chamados = preparar_chamados(rows, complementos)

    total = len(chamados)
    total_abertos = sum(1 for item in chamados if item.get("status") in ("open", "pending", "in_progress"))
    total_concluidos = sum(1 for item in chamados if item.get("status") in ("completed", "closed"))
    valor_total = sum(Decimal(item.get("amount") or 0) for item in chamados)
    has_oauth_client = has_google_oauth_client(db)
    has_token = has_google_token(db)

    return templates.TemplateResponse(
        "facilities/index.html",
        {
            "request": request,
            "chamados": chamados,
            "cards": {
                "total": total,
                "abertos": total_abertos,
                "concluidos": total_concluidos,
                "valor_total": fmt_money(valor_total),
            },
            "config": {
                "spreadsheet_id": SPREADSHEET_ID,
                "sheet_name": SHEET_NAME,
                "worksheet_name": WORKSHEET_NAME,
                "gid": GID_ABA_RESPOSTAS,
                "oauth_client_path": DEFAULT_OAUTH_CLIENT_PATH,
                "token_path": DEFAULT_TOKEN_PATH,
                "audit_dir": DEFAULT_AUDIT_DIR,
                "has_oauth_client": has_oauth_client,
                "has_token": has_token,
                "google_sync": google_sync_status(has_oauth_client, has_token),
                "cache_csv_path": DEFAULT_CACHE_CSV_PATH,
                "has_cache_csv": os.path.exists(DEFAULT_CACHE_CSV_PATH),
                "complementos_path": DEFAULT_COMPLEMENTOS_PATH,
                "complementar_xlsx_path": str(COMPLEMENTAR_XLSX_PATH),
                "has_complementar_xlsx": COMPLEMENTAR_XLSX_PATH.exists(),
                "complementar_xlsx_updated_at": fmt_file_date(COMPLEMENTAR_XLSX_PATH),
            },
            "opcoes": lista_opcoes_complementares(),
            "sync_result": sync_result,
            "sync_error": sync_error,
            "cache_notice": cache_notice,
            "is_admin": _is_admin(request),
            "pode_complementar": tem_acesso_modulo(request, "facilities"),
            **flash_from_request(request),
        },
    )


@router.post("/facilities/complementar")
def salvar_complemento(
    source_row: str = Form(...),
    area_servico: str = Form(""),
    unidade_local: str = Form(""),
    tipo_atendimento: str = Form(""),
    data_inicio: str = Form(""),
    data_finalizacao: str = Form(""),
    dentro_prazo: str = Form(""),
    retrabalho: str = Form(""),
    custo: str = Form(""),
    observacao: str = Form(""),
    rastreamento: str = Form(""),
    db: Session = Depends(get_db),
):
    chave = str(source_row).strip()
    dados = {
        "area_servico": area_servico.strip(),
        "unidade_local": unidade_local.strip(),
        "tipo_atendimento": tipo_atendimento.strip(),
        "data_inicio": data_inicio.strip(),
        "data_finalizacao": data_finalizacao.strip(),
        "dentro_prazo": dentro_prazo.strip(),
        "retrabalho": retrabalho.strip(),
        "custo": custo.strip(),
        "observacao": observacao.strip(),
        "rastreamento": rastreamento.strip(),
        "atualizado_em": datetime.utcnow().isoformat(timespec="seconds"),
    }
    try:
        _upsert_complemento(db, chave, dados)
        db.commit()
        return redirect_with_message("/facilities", success="Complemento salvo.")
    except Exception as exc:
        db.rollback()
        return redirect_with_message("/facilities", error=f"Erro ao salvar complemento: {exc}")


@router.post("/facilities/upload-complementar")
async def upload_complementar(request: Request, arquivo: UploadFile = File(...), db: Session = Depends(get_db)):
    if not _is_admin(request):
        return redirect_with_message(
            "/facilities",
            error="Somente o administrador pode carregar a planilha complementar.",
        )

    nome_arquivo = arquivo.filename or ""
    if not nome_arquivo.lower().endswith(".xlsx"):
        return redirect_with_message(
            "/facilities",
            error="Envie um arquivo Excel no formato .xlsx.",
        )

    conteudo = await arquivo.read()
    if not conteudo:
        return redirect_with_message(
            "/facilities",
            error="O arquivo enviado está vazio.",
        )

    COMPLEMENTAR_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = COMPLEMENTAR_XLSX_PATH.with_name(
        f"{COMPLEMENTAR_XLSX_PATH.stem}_upload_tmp{COMPLEMENTAR_XLSX_PATH.suffix}"
    )

    try:
        temp_path.write_bytes(conteudo)

        from openpyxl import load_workbook

        wb = load_workbook(temp_path, data_only=True, read_only=True)
        tem_lista_suspensa = "LISTA SUSPENSA" in wb.sheetnames
        tem_facilities = "FACILITIES" in wb.sheetnames

        if not tem_lista_suspensa:
            wb.close()
            return redirect_with_message(
                "/facilities",
                error="A planilha precisa ter a aba LISTA SUSPENSA.",
            )

        # Importa dados da aba FACILITIES para o JSON de complementos
        # Match por: data + nome (primeiro nome) + setor (abreviação)
        linhas_importadas = 0
        if tem_facilities:
            ws_fac = wb["FACILITIES"]

            # Mapa de setor completo (CSV) → abreviação (Excel)
            SETOR_MAP = {
                "RECURSOS HUMANOS": "RHU", "ADMINISTRATIVO": "ADM", "FINANCEIRO": "ADM",
                "DIRETORIA DE OPERAÇOES": "DOP", "DIRETORIA DE OPERAÇÕES": "DOP",
                "PLANEJAMENTO E RELACIONAMENTO": "PER", "LOGÍSTICA": "LOG", "LOGISTICA": "LOG",
                "RECUPERAÇÃO E DESENVOLVIMENTO": "RED", "RECUPERACAO E DESENVOLVIMENTO": "RED",
                "ENGENHARIA DE QUALIDADE": "ENQ", "ENGENHARIA DE OFICINA": "ENG",
                "PEÇAS": "PEC", "PECAS": "PEC", "PNEUS E RODAS": "PNR",
                "CHALLENGE": "CT1", "CARREIRA": "CT2", "FUNILARIA": "FUN",
                "PRESIDÊNCIA": "PRE", "PRESIDENCIA": "PRE", "OFICINA": "CT2",
            }

            def normalizar_nome(nome: str) -> str:
                """Extrai e normaliza o primeiro nome."""
                if not nome:
                    return ""
                # Se for email, pega a parte antes do @
                if "@" in nome:
                    nome = nome.split("@")[0].replace(".", " ")
                return nome.strip().upper().split()[0] if nome.strip() else ""

            def score_match(excel_sol: str, excel_setor: str, csv_nome: str, csv_email: str, csv_setor: str) -> int:
                score = 0
                primeiro_excel = normalizar_nome(str(excel_sol) if excel_sol else "")
                primeiro_csv_nome = normalizar_nome(csv_nome)
                primeiro_csv_email = normalizar_nome(csv_email)
                if primeiro_excel and (primeiro_excel == primeiro_csv_nome or primeiro_excel == primeiro_csv_email):
                    score += 2
                setor_abrev = SETOR_MAP.get(csv_setor.strip().upper(), "")
                if setor_abrev and excel_setor and setor_abrev == str(excel_setor).strip().upper():
                    score += 1
                return score

            from collections import defaultdict
            excel_por_data = defaultdict(list)
            for row in ws_fac.iter_rows(min_row=3, values_only=True):
                if row[0] is None:
                    continue
                data_abertura = row[4]
                if not data_abertura:
                    continue
                data_str = data_abertura.strftime("%Y-%m-%d") if hasattr(data_abertura, "strftime") else str(data_abertura)[:10]
                data_inicio = row[11]
                data_fin = row[12]
                excel_por_data[data_str].append({
                    "solicitante": str(row[2]) if row[2] else "",
                    "setor": str(row[3]) if row[3] else "",
                    "usado": False,
                    "dados": {
                        "data_inicio": data_inicio.strftime("%Y-%m-%d") if hasattr(data_inicio, "strftime") else (str(data_inicio) if data_inicio else ""),
                        "data_finalizacao": data_fin.strftime("%Y-%m-%d") if hasattr(data_fin, "strftime") else (str(data_fin) if data_fin else ""),
                        "prazo": str(row[15]) if row[15] else "",
                        "retrabalho": str(row[16]) if row[16] else "",
                        "custo": str(row[17]) if row[17] else "",
                        "observacao": str(row[18]) if row[18] else "",
                    },
                })

            complementos_novos = {}

            def _processar_linha_matching(idx, data_csv, csv_nome, csv_email, csv_setor):
                candidatos = excel_por_data.get(data_csv, [])
                melhor_idx = -1
                melhor_score = -1
                for ci, cand in enumerate(candidatos):
                    if cand["usado"]:
                        continue
                    s = score_match(cand["solicitante"], cand["setor"], csv_nome, csv_email, csv_setor)
                    if s > melhor_score:
                        melhor_score = s
                        melhor_idx = ci
                if melhor_idx >= 0:
                    candidatos[melhor_idx]["usado"] = True
                    complementos_novos[str(idx)] = candidatos[melhor_idx]["dados"]
                    return True
                return False

            csv_path = DEFAULT_CACHE_CSV_PATH
            if os.path.exists(csv_path):
                with open(csv_path, "r", encoding="utf-8-sig", newline="") as arq:
                    reader = csv.DictReader(arq)
                    for idx, row_csv in enumerate(reader, start=2):
                        vals = list(row_csv.values())
                        if not vals or not vals[0]:
                            continue
                        raw_date = vals[0]
                        try:
                            data_csv = datetime.strptime(raw_date[:10], "%d/%m/%Y").strftime("%Y-%m-%d")
                        except Exception:
                            try:
                                data_csv = datetime.strptime(raw_date[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
                            except Exception:
                                continue
                        csv_nome = vals[1] if len(vals) > 1 else ""
                        csv_email = vals[2] if len(vals) > 2 else ""
                        csv_setor = vals[3] if len(vals) > 3 else ""
                        if _processar_linha_matching(idx, data_csv, csv_nome, csv_email, csv_setor):
                            linhas_importadas += 1
            else:
                try:
                    from app.integrations.google_sheets.sync import maintenance_tickets as mt
                    db_rows = db.execute(
                        select(mt).order_by(mt.c.source_row)
                    ).mappings().all()
                    for row_db in db_rows:
                        created = row_db.get("created_at")
                        if not created:
                            continue
                        data_csv = created.strftime("%Y-%m-%d") if hasattr(created, "strftime") else str(created)[:10]
                        csv_nome = str(row_db.get("requester_name") or "")
                        csv_email = str(row_db.get("requester_email") or "")
                        csv_setor = str(row_db.get("department") or "")
                        idx = row_db.get("source_row") or 0
                        if _processar_linha_matching(idx, data_csv, csv_nome, csv_email, csv_setor):
                            linhas_importadas += 1
                except Exception as exc:
                    print(f"AVISO - fallback banco para complementar falhou: {exc}")

            for chave_novo, dados_novo in complementos_novos.items():
                _upsert_complemento(db, chave_novo, dados_novo)
            db.commit()

        wb.close()

        if COMPLEMENTAR_XLSX_PATH.exists():
            backup_path = COMPLEMENTAR_XLSX_PATH.with_name(
                f"{COMPLEMENTAR_XLSX_PATH.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}{COMPLEMENTAR_XLSX_PATH.suffix}"
            )
            shutil.copy2(COMPLEMENTAR_XLSX_PATH, backup_path)

        try:
            if COMPLEMENTAR_XLSX_PATH.exists():
                COMPLEMENTAR_XLSX_PATH.unlink()
            temp_path.replace(COMPLEMENTAR_XLSX_PATH)
        except OSError:
            COMPLEMENTAR_XLSX_PATH.write_bytes(conteudo)

        msg = f"Planilha complementar carregada com sucesso. {linhas_importadas} linhas importadas."
        return redirect_with_message("/facilities", success=msg)
    except Exception as exc:
        return redirect_with_message(
            "/facilities",
            error=f"Não consegui carregar a planilha complementar: {exc}",
        )
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


# Atualizar o espelho é operação de quem trabalha no Facilities, não só do
# admin: exigir admin barrava a operadora sem nenhuma mensagem. Conectar ou
# desconectar a conta Google continua restrito a admin — é credencial.
def _pode_operar_facilities(request) -> bool:
    return _is_admin(request) or tem_acesso_modulo(request, "facilities")


_ERRO_SEM_MODULO_FAC = (
    "Atualizar o espelho exige acesso ao módulo Facilities. Peça a um "
    "administrador para liberar o módulo para o seu usuário em Usuários."
)


@router.post("/facilities/sincronizar")
def sincronizar(
    request: Request,
    oauth_client_path: str = Form(DEFAULT_OAUTH_CLIENT_PATH),
    token_path: str = Form(DEFAULT_TOKEN_PATH),
    audit_dir: str = Form(DEFAULT_AUDIT_DIR),
    db: Session = Depends(get_db),
):
    if not _pode_operar_facilities(request):
        return redirect_with_message("/facilities", error=_ERRO_SEM_MODULO_FAC)
    _restaurar_credenciais_google(db)
    if not has_google_oauth_client(db):
        return redirect_with_message(
            "/facilities",
            error="Credencial OAuth do Google nao encontrada no ambiente.",
        )
    if not has_google_token(db):
        return RedirectResponse("/facilities/oauth/iniciar", status_code=303)
    try:
        result = sync_maintenance_tickets(
            db=db,
            oauth_client_path=oauth_client_path,
            token_path=token_path,
            output_dir=audit_dir,
        )
        _persistir_credenciais_google(db)
    except GoogleSheetsAuthRequiredError:
        return RedirectResponse("/facilities/oauth/iniciar", status_code=303)
    except GoogleSheetsPermissionError as exc:
        return redirect_with_message("/facilities", error=str(exc))
    except FileNotFoundError as exc:
        return redirect_with_message("/facilities", error=str(exc))
    except Exception as exc:
        return redirect_with_message(
            "/facilities",
            error=f"Erro ao sincronizar Google Sheets: {exc}",
        )

    return redirect_with_message(
        "/facilities",
        success=(
            "Google Sheets sincronizado: "
            f"{result['total']} linhas, "
            f"{result['created']} criadas e "
            f"{result['updated']} atualizadas."
        ),
    )


@router.get("/facilities/oauth/iniciar")
def facilities_oauth_iniciar(request: Request, db: Session = Depends(get_db)):
    if not _is_admin(request):
        return redirect_with_message(
            "/facilities",
            error="Conectar ou desconectar a conta Google é restrito a administradores.")

    ensure_google_oauth_client_config(db)
    client_config = get_google_oauth_client_config(db)
    redirect_uri = get_google_oauth_redirect_uri(db)
    if not client_config or not redirect_uri:
        return redirect_with_message(
            "/facilities",
            error="Credencial OAuth do Google nao encontrada no ambiente.",
        )

    flow = build_google_oauth_flow(client_config)
    flow.redirect_uri = redirect_uri
    authorization_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="select_account consent",
    )
    _set_config(db, GOOGLE_OAUTH_STATE_DB_KEY, state)
    if getattr(flow, "code_verifier", None):
        _set_config(db, GOOGLE_OAUTH_CODE_VERIFIER_DB_KEY, flow.code_verifier)
    db.commit()
    return RedirectResponse(authorization_url, status_code=303)


@router.get("/facilities/oauth/callback")
def facilities_oauth_callback(
    request: Request,
    code: str = "",
    state: str = "",
    db: Session = Depends(get_db),
):
    if not _is_admin(request):
        return redirect_with_message(
            "/facilities",
            error="Conectar ou desconectar a conta Google é restrito a administradores.")

    client_config = get_google_oauth_client_config(db)
    redirect_uri = get_google_oauth_redirect_uri(db)
    saved_state = _get_config(db, GOOGLE_OAUTH_STATE_DB_KEY) or ""
    saved_code_verifier = _get_config(db, GOOGLE_OAUTH_CODE_VERIFIER_DB_KEY) or ""
    if not client_config or not redirect_uri:
        return redirect_with_message(
            "/facilities",
            error="Credencial OAuth do Google nao encontrada no ambiente.",
        )
    if not code:
        return redirect_with_message("/facilities", error="Codigo OAuth do Google ausente.")
    if saved_state and state and state != saved_state:
        return redirect_with_message("/facilities", error="Estado OAuth do Google invalido.")

    try:
        flow = build_google_oauth_flow(
            client_config,
            state=state or saved_state,
            code_verifier=saved_code_verifier,
        )
        flow.redirect_uri = redirect_uri
        flow.fetch_token(code=code)
        credentials = flow.credentials
        _restaurar_credenciais_google(db)
        os.makedirs(os.path.dirname(DEFAULT_TOKEN_PATH), exist_ok=True)
        with open(DEFAULT_TOKEN_PATH, "w", encoding="utf-8") as token_file:
            token_file.write(credentials.to_json())
        _persistir_credenciais_google(db)
        result = sync_maintenance_tickets(
            db=db,
            oauth_client_path=DEFAULT_OAUTH_CLIENT_PATH,
            token_path=DEFAULT_TOKEN_PATH,
            output_dir=DEFAULT_AUDIT_DIR,
        )
        _persistir_credenciais_google(db)
    except GoogleSheetsPermissionError as exc:
        return redirect_with_message("/facilities", error=str(exc))
    except Exception as exc:
        return redirect_with_message(
            "/facilities",
            error=f"Erro ao concluir login Google: {exc}",
        )

    return redirect_with_message(
        "/facilities",
        success=(
            "Google conectado e espelho sincronizado: "
            f"{result['total']} linhas, "
            f"{result['created']} criadas e "
            f"{result['updated']} atualizadas."
        ),
    )


@router.post("/facilities/oauth/desconectar")
def facilities_oauth_desconectar(request: Request, db: Session = Depends(get_db)):
    if not _is_admin(request):
        return redirect_with_message(
            "/facilities",
            error="Conectar ou desconectar a conta Google é restrito a administradores.")

    _delete_config(db, GOOGLE_TOKEN_DB_KEY)
    _delete_config(db, GOOGLE_OAUTH_STATE_DB_KEY)
    _delete_config(db, GOOGLE_OAUTH_CODE_VERIFIER_DB_KEY)
    if os.path.exists(DEFAULT_TOKEN_PATH):
        os.remove(DEFAULT_TOKEN_PATH)
    db.commit()
    return redirect_with_message(
        "/facilities",
        success="Conta Google desconectada. Clique em 'Atualizar espelho' para conectar novamente.",
    )


@router.get("/facilities/diagnostico")
def diagnostico(db: Session = Depends(get_db)):
    try:
        total = db.execute(select(func.count()).select_from(maintenance_tickets)).scalar_one()
    except SQLAlchemyError:
        total = 0
    return {
        "spreadsheet_id": SPREADSHEET_ID,
        "gid": GID_ABA_RESPOSTAS,
        "worksheet_name": WORKSHEET_NAME,
        "oauth_client_path": DEFAULT_OAUTH_CLIENT_PATH,
        "token_path": DEFAULT_TOKEN_PATH,
        "audit_dir": DEFAULT_AUDIT_DIR,
        "cache_csv_path": DEFAULT_CACHE_CSV_PATH,
        "has_oauth_client": os.path.exists(DEFAULT_OAUTH_CLIENT_PATH),
        "has_token": os.path.exists(DEFAULT_TOKEN_PATH),
        "has_cache_csv": os.path.exists(DEFAULT_CACHE_CSV_PATH),
        "tickets": total,
    }
