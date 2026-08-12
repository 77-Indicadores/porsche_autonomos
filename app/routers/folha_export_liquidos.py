"""Exportação dos líquidos da folha no layout de títulos a pagar do Protheus.

Gera um arquivo por empresa, com uma linha por colaborador: o líquido vira um
título a pagar. As colunas fixas (prefixo, tipo, natureza, projeto…) ficam
editáveis na tela, porque mudam de exercício para exercício.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import get_db
from app.template_config import templates
from app.utils import data_para_date, empresa_curta, redirect_with_message

router = APIRouter(tags=["folha_export_liquidos"])


COLUNAS = ["Prefixo", "Número", "Parcela", "Tipo", "Fornecedor", "Loja", "Natureza",
           "Emissão", "Vencimento", " Valor ", "Histórico", "finalidade", "setor",
           "subsetor", "Projeto"]

# Valores que se repetem em todas as linhas do layout. Ficam aqui como padrão e
# podem ser trocados na tela sem mexer no código.
PADROES = {
    "prefixo": "DPE",
    "tipo": "FOL",
    "loja": "1",
    "natureza": "207003",
    "finalidade": "382010001",
    "projeto": f"SEDE{date.today().year}",
    # largura do código de fornecedor no Protheus; 0 desliga o preenchimento
    "digitos": "6",
}


def _competencias(db: Session) -> list[str]:
    """Competências disponíveis, da mais recente para a mais antiga."""
    rows = db.execute(text(
        "SELECT DISTINCT competencia FROM folha_funcionarios "
        "WHERE COALESCE(competencia, '') <> ''"
    )).fetchall()
    # a competência é gravada como MM/AAAA; ordena por ano e mês, não por texto
    def chave(c: str):
        partes = str(c).split("/")
        return (partes[-1], partes[0]) if len(partes) == 2 else ("", str(c))
    return sorted({r[0] for r in rows}, key=chave, reverse=True)


def _mesma_empresa(a: str, b: str) -> bool:
    """A folha grava a razão social truncada; compara pelo começo."""
    na = re.sub(r"[^A-Z0-9]", "", (a or "").upper())
    nb = re.sub(r"[^A-Z0-9]", "", (b or "").upper())
    if not na or not nb:
        return False
    return na.startswith(nb) or nb.startswith(na)


def _mapa_setor(db: Session) -> list[dict]:
    try:
        return [dict(r) for r in db.execute(text(
            "SELECT empresa, codigo, departamento_protheus "
            "FROM folha_centros_custo WHERE COALESCE(status,'Ativo') = 'Ativo'"
        )).mappings()]
    except Exception as exc:
        print(f"AVISO - mapa de setor Protheus: {exc}")
        return []


def _setor_de(mapa: list[dict], empresa: str, centro_custo: str) -> str:
    cc = str(centro_custo or "").strip()
    if not cc:
        return ""
    for r in mapa:
        if r["codigo"] and str(r["codigo"]).strip() == cc and _mesma_empresa(empresa, r["empresa"]):
            return (r["departamento_protheus"] or "").strip()
    # mesmo código em outra empresa serve de reforço antes de devolver vazio
    for r in mapa:
        if r["codigo"] and str(r["codigo"]).strip() == cc:
            return (r["departamento_protheus"] or "").strip()
    return ""


def _mapa_matricula_feedz(db: Session) -> dict[str, str]:
    """Nome e matrícula do empregado → matrícula do Feedz.

    O fornecedor no Protheus é a matrícula do Feedz. A folha traz a própria
    matrícula, que costuma ser a mesma, mas quando as duas divergem vale a do
    Feedz — por isso a busca passa por aqui antes de usar a da folha.
    """
    mapa: dict[str, str] = {}
    try:
        rows = db.execute(text(
            "SELECT matricula, nome FROM dho_empregados "
            "WHERE COALESCE(matricula, '') <> ''"
        )).mappings().all()
    except Exception as exc:
        print(f"AVISO - mapa de matrículas do Feedz: {exc}")
        return mapa
    for r in rows:
        mat = str(r["matricula"] or "").strip()
        if not mat:
            continue
        mapa[f"mat:{_so_numero(mat)}"] = mat
        nome = str(r["nome"] or "").strip().upper()
        if nome:
            mapa[f"nome:{nome}"] = mat
    return mapa


def _so_numero(valor: str) -> str:
    """Chave de comparação de matrícula: '000044' e '44' são a mesma pessoa."""
    limpo = str(valor or "").strip()
    return limpo.lstrip("0") or limpo


def _codigo_fornecedor(matricula: str, digitos: int) -> str:
    """Matrícula no formato do cadastro de fornecedores do Protheus.

    Os códigos são de largura fixa ('000241'), mas a matrícula chega ora com
    zeros ora sem ('000044' e '106'). Só completa o que é número: código com
    letra ('S09627') fica intacto.
    """
    mat = str(matricula or "").strip()
    if not mat or digitos <= 0 or not mat.isdigit():
        return mat
    return mat.zfill(digitos)


def _linhas_da_competencia(db: Session, competencia: str,
                           digitos: int = 6) -> list[dict]:
    """Uma linha por colaborador com líquido na competência, já com empresa."""
    rows = db.execute(text("""
        SELECT a.empresa_nome AS empresa, f.matricula, f.nome, f.centro_custo,
               f.liquido, f.situacao
        FROM folha_funcionarios f
        JOIN folha_arquivos a ON a.id_arquivo = f.id_arquivo
        WHERE f.competencia = :c
        ORDER BY a.empresa_nome, f.nome
    """), {"c": competencia}).mappings().all()

    mapa_setor = _mapa_setor(db)
    mapa_feedz = _mapa_matricula_feedz(db)

    resultado = []
    for r in rows:
        liquido = float(r["liquido"] or 0)
        # líquido zero ou negativo não vira título a pagar
        if liquido <= 0:
            continue
        mat_folha = str(r["matricula"] or "").strip()
        nome = str(r["nome"] or "").strip()
        # a matrícula do Feedz manda; a da folha entra quando a pessoa ainda
        # não foi sincronizada, porque quase sempre é o mesmo número
        mat = (mapa_feedz.get(f"mat:{_so_numero(mat_folha)}")
               or mapa_feedz.get(f"nome:{nome.upper()}")
               or mat_folha)
        resultado.append({
            "empresa": r["empresa"],
            "matricula": mat_folha,
            "nome": nome,
            "centro_custo": str(r["centro_custo"] or "").strip(),
            "liquido": round(liquido, 2),
            "fornecedor": _codigo_fornecedor(mat, digitos),
            "setor": _setor_de(mapa_setor, r["empresa"], r["centro_custo"]),
        })
    return resultado


def _historico(competencia: str) -> str:
    return f"FOLHA MENSAL REF {competencia}"


def _parametros(request: Request) -> dict:
    """Constantes do layout, com o que veio da tela sobrepondo o padrão."""
    q = request.query_params
    p = {k: (q.get(k) or v).strip() for k, v in PADROES.items()}
    emissao = data_para_date(q.get("emissao")) or date.today()
    vencimento = data_para_date(q.get("vencimento")) or emissao
    p["emissao"] = emissao
    p["vencimento"] = vencimento
    p["numero"] = (q.get("numero") or emissao.strftime("%Y%m%d")).strip()
    try:
        p["digitos"] = max(0, min(12, int(q.get("digitos") or PADROES["digitos"])))
    except ValueError:
        p["digitos"] = int(PADROES["digitos"])
    return p


# ─── tela ────────────────────────────────────────────────────────────
@router.get("/folha/exportacao-liquidos")
def index(request: Request, competencia: str = "", db: Session = Depends(get_db)):
    comps = _competencias(db)
    competencia = competencia or (comps[0] if comps else "")
    p = _parametros(request)
    linhas = _linhas_da_competencia(db, competencia, p["digitos"]) if competencia else []

    por_empresa: dict[str, dict] = {}
    for l in linhas:
        g = por_empresa.setdefault(l["empresa"], {
            "empresa": l["empresa"], "curta": empresa_curta(l["empresa"]),
            "qtd": 0, "total": 0.0, "sem_fornecedor": 0, "sem_setor": 0,
        })
        g["qtd"] += 1
        g["total"] += l["liquido"]
        if not l["fornecedor"]:
            g["sem_fornecedor"] += 1
        if not l["setor"]:
            g["sem_setor"] += 1

    grupos = sorted(por_empresa.values(), key=lambda g: g["empresa"])
    pendentes = [l for l in linhas if not l["fornecedor"] or not l["setor"]]

    return templates.TemplateResponse("folha/exportacao_liquidos.html", {
        "request": request,
        "competencias": comps,
        "competencia": competencia,
        "grupos": grupos,
        "total_linhas": len(linhas),
        "total_valor": round(sum(l["liquido"] for l in linhas), 2),
        "pendentes": pendentes[:40],
        "qtd_pendentes": len(pendentes),
        "parametros": p,
        "historico": _historico(competencia),
        "colunas": COLUNAS,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


# ─── download ────────────────────────────────────────────────────────
@router.get("/folha/exportacao-liquidos/baixar")
def baixar(request: Request, competencia: str = "", empresa: str = "",
           db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not competencia:
        return redirect_with_message("/folha/exportacao-liquidos",
                                     error="Escolha a competência antes de exportar.")

    p = _parametros(request)
    linhas = [l for l in _linhas_da_competencia(db, competencia, p["digitos"])
              if not empresa or l["empresa"] == empresa]
    if not linhas:
        return redirect_with_message(
            f"/folha/exportacao-liquidos?competencia={competencia}",
            error="Nada a exportar para essa competência e empresa.")

    hist = _historico(competencia)

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    ws.append(COLUNAS)

    cab_fill = PatternFill("solid", fgColor="D9D9D9")
    cab_font = Font(bold=True)
    for c in ws[1]:
        c.fill = cab_fill
        c.font = cab_font
        c.alignment = Alignment(horizontal="center")

    for l in linhas:
        ws.append([
            p["prefixo"],
            int(p["numero"]) if p["numero"].isdigit() else p["numero"],
            None,                       # parcela vai vazia no layout
            p["tipo"],
            l["fornecedor"],
            int(p["loja"]) if p["loja"].isdigit() else p["loja"],
            int(p["natureza"]) if p["natureza"].isdigit() else p["natureza"],
            p["emissao"],
            p["vencimento"],
            l["liquido"],
            hist,
            int(p["finalidade"]) if p["finalidade"].isdigit() else p["finalidade"],
            l["setor"],
            None,                       # subsetor não é usado hoje
            p["projeto"],
        ])

    for linha in ws.iter_rows(min_row=2):
        linha[7].number_format = "dd/mm/yyyy"
        linha[8].number_format = "dd/mm/yyyy"
        linha[9].number_format = "#,##0.00"

    for i, titulo in enumerate(COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(titulo) + 6)
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    comp = re.sub(r"[^0-9]", "", competencia) or "competencia"
    alvo = re.sub(r"[^a-zA-Z0-9]+", "", empresa_curta(empresa)).lower() if empresa else "todas"
    nome = f"Layout_fopag_{comp}_{alvo}.xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )
