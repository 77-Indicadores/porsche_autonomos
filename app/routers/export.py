"""Exportação da base de dados completa em Excel. v2"""

from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import get_db, engine
from app.template_config import templates

router = APIRouter(tags=["export"])

# Tabelas disponíveis para exportação: (chave, label, query SQL)
TABELAS = [
    ("autonomos",        "Autônomos",             "SELECT * FROM dim_autonomos ORDER BY nome_autonomo"),
    ("etapas",           "Etapas",                "SELECT * FROM dim_etapas ORDER BY temporada DESC, data_inicio"),
    ("categorias",       "Categorias",            "SELECT * FROM dim_provas ORDER BY nome_prova"),
    ("carros",           "Carros",                "SELECT * FROM dim_carros ORDER BY numero_carro"),
    ("pilotos",          "Pilotos",               "SELECT * FROM dim_pilotos ORDER BY nome_piloto"),
    ("cargos",           "Cargos Autônomo",       "SELECT * FROM dim_cargos_autonomos ORDER BY nome_cargo"),
    ("alocacoes",        "Alocações",             "SELECT * FROM fato_piloto_autonomo_prova ORDER BY id_etapa, id_prova"),
    ("equipe_geral",     "Equipe Geral",          "SELECT * FROM equipe_geral ORDER BY id_etapa"),
    ("planejamento",     "Planejamento Equipe",   "SELECT * FROM planejamento_equipe_etapa ORDER BY id_etapa"),
    ("autonomo_sede",    "Autônomo Sede",         "SELECT * FROM autonomo_sede_lancamentos ORDER BY criado_em DESC"),
    ("pesquisas",        "Pesquisas",             "SELECT * FROM pesquisas ORDER BY criado_em DESC"),
    ("composicao_padrao","Composição Padrão",     "SELECT * FROM composicao_padrao ORDER BY id_etapa"),
]


@router.get("/export")
def export_index(request: Request):
    return templates.TemplateResponse("export/index.html", {
        "request": request,
        "tabelas": [(k, label) for k, label, _ in TABELAS],
    })


@router.get("/export/download")
def export_download(tabelas_sel: str = "", db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    selecionadas = set(tabelas_sel.split(",")) if tabelas_sel else {k for k, _, _ in TABELAS}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba padrão

    red = PatternFill("solid", fgColor="DC2626")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt = PatternFill("solid", fgColor="F9FAFB")

    abas_criadas = 0
    for chave, label, sql in TABELAS:
        if chave not in selecionadas:
            continue
        try:
            rows = db.execute(text(sql)).mappings().all()
        except Exception:
            continue

        if not rows:
            ws = wb.create_sheet(title=label[:31])
            ws["A1"] = "Sem dados"
            abas_criadas += 1
            continue

        ws = wb.create_sheet(title=label[:31])
        cols = list(rows[0].keys())

        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = red
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for ri, row in enumerate(rows):
            fill = alt if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=ri + 2, column=ci, value=row[col])
                cell.fill = fill
                cell.border = border
                cell.font = Font(size=9)

        ws.freeze_panes = "A2"
        abas_criadas += 1

    if abas_criadas == 0:
        ws = wb.create_sheet("Vazio")
        ws["A1"] = "Nenhuma tabela selecionada ou sem dados."

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="export_porsche_{ts}.xlsx"'},
    )
