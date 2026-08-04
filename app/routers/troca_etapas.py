import io
from datetime import datetime

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import (
    Column,
    DateTime,
    Integer,
    MetaData,
    Table,
    Text,
    UniqueConstraint,
    delete,
    insert,
    select,
    text,
    update,
)
from sqlalchemy.orm import Session, joinedload

from app.database import engine, get_db
from app.models import DimAutonomo, DimEtapa, DimMotivoTroca, FatoPilotoAutonomoProva
from app.template_config import templates
from app.utils import redirect_with_message

router = APIRouter(tags=["troca_etapas"])

_metadata = MetaData()

troca_entre_etapas = Table(
    "troca_entre_etapas",
    _metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("id_etapa_anterior", Integer, nullable=False),
    Column("id_etapa_proxima", Integer, nullable=False),
    Column("id_fato", Integer, nullable=False),
    Column("id_motivo_troca", Integer, nullable=True),
    Column("id_autonomo_substituto", Integer, nullable=True),
    Column("justificativa", Text, nullable=True),
    Column("registrado_em", DateTime, default=datetime.utcnow),
    UniqueConstraint("id_etapa_anterior", "id_etapa_proxima", "id_fato", name="uq_troca_etapas_fato"),
)

_metadata.create_all(engine)


def _garantir_schema():
    try:
        with engine.begin() as conn:
            if conn.dialect.name == "postgresql":
                conn.execute(text("""
                    ALTER TABLE IF EXISTS troca_entre_etapas
                    ADD COLUMN IF NOT EXISTS id_autonomo_substituto INTEGER
                """))
    except Exception as exc:
        print(f"AVISO - troca_entre_etapas schema: {exc}")


_garantir_schema()


@router.get("/troca-etapas")
def troca_etapas_index(request: Request, db: Session = Depends(get_db)):
    etapas = db.query(DimEtapa).order_by(
        DimEtapa.temporada.desc(),
        DimEtapa.data_inicio,
        DimEtapa.nome_etapa,
    ).all()

    motivos = (
        db.query(DimMotivoTroca)
        .filter(DimMotivoTroca.status == "Ativo")
        .order_by(DimMotivoTroca.motivo_troca)
        .all()
    )
    motivos_map = {m.id_motivo_troca: m.motivo_troca for m in motivos}

    autonomos_map = {
        a.id_autonomo: a.nome_autonomo
        for a in db.query(DimAutonomo).all()
    }

    # Todos os fatos carregados de uma vez
    todos_fatos = (
        db.query(FatoPilotoAutonomoProva)
        .options(
            joinedload(FatoPilotoAutonomoProva.piloto),
            joinedload(FatoPilotoAutonomoProva.autonomo),
            joinedload(FatoPilotoAutonomoProva.prova),
            joinedload(FatoPilotoAutonomoProva.carro),
        )
        .all()
    )
    fatos_por_etapa: dict = {}
    for f in todos_fatos:
        fatos_por_etapa.setdefault(f.id_etapa, []).append(f)

    # Todos os registros de troca já salvos
    todas_trocas = db.execute(select(troca_entre_etapas)).mappings().all()
    trocas_db: dict = {}
    for t in todas_trocas:
        key = (t["id_etapa_anterior"], t["id_etapa_proxima"], t["id_fato"])
        trocas_db[key] = dict(t)

    # ── Agrupamento por categoria ───────────────────────────────────────
    # Cada categoria tem o seu proprio calendario: a Sprint Challenge pode
    # pular etapas que a Carrera Cup corre. Parear pelo calendario geral
    # perdia transicoes reais (ex.: etapa 2 -> 5 da Sprint), entao o
    # pareamento e feito dentro de cada categoria.
    def _nome_prova(f):
        return (f.prova.nome_prova if f.prova else None) or ""

    etapas_por_id = {e.id_etapa: e for e in etapas}
    ordem_etapa = {e.id_etapa: i for i, e in enumerate(etapas)}

    # categoria -> id_etapa -> fatos
    por_categoria: dict = {}
    for f in todos_fatos:
        cat = _nome_prova(f) or "Sem categoria"
        por_categoria.setdefault(cat, {}).setdefault(f.id_etapa, []).append(f)

    categorias = []
    for cat in sorted(por_categoria):
        etapas_cat = sorted(
            (etapas_por_id[eid] for eid in por_categoria[cat] if eid in etapas_por_id),
            key=lambda e: ordem_etapa[e.id_etapa],
        )

        transicoes = []
        for i in range(len(etapas_cat) - 1):
            ant, prox = etapas_cat[i], etapas_cat[i + 1]
            if ant.temporada != prox.temporada:
                continue

            fatos_ant = por_categoria[cat].get(ant.id_etapa, [])
            fatos_prox = por_categoria[cat].get(prox.id_etapa, [])

            chaves_prox = {(f.id_autonomo, f.id_piloto, f.id_carro) for f in fatos_prox}
            pilotos_na_prox = {f.id_piloto for f in fatos_prox}

            # quem ja estava na etapa anterior, por piloto/carro/funcao
            autonomos_na_ant: dict = {}
            for f in fatos_ant:
                chave = (f.id_piloto, f.id_carro, f.funcao_autonomo or "")
                autonomos_na_ant.setdefault(chave, set()).add(f.id_autonomo)

            # substituto 1-para-1: autonomo novo na mesma funcao
            substitutos_map: dict = {}
            for f in fatos_prox:
                chave = (f.id_piloto, f.id_carro, f.funcao_autonomo or "")
                if f.id_autonomo not in autonomos_na_ant.get(chave, set()):
                    nome_aut = getattr(f.autonomo, "nome_autonomo", None) or "—"
                    substitutos_map.setdefault(chave, []).append(nome_aut)

            # quem continua na categoria na etapa seguinte, mesmo em outro carro
            autonomos_na_prox = {f.id_autonomo for f in fatos_prox}

            ausentes = []
            ids_piloto_nao_correu = set()
            ids_remanejado = set()
            destinos: dict = {}
            substitutos: dict = {}
            for f in fatos_ant:
                if (f.id_autonomo, f.id_piloto, f.id_carro) in chaves_prox:
                    continue
                ausentes.append(f)
                if f.id_piloto not in pilotos_na_prox:
                    ids_piloto_nao_correu.add(f.id_fato)
                    continue
                if f.id_autonomo in autonomos_na_prox:
                    # nao saiu da equipe: apenas mudou de carro/piloto
                    ids_remanejado.add(f.id_fato)
                    destinos[f.id_fato] = [
                        (getattr(g.carro, "chassi", None) or str(g.id_carro or "—"),
                         getattr(g.piloto, "nome_piloto", None) or "—")
                        for g in fatos_prox if g.id_autonomo == f.id_autonomo
                    ]
                    continue
                subs = substitutos_map.get((f.id_piloto, f.id_carro, f.funcao_autonomo or ""), [])
                if not subs:
                    # a funcao pode mudar entre etapas; tenta so por piloto e carro
                    for (pid, cid, _fn), nomes in substitutos_map.items():
                        if pid == f.id_piloto and cid == f.id_carro:
                            subs = subs + nomes
                if subs:
                    substitutos[f.id_fato] = subs

            if not ausentes:
                continue

            trocas_salvas: dict = {}
            for f in ausentes:
                t = trocas_db.get((ant.id_etapa, prox.id_etapa, f.id_fato))
                if t:
                    trocas_salvas[f.id_fato] = {
                        "id_motivo_troca": t["id_motivo_troca"],
                        "justificativa": t["justificativa"] or "",
                        "motivo_label": motivos_map.get(t["id_motivo_troca"], ""),
                    }

            trocas_reais = [f for f in ausentes
                            if f.id_fato not in ids_piloto_nao_correu
                            and f.id_fato not in ids_remanejado]
            remanejados = [f for f in ausentes if f.id_fato in ids_remanejado]
            nao_correram = [f for f in ausentes if f.id_fato in ids_piloto_nao_correu]

            transicoes.append({
                "etapa_ant": ant,
                "etapa_prox": prox,
                "ausentes": ausentes,
                "trocas_salvas": trocas_salvas,
                "ids_piloto_nao_correu": ids_piloto_nao_correu,
                "substitutos": substitutos,
                "trocas_reais": trocas_reais,
                "remanejados": remanejados,
                "destinos": destinos,
                "nao_correram": nao_correram,
            })

        if transicoes:
            categorias.append({
                "nome": cat,
                "transicoes": transicoes,
                "qt_trocas": sum(len(t["trocas_reais"]) for t in transicoes),
                "qt_remanejados": sum(len(t["remanejados"]) for t in transicoes),
                "qt_justif": sum(
                    len([f for f in t["trocas_reais"] if f.id_fato in t["trocas_salvas"]])
                    for t in transicoes
                ),
                "etapas_seq": " → ".join(
                    [transicoes[0]["etapa_ant"].nome_etapa]
                    + [t["etapa_prox"].nome_etapa for t in transicoes]
                ),
            })

    # índice global das transições, usado pelo filtro de etapa no template
    idx = 0
    for c in categorias:
        for t in c["transicoes"]:
            idx += 1
            t["idx"] = idx
            t["label"] = f"{t['etapa_ant'].nome_etapa} → {t['etapa_prox'].nome_etapa}"

    return templates.TemplateResponse(
        "troca_etapas/index.html",
        {
            "request": request,
            "motivos": motivos,
            "categorias": categorias,
        },
    )


@router.get("/troca-etapas/exportar")
def exportar_excel(db: Session = Depends(get_db)):
    etapas = db.query(DimEtapa).order_by(
        DimEtapa.temporada.desc(),
        DimEtapa.data_inicio,
        DimEtapa.nome_etapa,
    ).all()

    motivos_map = {
        m.id_motivo_troca: m.motivo_troca
        for m in db.query(DimMotivoTroca).all()
    }

    todos_fatos = (
        db.query(FatoPilotoAutonomoProva)
        .options(
            joinedload(FatoPilotoAutonomoProva.piloto),
            joinedload(FatoPilotoAutonomoProva.autonomo),
            joinedload(FatoPilotoAutonomoProva.prova),
            joinedload(FatoPilotoAutonomoProva.carro),
        )
        .all()
    )
    fatos_por_etapa: dict = {}
    for f in todos_fatos:
        fatos_por_etapa.setdefault(f.id_etapa, []).append(f)

    todas_trocas = db.execute(select(troca_entre_etapas)).mappings().all()
    trocas_db: dict = {}
    for t in todas_trocas:
        key = (t["id_etapa_anterior"], t["id_etapa_proxima"], t["id_fato"])
        trocas_db[key] = dict(t)

    wb = Workbook()
    ws = wb.active
    ws.title = "Troca de Etapas"

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="F1F5F9", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Temporada", "Etapa Anterior", "Etapa Seguinte",
        "Tipo", "Piloto", "Carro / Chassi", "Categoria",
        "Autônomo", "Função", "Substituto (etapa seguinte)",
        "Motivo de Troca", "Justificativa", "Status",
    ]
    col_widths = [12, 28, 28, 22, 28, 16, 18, 32, 20, 32, 28, 32, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 30

    def _nome_prova(f):
        return (f.prova.nome_prova if f.prova else None) or ""

    # Mesmo pareamento da tela: dentro de cada categoria, etapas consecutivas
    # daquela categoria — e não do calendário geral.
    etapas_por_id = {e.id_etapa: e for e in etapas}
    ordem_etapa = {e.id_etapa: i for i, e in enumerate(etapas)}

    por_categoria: dict = {}
    for f in todos_fatos:
        cat = _nome_prova(f) or "Sem categoria"
        por_categoria.setdefault(cat, {}).setdefault(f.id_etapa, []).append(f)

    pares = []
    for cat in sorted(por_categoria):
        etapas_cat = sorted(
            (etapas_por_id[eid] for eid in por_categoria[cat] if eid in etapas_por_id),
            key=lambda e: ordem_etapa[e.id_etapa],
        )
        for i in range(len(etapas_cat) - 1):
            a, p = etapas_cat[i], etapas_cat[i + 1]
            if a.temporada == p.temporada:
                pares.append((cat, a, p))

    row_num = 2
    for nome, ant, prox in pares:
        fatos_ant = por_categoria[nome].get(ant.id_etapa, [])
        fatos_prox_list = por_categoria[nome].get(prox.id_etapa, [])

        chaves_prox = {(f.id_autonomo, f.id_piloto, f.id_carro) for f in fatos_prox_list}
        pilotos_na_prox = {f.id_piloto for f in fatos_prox_list}
        autonomos_na_prox = {f.id_autonomo for f in fatos_prox_list}

        autonomos_na_ant: dict = {}
        for f in fatos_ant:
            chave = (f.id_piloto, f.id_carro, f.funcao_autonomo or "")
            autonomos_na_ant.setdefault(chave, set()).add(f.id_autonomo)

        substitutos_map: dict = {}
        for f in fatos_prox_list:
            chave = (f.id_piloto, f.id_carro, f.funcao_autonomo or "")
            if f.id_autonomo not in autonomos_na_ant.get(chave, set()):
                nome_aut = getattr(f.autonomo, "nome_autonomo", None) or "—"
                substitutos_map.setdefault(chave, []).append(nome_aut)

        for f in fatos_ant:
            if (f.id_autonomo, f.id_piloto, f.id_carro) in chaves_prox:
                continue

            nao_correu = f.id_piloto not in pilotos_na_prox
            remanejado = not nao_correu and f.id_autonomo in autonomos_na_prox
            if nao_correu:
                tipo = "Piloto não correu"
            elif remanejado:
                tipo = "Remanejado para outro carro"
            else:
                tipo = "Troca de autônomo"

            subs_chave = (f.id_piloto, f.id_carro, f.funcao_autonomo or "")
            substitutos = ", ".join(substitutos_map.get(subs_chave, [])) or "—"

            key = (ant.id_etapa, prox.id_etapa, f.id_fato)
            troca = trocas_db.get(key)
            motivo_label = motivos_map.get(troca["id_motivo_troca"], "") if troca and troca.get("id_motivo_troca") else ""
            justificativa = (troca.get("justificativa") or "") if troca else ""
            status = "Justificado" if troca else "Não justificado"

            ws.append([
                ant.temporada,
                ant.nome_etapa,
                prox.nome_etapa,
                tipo,
                getattr(f.piloto, "nome_piloto", "") or "",
                getattr(f.carro, "chassi", "") or str(f.id_carro or ""),
                nome,
                getattr(f.autonomo, "nome_autonomo", "") or "",
                f.funcao_autonomo or "",
                substitutos,
                motivo_label,
                justificativa,
                status,
            ])

            if nao_correu:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = PatternFill("solid", fgColor="1E3A5F")

            row_num += 1

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"troca_etapas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/troca-etapas/registrar")
def registrar_motivo(
    request: Request,
    id_etapa_anterior: int = Form(...),
    id_etapa_proxima: int = Form(...),
    id_fato: int = Form(...),
    id_motivo_troca: str = Form(""),
    justificativa: str = Form(""),
    db: Session = Depends(get_db),
):
    motivo_int = int(id_motivo_troca) if id_motivo_troca.strip() else None
    justificativa = justificativa.strip()

    existing = db.execute(
        select(troca_entre_etapas).where(
            troca_entre_etapas.c.id_etapa_anterior == id_etapa_anterior,
            troca_entre_etapas.c.id_etapa_proxima == id_etapa_proxima,
            troca_entre_etapas.c.id_fato == id_fato,
        )
    ).first()

    if existing:
        db.execute(
            update(troca_entre_etapas)
            .where(
                troca_entre_etapas.c.id_etapa_anterior == id_etapa_anterior,
                troca_entre_etapas.c.id_etapa_proxima == id_etapa_proxima,
                troca_entre_etapas.c.id_fato == id_fato,
            )
            .values(
                id_motivo_troca=motivo_int,
                justificativa=justificativa,
                registrado_em=datetime.utcnow(),
            )
        )
    else:
        db.execute(
            insert(troca_entre_etapas).values(
                id_etapa_anterior=id_etapa_anterior,
                id_etapa_proxima=id_etapa_proxima,
                id_fato=id_fato,
                id_motivo_troca=motivo_int,
                justificativa=justificativa,
                registrado_em=datetime.utcnow(),
            )
        )

    db.commit()
    if request.headers.get("x-fetch") == "1":
        return JSONResponse({"ok": True})
    return redirect_with_message("/troca-etapas", success="Registrado com sucesso.")


@router.post("/troca-etapas/excluir")
def excluir_motivo(
    id_etapa_anterior: int = Form(...),
    id_etapa_proxima: int = Form(...),
    id_fato: int = Form(...),
    db: Session = Depends(get_db),
):
    db.execute(
        delete(troca_entre_etapas).where(
            troca_entre_etapas.c.id_etapa_anterior == id_etapa_anterior,
            troca_entre_etapas.c.id_etapa_proxima == id_etapa_proxima,
            troca_entre_etapas.c.id_fato == id_fato,
        )
    )
    db.commit()
    return redirect_with_message("/troca-etapas", success="Registro removido.")
