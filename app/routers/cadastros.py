from pathlib import Path
import base64
import mimetypes
from fastapi import APIRouter, Depends, Form, Request, UploadFile, File
from sqlalchemy import or_, text
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import DimCargoAutonomo, DimAutonomo, DimEtapa, DimMotivoTroca, DimPiloto, DimProva, DimTipoProva, FatoPilotoAutonomoProva
from app.template_config import templates
from app.utils import flash_from_request, parse_date, redirect_with_message

router = APIRouter(tags=["cadastros"])



# Mantido apenas por compatibilidade com fotos antigas que já estavam salvas em /static/uploads.
UPLOAD_BASE = Path(__file__).resolve().parents[1] / "static" / "uploads"


def salvar_upload_foto(arquivo: UploadFile, pasta: str):
    """
    Salva a imagem diretamente no banco em formato Base64 Data URI.

    Antes:
        - salvava arquivo físico em app/static/uploads
        - gravava no banco apenas /static/uploads/...

    Agora:
        - não grava arquivo local
        - grava no banco data:image/...;base64,...
    """
    if not arquivo or not arquivo.filename:
        return None

    extensao = Path(arquivo.filename).suffix.lower()

    mime_por_extensao = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
    }

    mime_type = mime_por_extensao.get(extensao)

    if not mime_type:
        return None

    conteudo = arquivo.file.read()

    if not conteudo:
        return None

    # Limite preventivo de 3 MB para não estourar banco/renderização.
    # Se precisar, podemos aumentar depois.
    limite_mb = 3
    if len(conteudo) > limite_mb * 1024 * 1024:
        return None

    base64_str = base64.b64encode(conteudo).decode("utf-8")

    return f"{mime_type};base64,{base64_str}".join(["data:", ""])



def lists(db: Session):
    return {
        "pilotos": db.query(DimPiloto).order_by(DimPiloto.nome_piloto).all(),
        "autonomos": db.query(DimAutonomo).order_by(DimAutonomo.nome_autonomo).all(),
        "etapas": db.query(DimEtapa).order_by(DimEtapa.temporada.desc(), DimEtapa.nome_etapa).all(),
        "tipos": db.query(DimTipoProva).order_by(DimTipoProva.nome_tipo_prova).all(),
        "motivos": db.query(DimMotivoTroca).order_by(DimMotivoTroca.motivo_troca).all(),
        "cargos_autonomos": db.query(DimCargoAutonomo).filter(DimCargoAutonomo.status == "Ativo").order_by(DimCargoAutonomo.nome_cargo).all(),
    }


def parse_date_safe(value: str, msg: str):
    if not value:
        return None
    try:
        return parse_date(value)
    except ValueError:
        raise ValueError(msg)


def aplicar_id_manual(db: Session, obj, model, pk_name: str, requested_id: str, current_id: str = ""):
    requested_id = str(requested_id or "").strip()
    current_id = str(current_id or "").strip()

    if not requested_id:
        return

    try:
        requested = int(requested_id)
    except ValueError:
        raise ValueError("ID manual invalido.")

    if requested <= 0:
        raise ValueError("ID manual deve ser maior que zero.")

    current = int(current_id) if current_id else getattr(obj, pk_name, None)
    if current and requested == int(current):
        setattr(obj, pk_name, requested)
        return

    if db.get(model, requested):
        raise ValueError(f"ID {requested} ja esta em uso.")

    setattr(obj, pk_name, requested)


@router.get("/pilotos")
def pilotos(request: Request, q: str = "", db: Session = Depends(get_db)):
    query = db.query(DimPiloto)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(DimPiloto.nome_piloto.ilike(like), DimPiloto.status_piloto.ilike(like)))
    return templates.TemplateResponse("cadastros/pilotos.html", {"request": request, "items": query.order_by(DimPiloto.nome_piloto).all(), "q": q, **flash_from_request(request)})


@router.post("/pilotos")
def salvar_piloto(
    request: Request,
    id_piloto: str = Form(""),
    current_id_piloto: str = "",
    nome_piloto: str = Form(...),
    cpf: str = Form(""),
    telefone: str = Form(""),
    email: str = Form(""),
    data_inclusao: str = Form(""),
    status_piloto: str = Form("Ativo"),
    observacoes: str = Form(""),
    foto_url: str = Form(""),
    foto_upload: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    current_id_piloto = (current_id_piloto or request.query_params.get("current_id_piloto", "")).strip()
    piloto = db.get(DimPiloto, int(current_id_piloto or id_piloto)) if (current_id_piloto or id_piloto) else DimPiloto()
    try:
        aplicar_id_manual(db, piloto, DimPiloto, "id_piloto", id_piloto, current_id_piloto)
    except ValueError as e:
        return redirect_with_message("/pilotos", error=str(e))
    piloto.nome_piloto = nome_piloto
    piloto.cpf = cpf
    piloto.telefone = telefone
    piloto.email = email
    try:
        data = parse_date_safe(data_inclusao, "Data de inclusao invalida.")
    except ValueError as e:
        return redirect_with_message("/pilotos", error=str(e))
    piloto.data_inclusao = data or piloto.data_inclusao
    piloto.status_piloto = status_piloto
    piloto.observacoes = observacoes
    foto_salva = salvar_upload_foto(foto_upload, "pilotos")

    if foto_salva:
        piloto.foto_url = foto_salva
    elif foto_url:
        piloto.foto_url = foto_url
    db.add(piloto)
    db.commit()
    return redirect_with_message("/pilotos", success="Piloto salvo com sucesso.")


@router.post("/pilotos/{id_piloto}/desligar")
def desligar_piloto(id_piloto: int, data_desligamento: str = Form(...), motivo_desligamento: str = Form(...), db: Session = Depends(get_db)):
    piloto = db.get(DimPiloto, id_piloto)
    if not piloto:
        return redirect_with_message("/pilotos", error="Piloto nao encontrado.")
    piloto.status_piloto = "Desligado"
    try:
        piloto.data_desligamento = parse_date(data_desligamento)
    except ValueError:
        return redirect_with_message("/pilotos", error="Data de desligamento invalida.")
    piloto.motivo_desligamento = motivo_desligamento
    db.commit()
    return redirect_with_message("/pilotos", success="Piloto desligado sem exclusao fisica.")


@router.get("/autonomos")
def autonomos(request: Request, q: str = "", db: Session = Depends(get_db)):
    query = db.query(DimAutonomo)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(DimAutonomo.nome_autonomo.ilike(like), DimAutonomo.tipo_autonomo.ilike(like), DimAutonomo.status_autonomo.ilike(like)))
    return templates.TemplateResponse("cadastros/autonomos.html", {"request": request, "items": query.order_by(DimAutonomo.nome_autonomo).all(), "q": q, **lists(db), **flash_from_request(request)})


@router.post("/autonomos")
def salvar_autonomo(
    request: Request,
    id_autonomo: str = Form(""),
    current_id_autonomo: str = Form(""),
    nome_autonomo: str = Form(...),
    cpf: str = Form(""),
    telefone: str = Form(""),
    email: str = Form(""),
    tipo_autonomo: str = Form(""),
    id_cargo_autonomo: str = Form(""),
    data_inclusao: str = Form(""),
    status_autonomo: str = Form("Ativo"),
    observacoes: str = Form(""),
    foto_url: str = Form(""),
    foto_upload: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    current_id_autonomo = (current_id_autonomo or request.query_params.get("current_id_autonomo", "")).strip()
    autonomo = db.get(DimAutonomo, int(current_id_autonomo or id_autonomo)) if (current_id_autonomo or id_autonomo) else DimAutonomo()
    try:
        aplicar_id_manual(db, autonomo, DimAutonomo, "id_autonomo", id_autonomo, current_id_autonomo)
    except ValueError as e:
        return redirect_with_message("/autonomos", error=str(e))
    autonomo.nome_autonomo = nome_autonomo
    autonomo.cpf = cpf
    autonomo.telefone = telefone
    autonomo.email = email

    cargo_obj = db.get(DimCargoAutonomo, int(id_cargo_autonomo)) if id_cargo_autonomo else None
    autonomo.id_cargo_autonomo = int(id_cargo_autonomo) if id_cargo_autonomo else None
    autonomo.tipo_autonomo = cargo_obj.nome_cargo if cargo_obj else tipo_autonomo
    try:
        data = parse_date_safe(data_inclusao, "Data de inclusao invalida.")
    except ValueError as e:
        return redirect_with_message("/autonomos", error=str(e))
    autonomo.data_inclusao = data or autonomo.data_inclusao
    autonomo.status_autonomo = status_autonomo
    autonomo.observacoes = observacoes

    foto_salva = salvar_upload_foto(foto_upload, "autonomos")

    if foto_salva:
        autonomo.foto_url = foto_salva
    elif foto_url:
        autonomo.foto_url = foto_url
    db.add(autonomo)
    db.commit()
    return redirect_with_message("/autonomos", success="Autonomo salvo com sucesso.")


@router.get("/autonomos/atualizar-excel")
def atualizar_excel_form(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("cadastros/atualizar_autonomos.html", {
        "request": request,
        **flash_from_request(request),
    })


@router.post("/autonomos/atualizar-excel")
async def atualizar_excel_post(request: Request, arquivo: UploadFile = File(...), db: Session = Depends(get_db)):
    import io
    import openpyxl

    conteudo = await arquivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active

    # detecta linha de cabeçalho (primeira que tem "cpf" ou "nome" na primeira coluna)
    header_row = 1
    for r in ws.iter_rows(min_row=1, max_row=10):
        vals = [str(c.value or "").strip().lower() for c in r]
        if any("cpf" in v or "nome" in v for v in vals):
            header_row = r[0].row
            break

    headers = [str(c.value or "").strip().lower() for c in ws[header_row]]

    def _col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return None

    idx_cpf      = _col("cpf")
    idx_email    = _col("email")
    idx_nome     = _col("nome")
    idx_tel      = _col("telefone", "tel")
    idx_status   = _col("status")
    idx_data_inc = _col("data_inclusao", "data inclusao", "inclusao")

    atualizados = 0
    ignorados = 0
    nao_encontrados = []

    CAMPOS_STATUS = {'Ativo','Inativo','Hiato','Treinamento','Bloqueado','Desligado','Suspenso'}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue

        def cel(idx):
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        cpf   = cel(idx_cpf).replace(".", "").replace("-", "").replace(" ", "")
        email = cel(idx_email).lower()
        nome  = cel(idx_nome)

        # busca por CPF primeiro, depois email
        autonomo = None
        if cpf:
            autonomo = db.query(DimAutonomo).filter(
                text("REPLACE(REPLACE(REPLACE(cpf,'.',''),'-',''),' ','') = :cpf")
            ).params(cpf=cpf).first()
        if not autonomo and email:
            autonomo = db.query(DimAutonomo).filter(
                DimAutonomo.email == email
            ).first()

        if not autonomo:
            nao_encontrados.append(nome or cpf or email or "?")
            ignorados += 1
            continue

        # atualiza apenas os campos presentes na planilha
        if idx_tel is not None:
            v = cel(idx_tel)
            if v:
                autonomo.telefone = v
        if idx_status is not None:
            v = cel(idx_status)
            if v and v in CAMPOS_STATUS:
                autonomo.status_autonomo = v
        if idx_data_inc is not None:
            v = cel(idx_data_inc)
            if v:
                # normaliza datetime do Excel
                if len(v) >= 10 and v[4] == "-":
                    v = v[:10]
                autonomo.data_inclusao = v
        if idx_email is not None and not email:
            pass  # email veio vazio, não sobrescreve
        elif idx_email is not None:
            autonomo.email = email

        atualizados += 1

    db.commit()

    msg = f"{atualizados} autônomo(s) atualizado(s)."
    if nao_encontrados:
        msg += f" {ignorados} não encontrado(s): {', '.join(nao_encontrados[:5])}"
        if len(nao_encontrados) > 5:
            msg += f" e mais {len(nao_encontrados)-5}."
    return redirect_with_message("/autonomos/atualizar-excel", success=msg)


@router.post("/autonomos/{id_autonomo}/desligar")
def desligar_autonomo(id_autonomo: int, data_saida: str = Form(...), motivo_saida: str = Form(...), db: Session = Depends(get_db)):
    autonomo = db.get(DimAutonomo, id_autonomo)
    if not autonomo:
        return redirect_with_message("/autonomos", error="Autonomo nao encontrado.")
    autonomo.status_autonomo = "Desligado"
    try:
        autonomo.data_saida = parse_date(data_saida)
    except ValueError:
        return redirect_with_message("/autonomos", error="Data de saida invalida.")
    autonomo.motivo_saida = motivo_saida
    db.commit()
    return redirect_with_message("/autonomos", success="Autonomo desligado sem exclusao fisica.")


@router.get("/etapas")
def etapas(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("cadastros/etapas.html", {"request": request, "items": db.query(DimEtapa).order_by(DimEtapa.temporada.desc(), DimEtapa.nome_etapa).all(), **flash_from_request(request)})


@router.post("/etapas")
def salvar_etapa(
    request: Request,
    id_etapa: str = Form(""),
    current_id_etapa: str = "",
    temporada: str = Form(...),
    nome_etapa: str = Form(...),
    local: str = Form(""),
    data_inicio: str = Form(""),
    data_fim: str = Form(""),
    status_etapa: str = Form("Planejada"),
    observacoes: str = Form(""),
    foto_url: str = Form(""),
    foto_upload: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    current_id_etapa = (current_id_etapa or request.query_params.get("current_id_etapa", "")).strip()
    etapa = db.get(DimEtapa, int(current_id_etapa or id_etapa)) if (current_id_etapa or id_etapa) else DimEtapa()
    try:
        aplicar_id_manual(db, etapa, DimEtapa, "id_etapa", id_etapa, current_id_etapa)
    except ValueError as e:
        return redirect_with_message("/etapas", error=str(e))

    etapa.temporada = temporada
    etapa.nome_etapa = nome_etapa
    etapa.local = local
    try:
        data_inicio_parsed = parse_date_safe(data_inicio, "Data de inicio invalida.")
        data_fim_parsed = parse_date_safe(data_fim, "Data de fim invalida.")
    except ValueError as e:
        return redirect_with_message("/etapas", error=str(e))
    etapa.data_inicio = data_inicio_parsed or etapa.data_inicio
    etapa.data_fim = data_fim_parsed or etapa.data_fim
    etapa.status_etapa = status_etapa
    etapa.observacoes = observacoes

    db.add(etapa)
    db.commit()
    return redirect_with_message("/etapas", success="Etapa salva.")


@router.post("/etapas/{id_etapa}/inativar")
def inativar_etapa(id_etapa: int, db: Session = Depends(get_db)):
    etapa = db.get(DimEtapa, id_etapa)
    if not etapa:
        return redirect_with_message("/etapas", error="Etapa nao encontrada.")
    etapa.status_etapa = "Cancelada"
    db.commit()
    return redirect_with_message("/etapas", success="Etapa inativada.")


@router.get("/categorias")
@router.get("/provas")
def provas(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("cadastros/provas.html", {"request": request, "items": db.query(DimProva).order_by(DimProva.data_prova.desc()).all(), **lists(db), **flash_from_request(request)})


@router.post("/categorias")
@router.post("/provas")
def salvar_prova(id_prova: str = Form(""), current_id_prova: str = Form(""), id_etapa: int = Form(...), id_tipo_prova: int = Form(...), nome_prova: str = Form(...), data_prova: str = Form(""), status_prova: str = Form("Planejada"), observacoes: str = Form(""), db: Session = Depends(get_db)):
    prova = db.get(DimProva, int(current_id_prova or id_prova)) if (current_id_prova or id_prova) else DimProva()
    try:
        aplicar_id_manual(db, prova, DimProva, "id_prova", id_prova, current_id_prova)
    except ValueError as e:
        return redirect_with_message("/categorias", error=str(e))

    prova.id_etapa = id_etapa
    prova.id_tipo_prova = id_tipo_prova
    prova.nome_prova = nome_prova
    try:
        data_prova_parsed = parse_date_safe(data_prova, "Data da categoria invalida.")
    except ValueError as e:
        return redirect_with_message("/categorias", error=str(e))
    prova.data_prova = data_prova_parsed or prova.data_prova
    prova.status_prova = status_prova
    prova.observacoes = observacoes

    db.add(prova)
    db.commit()
    return redirect_with_message("/categorias", success="Categoria salva.")


@router.post("/provas/{id_prova}/inativar")
@router.post("/categorias/{id_prova}/inativar")
def inativar_prova(id_prova: int, db: Session = Depends(get_db)):
    prova = db.get(DimProva, id_prova)
    if not prova:
        return redirect_with_message("/categorias", error="Categoria nao encontrada.")
    prova.status_prova = "Cancelada"
    db.commit()
    return redirect_with_message("/categorias", success="Categoria inativada.")


@router.get("/tipos-categoria")
@router.get("/tipos-prova")
def tipos_prova(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("cadastros/tipos.html", {"request": request, "items": db.query(DimTipoProva).order_by(DimTipoProva.nome_tipo_prova).all(), **flash_from_request(request)})


@router.post("/tipos-categoria")
@router.post("/tipos-prova")
def salvar_tipo(
    request: Request,
    id_tipo_prova: str = Form(""),
    current_id_tipo_prova: str = "",
    nome_tipo_prova: str = Form(...),
    descricao: str = Form(""),
    status_tipo_prova: str = Form("Ativo"),
    db: Session = Depends(get_db),
):
    current_id_tipo_prova = (current_id_tipo_prova or request.query_params.get("current_id_tipo_prova", "")).strip()
    tipo = db.get(DimTipoProva, int(current_id_tipo_prova or id_tipo_prova)) if (current_id_tipo_prova or id_tipo_prova) else DimTipoProva()
    try:
        aplicar_id_manual(db, tipo, DimTipoProva, "id_tipo_prova", id_tipo_prova, current_id_tipo_prova)
    except ValueError as e:
        return redirect_with_message("/tipos-prova", error=str(e))

    if not tipo:
        return redirect_with_message("/tipos-prova", error="Tipo de prova não encontrado.")

    tipo.nome_tipo_prova = nome_tipo_prova
    tipo.descricao = descricao
    tipo.status_tipo_prova = status_tipo_prova

    db.add(tipo)
    db.commit()

    msg = "Tipo de prova atualizado." if (current_id_tipo_prova or id_tipo_prova) else "Tipo de prova cadastrado."
    return redirect_with_message("/tipos-prova", success=msg)


@router.get("/motivos-troca")
def motivos_troca(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("cadastros/motivos.html", {"request": request, "items": db.query(DimMotivoTroca).order_by(DimMotivoTroca.motivo_troca).all(), **flash_from_request(request)})


@router.post("/motivos-troca")
def salvar_motivo(
    request: Request,
    id_motivo_troca: str = Form(""),
    current_id_motivo_troca: str = Form(""),
    motivo_troca: str = Form(...),
    descricao: str = Form(""),
    status: str = Form("Ativo"),
    db: Session = Depends(get_db),
):
    current_id_motivo_troca = current_id_motivo_troca.strip()
    motivo = db.get(DimMotivoTroca, int(current_id_motivo_troca or id_motivo_troca)) if (current_id_motivo_troca or id_motivo_troca) else DimMotivoTroca()
    try:
        aplicar_id_manual(db, motivo, DimMotivoTroca, "id_motivo_troca", id_motivo_troca, current_id_motivo_troca)
    except ValueError as e:
        return redirect_with_message("/motivos-troca", error=str(e))

    if not motivo:
        return redirect_with_message("/motivos-troca", error="Motivo de troca não encontrado.")

    motivo.motivo_troca = motivo_troca
    motivo.descricao = descricao
    motivo.status = status

    db.add(motivo)
    db.commit()

    msg = "Motivo de troca atualizado." if (current_id_motivo_troca or id_motivo_troca) else "Motivo cadastrado."
    return redirect_with_message("/motivos-troca", success=msg)


@router.get("/cargos-autonomos")
def cargos_autonomos(request: Request, db: Session = Depends(get_db)):
    items = db.query(DimCargoAutonomo).order_by(DimCargoAutonomo.nome_cargo).all()

    return templates.TemplateResponse(
        "cadastros/cargos_autonomos.html",
        {
            "request": request,
            "items": items,
            **flash_from_request(request),
        },
    )


@router.post("/cargos-autonomos")
def salvar_cargo_autonomo(
    request: Request,
    id_cargo_autonomo: str = Form(""),
    current_id_cargo_autonomo: str = Form(""),
    nome_cargo: str = Form(...),
    descricao: str = Form(""),
    status: str = Form("Ativo"),
    db: Session = Depends(get_db),
):
    # fallback: lê da query string se não vier no body (compatibilidade)
    current_id_cargo_autonomo = (current_id_cargo_autonomo or request.query_params.get("current_id_cargo_autonomo", "")).strip()
    cargo = db.get(DimCargoAutonomo, int(current_id_cargo_autonomo or id_cargo_autonomo)) if (current_id_cargo_autonomo or id_cargo_autonomo) else DimCargoAutonomo()
    try:
        aplicar_id_manual(db, cargo, DimCargoAutonomo, "id_cargo_autonomo", id_cargo_autonomo, current_id_cargo_autonomo)
    except ValueError as e:
        return redirect_with_message("/cargos-autonomos", error=str(e))

    if not cargo:
        return redirect_with_message("/cargos-autonomos", error="Cargo não encontrado.")

    cargo.nome_cargo = nome_cargo
    cargo.descricao = descricao
    cargo.status = status

    db.add(cargo)
    db.commit()

    msg = "Cargo atualizado." if (current_id_cargo_autonomo or id_cargo_autonomo) else "Cargo cadastrado."
    return redirect_with_message("/cargos-autonomos", success=msg)

