from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal
from html import escape
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, inspect, text

from app.database import DATABASE_URL

try:
    from app.logging_utils import log_error
except Exception:
    def log_error(contexto, exc, extra=None, excel=False):
        print(contexto, exc, extra)
        return "sem_log"


router = APIRouter(prefix="/excel", tags=["Excel"])

BASE_DIR = Path(__file__).resolve().parents[2]
DB_PATH = BASE_DIR / "data" / "app.db"
IMPORT_ENGINE = create_engine(DATABASE_URL, future=True)


def extract_db_error_details(exc: Exception) -> dict[str, Any]:
    details: dict[str, Any] = {"erro": str(exc)}
    orig = getattr(exc, "orig", None)
    if orig is None:
        return details

    details["db_error_type"] = type(orig).__name__
    details["db_error_message"] = str(orig)
    details["pgcode"] = getattr(orig, "pgcode", None)

    diag = getattr(orig, "diag", None)
    if diag is not None:
        details["constraint_name"] = getattr(diag, "constraint_name", None)
        details["table_name"] = getattr(diag, "table_name", None)
        details["column_name"] = getattr(diag, "column_name", None)
        details["schema_name"] = getattr(diag, "schema_name", None)
        details["detail"] = getattr(diag, "message_detail", None)
        details["hint"] = getattr(diag, "message_hint", None)

    return details


ENTIDADES = {
    "pilotos": {
        "label": "Pilotos",
        "table": "dim_pilotos",
        "unique": ["cpf"],
        "columns": [
            "nome_piloto",
            "cpf",
            "telefone",
            "email",
            "foto_url",
            "data_inclusao",
            "data_desligamento",
            "motivo_desligamento",
            "status_piloto",
            "observacoes",
        ],
        "example": [
            "Rafael Martins",
            "111.111.111-11",
            "(11) 99999-1001",
            "rafael@email.com",
            "https://link-da-foto.com/foto.jpg",
            "2026-01-10",
            "",
            "",
            "Ativo",
            "Exemplo",
        ],
        "tips": [
            "Use o CPF para evitar duplicidade do mesmo piloto.",
        ],
    },
    "cargos-autonomos": {
        "label": "Cargos de Autonomos",
        "table": "dim_cargos_autonomos",
        "unique": ["nome_cargo"],
        "columns": ["nome_cargo", "descricao", "status"],
        "example": [
            "Mecanico",
            "Profissional responsavel pela parte mecanica",
            "Ativo",
        ],
    },
    "autonomos": {
        "label": "Autonomos",
        "table": "dim_autonomos",
        "unique": ["cpf"],
        "columns": [
            "nome_autonomo",
            "cpf",
            "telefone",
            "email",
            "id_cargo_autonomo",
            "tipo_autonomo",
            "data_inclusao",
            "data_saida",
            "motivo_saida",
            "status_autonomo",
            "observacoes",
        ],
        "example": [
            "Joao Silva",
            "555.555.555-55",
            "(11) 98888-1001",
            "joao@email.com",
            "1",
            "Mecanico",
            "2026-01-05",
            "",
            "",
            "Ativo",
            "Exemplo",
        ],
        "references": ["cargos"],
        "tips": [
            "O campo id_cargo_autonomo deve ser buscado na aba REF_CARGOS deste arquivo.",
        ],
    },
    "etapas": {
        "label": "Etapas",
        "table": "dim_etapas",
        "unique": ["temporada", "nome_etapa"],
        "columns": [
            "temporada",
            "nome_etapa",
            "local",
            "data_inicio",
            "data_fim",
            "status_etapa",
            "observacoes",
        ],
        "example": [
            "2026",
            "Etapa 01 - Interlagos",
            "Sao Paulo/SP",
            "2026-03-13",
            "2026-03-15",
            "Planejada",
            "Exemplo",
        ],
    },
    "tipos-prova": {
        "label": "Tipos de Categoria",
        "table": "dim_tipos_categoria",
        "unique": ["nome_tipo_prova"],
        "columns": ["nome_tipo_prova", "descricao", "status_tipo_prova"],
        "example": [
            "Carrera Cup",
            "Tipo de categoria",
            "Ativo",
        ],
    },
    "carros": {
        "label": "Carros",
        "table": "dim_carros",
        "unique": ["numero_carro"],
        "columns": [
            "numero_carro",
            "modelo",
            "categoria_padrao",
            "chassi",
            "status_carro",
            "observacoes",
        ],
        "example": [
            "77",
            "911 GT3 Cup",
            "Cup",
            "CHASSI-001",
            "Ativo",
            "Exemplo",
        ],
    },
    "provas": {
        "label": "Categorias",
        "table": "dim_categorias",
        "unique": ["id_etapa", "id_tipo_prova", "nome_prova"],
        "columns": [
            "id_etapa",
            "id_tipo_prova",
            "nome_prova",
            "data_prova",
            "status_prova",
            "observacoes",
        ],
        "example": [
            "1",
            "1",
            "Sprint Challenge",
            "2026-03-14",
            "Planejada",
            "Exemplo",
        ],
        "references": ["etapas", "tipos_prova"],
        "tips": [
            "Para relacionamentos, escolha os IDs das abas REF_ETAPAS e REF_TIPOS_PROVA.",
        ],
    },
    "motivos-troca": {
        "label": "Motivos de Troca",
        "table": "dim_motivos_troca",
        "unique": ["motivo_troca"],
        "columns": ["motivo_troca", "descricao", "status"],
        "example": [
            "Solicitacao do piloto",
            "Troca solicitada pelo piloto",
            "Ativo",
        ],
    },
    "status-pagamento": {
        "label": "Status Pagamento",
        "table": "dim_status_pagamento",
        "unique": ["status_pagamento"],
        "columns": ["status_pagamento"],
        "example": ["Pendente"],
    },
    "alocacoes": {
        "label": "Alocacoes",
        "table": "fato_piloto_autonomo_prova",
        "unique": [
            "id_piloto",
            "id_autonomo",
            "id_etapa",
            "id_prova",
            "id_carro",
        ],
        "columns": [
            "id_piloto",
            "id_autonomo",
            "id_etapa",
            "id_prova",
            "id_carro",
            "id_cargo_autonomo",
            "status_vinculo",
            "foi_substituido",
            "id_autonomo_substituto",
            "data_troca",
            "id_motivo_troca",
            "justificativa_troca",
            "valor_fechado_etapa",
            "dias_trabalhados",
            "link_avaliacao_externa",
            "documento",
            "observacoes",
        ],
        "example": [
            "1",
            "1",
            "1",
            "1",
            "12",
            "1",
            "Ativo",
            "Nao",
            "",
            "",
            "",
            "",
            "3300",
            "3",
            "https://forms.gle/exemplo",
            "NF-001",
            "Exemplo",
        ],
        "references": ["pilotos", "autonomos", "etapas", "provas", "carros", "cargos", "motivos"],
        "tips": [
            "Nao informe o nome da funcao. Informe apenas id_cargo_autonomo e o sistema grava funcao_autonomo automaticamente.",
            "Os IDs usados nos relacionamentos estao nas abas REF_* deste arquivo.",
            "valor_fechado_etapa e o valor total do pacote, nao quantidade x valor unitario.",
        ],
    },
}


REFERENCE_SHEETS = {
    "pilotos": {
        "title": "REF_PILOTOS",
        "table": "dim_pilotos",
        "columns": ["id_piloto", "nome_piloto", "status_piloto"],
        "labels": ["ID Piloto", "Piloto", "Status"],
        "order_by": ["nome_piloto"],
        "field": "id_piloto",
    },
    "autonomos": {
        "title": "REF_AUTONOMOS",
        "table": "dim_autonomos",
        "columns": ["id_autonomo", "nome_autonomo", "id_cargo_autonomo", "status_autonomo"],
        "labels": ["ID Autonomo", "Autonomo", "ID Cargo", "Status"],
        "order_by": ["nome_autonomo"],
        "field": "id_autonomo",
    },
    "etapas": {
        "title": "REF_ETAPAS",
        "table": "dim_etapas",
        "columns": ["id_etapa", "temporada", "nome_etapa"],
        "labels": ["ID Etapa", "Temporada", "Etapa"],
        "order_by": ["temporada", "nome_etapa"],
        "field": "id_etapa",
    },
    "tipos_prova": {
        "title": "REF_TIPOS_PROVA",
        "table": "dim_tipos_categoria",
        "columns": ["id_tipo_prova", "nome_tipo_prova", "status_tipo_prova"],
        "labels": ["ID Tipo", "Tipo de Prova", "Status"],
        "order_by": ["nome_tipo_prova"],
        "field": "id_tipo_prova",
    },
    "provas": {
        "title": "REF_PROVAS",
        "table": "dim_categorias",
        "columns": ["id_prova", "nome_prova", "id_etapa", "id_tipo_prova", "status_prova"],
        "labels": ["ID Prova", "Categoria", "ID Etapa", "ID Tipo", "Status"],
        "order_by": ["nome_prova"],
        "field": "id_prova",
    },
    "carros": {
        "title": "REF_CARROS",
        "table": "dim_carros",
        "columns": ["id_carro", "numero_carro", "modelo", "status_carro"],
        "labels": ["ID Carro", "Numero", "Modelo", "Status"],
        "order_by": ["numero_carro"],
        "field": "id_carro",
    },
    "cargos": {
        "title": "REF_CARGOS",
        "table": "dim_cargos_autonomos",
        "columns": ["id_cargo_autonomo", "nome_cargo", "status"],
        "labels": ["ID Cargo", "Cargo", "Status"],
        "order_by": ["nome_cargo"],
        "field": "id_cargo_autonomo",
    },
    "motivos": {
        "title": "REF_MOTIVOS_TROCA",
        "table": "dim_motivos_troca",
        "columns": ["id_motivo_troca", "motivo_troca", "status"],
        "labels": ["ID Motivo", "Motivo da Troca", "Status"],
        "order_by": ["motivo_troca"],
        "field": "id_motivo_troca",
    },
    "status_pagamento": {
        "title": "REF_STATUS_PAGAMENTO",
        "table": "dim_status_pagamento",
        "columns": ["id_status_pagamento", "status_pagamento"],
        "labels": ["ID Status", "Status Pagamento"],
        "order_by": ["status_pagamento"],
        "field": "id_status_pagamento",
    },
}


FIELD_HINTS = {
    "id_piloto": "Use o ID da aba REF_PILOTOS.",
    "id_autonomo": "Use o ID da aba REF_AUTONOMOS.",
    "id_etapa": "Use o ID da aba REF_ETAPAS.",
    "id_prova": "Use o ID da aba REF_PROVAS.",
    "id_carro": "Use o ID da aba REF_CARROS.",
    "id_cargo_autonomo": "Use o ID da aba REF_CARGOS. Nao precisa repetir o nome do cargo.",
    "id_autonomo_substituto": "Quando houver troca, use o ID da aba REF_AUTONOMOS.",
    "id_motivo_troca": "Quando houver troca, use o ID da aba REF_MOTIVOS_TROCA.",
    "valor_fechado_etapa": "Informe o valor total do pacote da etapa.",
    "dias_trabalhados": "Informe a quantidade de dias trabalhados na etapa.",
    "status_pagamento": "Use um valor cadastrado em status-pagamento, como Pendente, Aprovado ou Pago.",
}


def prettify_field_name(field_name: str) -> str:
    return str(field_name or "").replace("_", " ")


def build_error_block(title: str, message: str, error_id: str, back_href: str = "/excel/") -> str:
    return f"""
    <!doctype html>
    <html lang="pt-br">
    <head>
        <meta charset="utf-8">
        <title>{escape(title)}</title>
        <style>
            body {{ margin:0; font-family:Arial; background:#0f172a; color:#e5e7eb; }}
            .page {{ max-width:900px; margin:0 auto; padding:32px; }}
            .card {{ background:#111827; border:1px solid #1f2937; border-radius:18px; padding:22px; }}
            .alert {{ background:#1f2937; border:1px solid #374151; border-radius:12px; padding:14px; margin:16px 0; }}
            a {{ color:#fca5a5; font-weight:700; }}
            code {{ color:#fde68a; }}
            p, li {{ line-height:1.5; }}
        </style>
    </head>
    <body>
        <div class="page">
            <div class="card">
                <h1>{escape(title)}</h1>
                <div class="alert">{message}</div>
                <p><b>ID do erro:</b> <code>{escape(error_id)}</code></p>
                <p><a href="{escape(back_href)}">Voltar para Excel</a> | <a href="/logs/excel">Ver log Excel</a></p>
            </div>
        </div>
    </body>
    </html>
    """


def explain_read_error(exc: Exception) -> str:
    message = str(exc).lower()
    if "file is not a zip file" in message:
        return "O arquivo enviado nao parece ser um .xlsx valido. Salve novamente a planilha em formato Excel antes de importar."
    if "worksheet" in message or "sheet" in message:
        return "Nao foi possivel localizar a primeira aba da planilha. Confira se o arquivo possui pelo menos uma aba preenchida."
    return "Nao foi possivel ler o arquivo Excel. Confira se ele foi salvo como .xlsx e se a primeira aba contem o layout da importacao."


def explain_import_error(exc: Exception, row: dict[str, Any] | None = None) -> str:
    raw_message = str(exc)
    lowered = raw_message.lower()

    if isinstance(exc, ValueError):
        return raw_message

    match = re.search(r'null value in column "([^"]+)"', raw_message, re.IGNORECASE)
    if match:
        field_name = match.group(1)
        return f"O campo obrigatorio '{prettify_field_name(field_name)}' ficou vazio para esta linha."

    match = re.search(r'foreign key constraint', raw_message, re.IGNORECASE)
    if match:
        return "Um dos IDs informados nao existe no cadastro relacionado. Revise os campos de relacionamento usando as abas REF_* do modelo."

    if "duplicate key value violates unique constraint" in lowered:
        return "Ja existe um registro com essa combinacao de dados unicos. Revise a linha para evitar duplicidade."

    if "invalid input syntax for type integer" in lowered:
        return "Ha um campo de ID ou numero com formato invalido. Use apenas numeros inteiros nos campos de identificacao."

    if "invalid input syntax for type numeric" in lowered or "numeric field overflow" in lowered:
        return "Ha um valor numerico invalido na linha. Revise campos como valor_fechado_etapa e dias_trabalhados."

    if "date/time field value out of range" in lowered or "invalid input syntax for type date" in lowered:
        return "Ha uma data invalida na linha. Use o formato DD/MM/AAAA ou AAAA-MM-DD."

    if row:
        known_ids = [key for key, value in row.items() if key.startswith("id_") and value not in [None, ""]]
        if known_ids and ("violates" in lowered or "constraint" in lowered):
            return "Existe um problema de relacionamento nesta linha. Revise os IDs informados e compare com as abas REF_* do arquivo."

    return "Nao foi possivel importar esta linha. Revise os valores preenchidos e, se precisar, consulte o log tecnico pelo ID do erro."


def validate_import_headers(entity_key: str, headers: list[str]) -> list[str]:
    expected_headers = ENTIDADES[entity_key]["columns"]
    normalized_expected = {norm_header(header): header for header in expected_headers}
    missing = [original for normalized, original in normalized_expected.items() if normalized not in headers]
    return missing


def get_conn():
    return IMPORT_ENGINE.connect()


def norm_header(value: Any) -> str:
    value = str(value or "").strip().lower()
    value = value.replace(" ", "_").replace("-", "_")
    return re.sub(r"[^a-z0-9_]", "", value)


def normalize_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        text_value = value.strip()
        if text_value == "":
            return None

        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text_value, fmt).date()
            except Exception:
                pass

        money = text_value.replace("R$", "").replace(".", "").replace(",", ".").strip()
        if re.fullmatch(r"-?\d+(\.\d+)?", money) and any(token in text_value for token in ["R$", ","]):
            try:
                return float(Decimal(money))
            except Exception:
                return text_value

        return text_value

    return value


def table_exists(conn, table: str) -> bool:
    inspector = inspect(conn)
    return table in inspector.get_table_names()


def get_table_columns(conn, table: str) -> list[str]:
    inspector = inspect(conn)
    return [col["name"] for col in inspector.get_columns(table)]


def get_table_column_defs(conn, table: str) -> dict[str, Any]:
    inspector = inspect(conn)
    return {col["name"]: col for col in inspector.get_columns(table)}


def get_pk_column(conn, table: str) -> str | None:
    inspector = inspect(conn)
    for col in inspector.get_columns(table):
        if col.get("primary_key", 0):
            return col["name"]
    return None


def quote_ident(identifier: str) -> str:
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", identifier or ""):
        raise ValueError(f"Identificador SQL invalido: {identifier}")
    return f'"{identifier}"'


def safe_sheet_title(value: str) -> str:
    title = str(value or "Modelo")
    for invalid in [":", "\\", "/", "?", "*", "[", "]"]:
        title = title.replace(invalid, "-")
    title = title.strip() or "Modelo"
    return title[:31]


def fetch_reference_rows(conn, ref_key: str) -> list[dict[str, Any]]:
    cfg = REFERENCE_SHEETS[ref_key]
    if not table_exists(conn, cfg["table"]):
        return []

    select_list = ", ".join(quote_ident(col) for col in cfg["columns"])
    order_clause = ", ".join(quote_ident(col) for col in cfg["order_by"])
    sql = text(f"SELECT {select_list} FROM {quote_ident(cfg['table'])} ORDER BY {order_clause}")
    result = conn.execute(sql)
    return [dict(row._mapping) for row in result]


def autosize_worksheet(ws):
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_len = 0
        for cell in column:
            current = len(str(cell.value or ""))
            if current > max_len:
                max_len = current
        ws.column_dimensions[letter].width = min(max(max_len + 3, 16), 48)


def style_header_row(ws, row_number: int = 1):
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def append_reference_sheet(wb: Workbook, conn, ref_key: str):
    cfg = REFERENCE_SHEETS[ref_key]
    ws = wb.create_sheet(safe_sheet_title(cfg["title"]))
    ws.append(cfg["labels"])
    style_header_row(ws)

    for row in fetch_reference_rows(conn, ref_key):
        ws.append([normalize_value(row.get(col)) for col in cfg["columns"]])

    ws.freeze_panes = "A2"
    autosize_worksheet(ws)


def build_readme_sheet(wb: Workbook, entity_key: str):
    cfg = ENTIDADES[entity_key]
    leia = wb.create_sheet("LEIA-ME")

    leia["A1"] = f"Como preencher o modelo de {cfg['label']}"
    leia["A1"].font = Font(bold=True, size=14)
    leia["A3"] = "1. Preencha somente a primeira aba do arquivo."
    leia["A4"] = "2. Nao renomeie colunas nem exclua cabecalhos."
    leia["A5"] = "3. Datas podem estar em DD/MM/AAAA ou AAAA-MM-DD."
    leia["A6"] = "4. Os IDs de relacionamento estao nas abas REF_* deste mesmo arquivo."
    leia["A7"] = "5. Se ja existe ID para relacionar, preencha o ID. Nao repita o nome da dimensao, salvo quando o layout pedir explicitamente."

    extra_tips = cfg.get("tips", [])
    tip_row = 9
    if extra_tips:
        leia[f"A{tip_row}"] = "Observacoes importantes"
        leia[f"A{tip_row}"].font = Font(bold=True)
        tip_row += 1
        for index, tip in enumerate(extra_tips, start=1):
            leia[f"A{tip_row}"] = f"- {tip}"
            tip_row += 1
        tip_row += 1

    leia[f"A{tip_row}"] = "Campos do layout"
    leia[f"A{tip_row}"].font = Font(bold=True)
    header_row = tip_row + 1
    leia[f"A{header_row}"] = "Campo"
    leia[f"B{header_row}"] = "Como preencher"
    style_header_row(leia, header_row)

    current_row = header_row + 1
    for column in cfg["columns"]:
        leia[f"A{current_row}"] = column
        leia[f"B{current_row}"] = FIELD_HINTS.get(column, "Preencha conforme a regra do cadastro.")
        current_row += 1

    current_row += 1
    references = cfg.get("references", [])
    if references:
        leia[f"A{current_row}"] = "Tabelas dimensao para relacionamento"
        leia[f"A{current_row}"].font = Font(bold=True)
        current_row += 1
        leia[f"A{current_row}"] = "Campo"
        leia[f"B{current_row}"] = "Aba de referencia"
        leia[f"C{current_row}"] = "Tabela origem"
        style_header_row(leia, current_row)
        current_row += 1
        for ref_key in references:
            ref_cfg = REFERENCE_SHEETS[ref_key]
            leia[f"A{current_row}"] = ref_cfg["field"]
            leia[f"B{current_row}"] = ref_cfg["title"]
            leia[f"C{current_row}"] = ref_cfg["table"]
            current_row += 1

    leia.freeze_panes = "A3"
    autosize_worksheet(leia)
    leia.column_dimensions["B"].width = 70


def make_template(entity_key: str):
    cfg = ENTIDADES[entity_key]
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(cfg["label"])

    ws.append(cfg["columns"])
    ws.append(cfg["example"])
    style_header_row(ws)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    with get_conn() as conn:
        build_readme_sheet(wb, entity_key)
        for ref_key in cfg.get("references", []):
            append_reference_sheet(wb, conn, ref_key)

    return wb


def workbook_stream(wb: Workbook):
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def read_rows_from_excel(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [norm_header(cell.value) for cell in ws[1]]
    rows = []

    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
        if all(value is None or str(value).strip() == "" for value in values):
            continue

        row = {}
        for header, value in zip(headers, values):
            if header:
                row[header] = normalize_value(value)

        row["_linha_excel"] = row_idx
        rows.append(row)

    return headers, rows


def resolve_cargo_nome(conn, cargo_id: Any) -> str | None:
    if cargo_id in [None, ""]:
        return None

    sql = text(
        "SELECT nome_cargo FROM dim_cargos_autonomos WHERE id_cargo_autonomo = :cargo_id LIMIT 1"
    )
    result = conn.execute(sql, {"cargo_id": cargo_id}).scalar()
    return result


def resolve_cargo_nome_from_autonomo(conn, autonomo_id: Any) -> str | None:
    if autonomo_id in [None, ""]:
        return None

    sql = text(
        """
        SELECT c.nome_cargo
        FROM dim_autonomos a
        LEFT JOIN dim_cargos_autonomos c
            ON c.id_cargo_autonomo = a.id_cargo_autonomo
        WHERE a.id_autonomo = :autonomo_id
        LIMIT 1
        """
    )
    result = conn.execute(sql, {"autonomo_id": autonomo_id}).scalar()
    return result


def resolve_autonomo_cargo_id(conn, autonomo_id: Any) -> Any:
    if autonomo_id in [None, ""]:
        return None

    sql = text(
        "SELECT id_cargo_autonomo FROM dim_autonomos WHERE id_autonomo = :autonomo_id LIMIT 1"
    )
    return conn.execute(sql, {"autonomo_id": autonomo_id}).scalar()


def prepare_row_for_entity(conn, entity_key: str, row: dict[str, Any]) -> dict[str, Any]:
    prepared = dict(row)

    if entity_key == "alocacoes":
        cargo_id = prepared.get("id_cargo_autonomo")
        autonomo_cargo_id = resolve_autonomo_cargo_id(conn, prepared.get("id_autonomo"))

        if cargo_id in [None, ""]:
            raise ValueError("Informe um id_cargo_autonomo valido para a alocacao.")

        if autonomo_cargo_id not in [None, ""] and str(autonomo_cargo_id) != str(cargo_id):
            raise ValueError(
                "O id_cargo_autonomo informado nao corresponde ao cargo cadastrado para o autonomo."
            )

        cargo_nome = resolve_cargo_nome(conn, cargo_id)

        if cargo_nome is None:
            raise ValueError(
                "Nao foi possivel descobrir a funcao_autonomo. Informe um id_cargo_autonomo valido."
            )

        prepared["funcao_autonomo"] = cargo_nome
        prepared["criado_em"] = prepared.get("criado_em") or datetime.utcnow()
        prepared["atualizado_em"] = datetime.utcnow()
        prepared["data_inicio_vinculo"] = prepared.get("data_inicio_vinculo") or date.today()
        prepared.pop("id_cargo_autonomo", None)

    return prepared


def insert_or_update(conn, table: str, row: dict[str, Any], unique_keys: list[str]):
    db_col_defs = get_table_column_defs(conn, table)
    db_cols = set(db_col_defs.keys())
    pk = get_pk_column(conn, table)

    clean = {
        key: normalize_value(value)
        for key, value in row.items()
        if key in db_cols and key != pk and not key.startswith("_")
    }

    if not clean:
        return "ignorado"

    for key, value in list(clean.items()):
        col_def = db_col_defs.get(key) or {}
        col_type = str(col_def.get("type", "")).lower()
        if value is not None and any(token in col_type for token in ["char", "text", "varchar"]):
            clean[key] = str(value)

    usable_keys = [key for key in unique_keys if key in clean and clean.get(key) not in [None, ""]]
    existing = None

    if usable_keys:
        where = " AND ".join([f"{quote_ident(key)} = :w_{key}" for key in usable_keys])
        where_params = {f"w_{key}": clean[key] for key in usable_keys}
        existing = conn.execute(
            text(f"SELECT 1 FROM {quote_ident(table)} WHERE {where} LIMIT 1"),
            where_params,
        ).first()

    if existing:
        set_cols = [key for key in clean.keys() if key not in usable_keys]
        if set_cols:
            set_clause = ", ".join([f"{quote_ident(col)} = :s_{col}" for col in set_cols])
            sql = text(f"UPDATE {quote_ident(table)} SET {set_clause} WHERE {where}")
            params = {f"s_{col}": clean[col] for col in set_cols}
            params.update(where_params)
            conn.execute(sql, params)
        return "atualizado"

    cols = list(clean.keys())
    col_list = ", ".join([quote_ident(col) for col in cols])
    val_list = ", ".join([f":i_{col}" for col in cols])
    sql = text(f"INSERT INTO {quote_ident(table)} ({col_list}) VALUES ({val_list})")
    conn.execute(sql, {f"i_{col}": clean[col] for col in cols})
    return "criado"


@router.get("/", response_class=HTMLResponse)
def excel_home():
    cards = ""

    for key, cfg in ENTIDADES.items():
        cards += f"""
        <div class="card">
            <h3>{cfg["label"]}</h3>
            <p>Tabela: <code>{cfg["table"]}</code></p>
            <p>O modelo ja inclui aba LEIA-ME e abas REF_* com as dimensoes relacionadas.</p>
            <a class="btn primary" href="/excel/modelo/{key}">Baixar modelo Excel</a>
            <form action="/excel/importar/{key}" method="post" enctype="multipart/form-data">
                <input type="file" name="arquivo" accept=".xlsx" required>
                <button class="btn" type="submit">Importar Excel</button>
            </form>
        </div>
        """

    return HTMLResponse(
        f"""
        <!doctype html>
        <html lang="pt-br">
        <head>
            <meta charset="utf-8">
            <title>Importacoes Excel</title>
            <style>
                body {{ margin:0; font-family:Arial; background:#0f172a; color:#e5e7eb; }}
                .page {{ max-width:1200px; margin:0 auto; padding:32px; }}
                .top {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:22px; gap:16px; }}
                .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(300px,1fr)); gap:16px; }}
                .card {{ background:#111827; border:1px solid #1f2937; border-radius:18px; padding:20px; }}
                .btn {{ display:inline-block; border:0; border-radius:10px; padding:10px 14px; margin:8px 0; color:white; background:#374151; text-decoration:none; font-weight:700; cursor:pointer; }}
                .primary {{ background:#dc2626; }}
                input {{ display:block; margin:8px 0; color:#e5e7eb; }}
                code {{ color:#fca5a5; }}
                a {{ color:#fca5a5; }}
                p {{ line-height:1.45; }}
            </style>
        </head>
        <body>
            <div class="page">
                <div class="top">
                    <div>
                        <h1>Importacoes e Modelos Excel</h1>
                        <p>Baixe os modelos, veja a aba LEIA-ME, consulte as abas REF_* e depois importe.</p>
                    </div>
                    <a href="/">Voltar</a>
                </div>
                <div class="grid">{cards}</div>
            </div>
        </body>
        </html>
        """
    )


@router.get("/modelo/{entity_key}")
def baixar_modelo(entity_key: str):
    if entity_key not in ENTIDADES:
        return HTMLResponse("Modelo nao encontrado.", status_code=404)

    wb = make_template(entity_key)
    filename = f"modelo_{entity_key}.xlsx"

    return StreamingResponse(
        workbook_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/importar/{entity_key}", response_class=HTMLResponse)
async def importar_excel(entity_key: str, arquivo: UploadFile = File(...)):
    if entity_key not in ENTIDADES:
        return HTMLResponse("Importacao nao encontrada.", status_code=404)

    if not arquivo.filename.lower().endswith(".xlsx"):
        return HTMLResponse("Envie um arquivo .xlsx.", status_code=400)

    cfg = ENTIDADES[entity_key]
    content = await arquivo.read()

    try:
        headers, rows = read_rows_from_excel(content)
    except Exception as exc:
        error_id = log_error(
            "ERRO_AO_LER_EXCEL",
            exc,
            {"entity_key": entity_key, "arquivo": arquivo.filename},
            excel=True,
        )
        return HTMLResponse(
            build_error_block(
                "Erro ao ler Excel",
                f"<p>{escape(explain_read_error(exc))}</p><p>Arquivo: <b>{escape(arquivo.filename)}</b></p>",
                error_id,
            ),
            status_code=400,
        )

    missing_headers = validate_import_headers(entity_key, headers)
    if missing_headers:
        error_id = log_error(
            "ERRO_CABECALHO_IMPORTACAO_EXCEL",
            ValueError(f"Colunas ausentes: {', '.join(missing_headers)}"),
            {
                "entity_key": entity_key,
                "arquivo": arquivo.filename,
                "headers_lidos": headers,
                "headers_ausentes": missing_headers,
            },
            excel=True,
        )
        missing_list = "".join([f"<li><code>{escape(col)}</code></li>" for col in missing_headers])
        return HTMLResponse(
            build_error_block(
                "Cabecalho incompleto",
                (
                    "<p>O arquivo nao contem todas as colunas esperadas para esta importacao.</p>"
                    f"<p>Colunas ausentes:</p><ul>{missing_list}</ul>"
                    "<p>Baixe novamente o modelo Excel desta entidade e copie os dados para o layout correto.</p>"
                ),
                error_id,
            ),
            status_code=400,
        )

    if not rows:
        error_id = log_error(
            "ERRO_PLANILHA_SEM_DADOS",
            ValueError("Planilha sem dados para importar."),
            {"entity_key": entity_key, "arquivo": arquivo.filename, "headers": headers},
            excel=True,
        )
        return HTMLResponse(
            build_error_block(
                "Planilha sem dados",
                "<p>Nenhuma linha de dados foi encontrada na primeira aba do arquivo.</p><p>Preencha pelo menos uma linha abaixo do cabecalho antes de importar.</p>",
                error_id,
            ),
            status_code=400,
        )

    criados = 0
    atualizados = 0
    ignorados = 0
    erros = []

    conn = get_conn()
    trans = conn.begin()

    try:
        if not table_exists(conn, cfg["table"]):
            raise Exception(f"Tabela nao encontrada: {cfg['table']}")

        for row in rows:
            linha = row.get("_linha_excel")
            row_tx = conn.begin_nested()

            try:
                prepared_row = prepare_row_for_entity(conn, entity_key, row)
                status = insert_or_update(conn, cfg["table"], prepared_row, cfg["unique"])
                row_tx.commit()

                if status == "criado":
                    criados += 1
                elif status == "atualizado":
                    atualizados += 1
                else:
                    ignorados += 1
            except Exception as exc:
                row_tx.rollback()
                db_error_details = extract_db_error_details(exc)
                error_id = log_error(
                    "ERRO_LINHA_IMPORTACAO_EXCEL",
                    exc,
                    {
                        "entity_key": entity_key,
                        "table": cfg["table"],
                        "arquivo": arquivo.filename,
                        "linha_excel": linha,
                        "headers": headers,
                        "row": row,
                        "db_error": db_error_details,
                    },
                    excel=True,
                )
                erros.append((linha, error_id, explain_import_error(exc, row)))

        trans.commit()
    except Exception as exc:
        trans.rollback()
        error_id = log_error(
            "ERRO_IMPORTACAO_EXCEL",
            exc,
            {"entity_key": entity_key, "table": cfg["table"], "arquivo": arquivo.filename},
            excel=True,
        )
        return HTMLResponse(
            build_error_block(
                "Erro na importacao",
                (
                    f"<p>{escape(explain_import_error(exc))}</p>"
                    f"<p>Entidade: <b>{escape(cfg['label'])}</b></p>"
                    f"<p>Arquivo: <b>{escape(arquivo.filename)}</b></p>"
                ),
                error_id,
            ),
            status_code=400,
        )
    finally:
        conn.close()

    linhas_erro = "".join(
        [
            f"<tr><td>{linha}</td><td><code>{escape(error_id)}</code></td><td>{escape(erro)}</td></tr>"
            for linha, error_id, erro in erros
        ]
    )

    return HTMLResponse(
        f"""
        <!doctype html>
        <html lang="pt-br">
        <head>
            <meta charset="utf-8">
            <title>Resultado Importacao</title>
            <style>
                body {{ margin:0; font-family:Arial; background:#0f172a; color:#e5e7eb; }}
                .page {{ max-width:900px; margin:0 auto; padding:32px; }}
                .card {{ background:#111827; border:1px solid #1f2937; border-radius:18px; padding:22px; }}
                table {{ width:100%; border-collapse:collapse; margin-top:16px; }}
                th, td {{ border-bottom:1px solid #374151; padding:10px; text-align:left; }}
                a {{ color:#fca5a5; font-weight:700; }}
                .summary {{ background:#1f2937; border:1px solid #374151; border-radius:12px; padding:14px; margin:16px 0; }}
                .ok {{ color:#86efac; }}
                .warn {{ color:#fcd34d; }}
            </style>
        </head>
        <body>
            <div class="page">
                <div class="card">
                    <h1>Resultado da Importacao</h1>
                    <p><b>Arquivo:</b> {arquivo.filename}</p>
                    <p><b>Entidade:</b> {cfg["label"]}</p>
                    <p>Linhas lidas: <b>{len(rows)}</b></p>
                    <div class="summary">
                        <p class="ok">Criados: <b>{criados}</b> | Atualizados: <b>{atualizados}</b> | Ignorados: <b>{ignorados}</b> | Erros: <b>{len(erros)}</b></p>
                        <p class="warn">Quando houver erro, corrija a linha indicada no Excel e importe novamente o arquivo.</p>
                    </div>
                    <p><a href="/excel/">Voltar para Excel</a> | <a href="/logs/excel">Ver log Excel</a></p>
                    <table>
                        <thead><tr><th>Linha</th><th>ID Erro</th><th>Como corrigir</th></tr></thead>
                        <tbody>{linhas_erro}</tbody>
                    </table>
                </div>
            </div>
        </body>
        </html>
        """
    )
