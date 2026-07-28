"""Módulo Atestados Médicos — registro e importação de atestados de colaboradores."""

from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import engine, get_db
from app.template_config import templates
from app.utils import flash_from_request, redirect_with_message

router = APIRouter(tags=["atestados"])

TIPOS_ATESTADO = ["Médico", "Odontológico", "Saúde Ocupacional", "Sáude", "Declaração de Horas", "Outro"]

# ─── tabela ──────────────────────────────────────────────────────────────────

def _garantir_tabela():
    with engine.connect() as conn:
        dialect = conn.dialect.name
        if dialect == "postgresql":
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS atestados_medicos (
                    id SERIAL PRIMARY KEY,
                    data_atestado VARCHAR(20) NOT NULL,
                    data_entrega VARCHAR(20) NOT NULL,
                    empresa VARCHAR(100) NOT NULL,
                    departamento VARCHAR(120) NOT NULL,
                    matricula VARCHAR(30),
                    nome_colaborador VARCHAR(200) NOT NULL,
                    cargo VARCHAR(150),
                    data_inicio_afastamento VARCHAR(20),
                    data_fim_afastamento VARCHAR(20),
                    cid VARCHAR(20),
                    medico_crm VARCHAR(200),
                    tipo_atestado VARCHAR(60) NOT NULL,
                    observacoes TEXT,
                    criado_em TIMESTAMP DEFAULT NOW()
                )
            """))
        else:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS atestados_medicos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    data_atestado VARCHAR(20) NOT NULL,
                    data_entrega VARCHAR(20) NOT NULL,
                    empresa VARCHAR(100) NOT NULL,
                    departamento VARCHAR(120) NOT NULL,
                    matricula VARCHAR(30),
                    nome_colaborador VARCHAR(200) NOT NULL,
                    cargo VARCHAR(150),
                    data_inicio_afastamento VARCHAR(20),
                    data_fim_afastamento VARCHAR(20),
                    cid VARCHAR(20),
                    medico_crm VARCHAR(200),
                    tipo_atestado VARCHAR(60) NOT NULL,
                    observacoes TEXT,
                    criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
        conn.commit()

_garantir_tabela()


# ─── helpers ─────────────────────────────────────────────────────────────────

def _fmt_data(v: str) -> str:
    """Converte yyyy-mm-dd (ou datetime string) para dd/mm/yyyy; mantém outros formatos."""
    s = str(v or "").strip().split(" ")[0]  # remove hora se houver
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
    return s


def _normalizar_data(v: str) -> str:
    """Normaliza datas de planilha (datetime Excel, ISO, dd/mm/yyyy) para dd/mm/yyyy."""
    s = str(v or "").strip()
    if not s or s == "-" or s.lower() == "none":
        return ""
    s = s.split(" ")[0]  # remove hora
    if len(s) == 10 and s[4] == "-":
        return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
    return s


def _is_duplicado(conn, data_atestado: str, nome: str, tipo: str) -> bool:
    r = conn.execute(text("""
        SELECT COUNT(*) FROM atestados_medicos
        WHERE data_atestado = :d AND nome_colaborador = :n AND tipo_atestado = :t
    """), {"d": data_atestado, "n": nome, "t": tipo}).scalar()
    return (r or 0) > 0


# ─── listagem ────────────────────────────────────────────────────────────────

@router.get("/folha/atestados")
def atestados_list(request: Request, q: str = "", db: Session = Depends(get_db)):
    base = """
        SELECT * FROM atestados_medicos
        WHERE 1=1
    """
    params: dict = {}
    if q:
        base += " AND (LOWER(nome_colaborador) LIKE :q OR LOWER(departamento) LIKE :q OR LOWER(empresa) LIKE :q OR LOWER(tipo_atestado) LIKE :q)"
        params["q"] = f"%{q.lower()}%"
    base += " ORDER BY id DESC LIMIT 200"

    items = db.execute(text(base), params).mappings().all()

    return templates.TemplateResponse("folha/atestados.html", {
        "request": request,
        "items": items,
        "q": q,
        "tipos_atestado": TIPOS_ATESTADO,
        **flash_from_request(request),
    })


# ─── formulário manual ───────────────────────────────────────────────────────

@router.post("/folha/atestados")
def atestados_salvar(
    request: Request,
    id_atestado: str = Form(""),
    data_atestado: str = Form(...),
    data_entrega: str = Form(...),
    empresa: str = Form(...),
    departamento: str = Form(...),
    matricula: str = Form(""),
    nome_colaborador: str = Form(...),
    cargo: str = Form(""),
    data_inicio_afastamento: str = Form(""),
    data_fim_afastamento: str = Form(""),
    cid: str = Form(""),
    medico_crm: str = Form(""),
    tipo_atestado: str = Form(...),
    observacoes: str = Form(""),
    db: Session = Depends(get_db),
):
    id_atestado = id_atestado.strip()

    if id_atestado:
        db.execute(text("""
            UPDATE atestados_medicos SET
                data_atestado=:da, data_entrega=:de, empresa=:emp, departamento=:dep,
                matricula=:mat, nome_colaborador=:nome, cargo=:cargo,
                data_inicio_afastamento=:di, data_fim_afastamento=:df,
                cid=:cid, medico_crm=:med, tipo_atestado=:tipo, observacoes=:obs
            WHERE id=:id
        """), {
            "da": data_atestado, "de": data_entrega, "emp": empresa, "dep": departamento,
            "mat": matricula or None, "nome": nome_colaborador, "cargo": cargo or None,
            "di": data_inicio_afastamento or None, "df": data_fim_afastamento or None,
            "cid": cid or None, "med": medico_crm or None, "tipo": tipo_atestado,
            "obs": observacoes or None, "id": int(id_atestado),
        })
        db.commit()
        return redirect_with_message("/folha/atestados", success="Atestado atualizado.")
    else:
        db.execute(text("""
            INSERT INTO atestados_medicos
                (data_atestado, data_entrega, empresa, departamento, matricula,
                 nome_colaborador, cargo, data_inicio_afastamento, data_fim_afastamento,
                 cid, medico_crm, tipo_atestado, observacoes)
            VALUES (:da,:de,:emp,:dep,:mat,:nome,:cargo,:di,:df,:cid,:med,:tipo,:obs)
        """), {
            "da": data_atestado, "de": data_entrega, "emp": empresa, "dep": departamento,
            "mat": matricula or None, "nome": nome_colaborador, "cargo": cargo or None,
            "di": data_inicio_afastamento or None, "df": data_fim_afastamento or None,
            "cid": cid or None, "med": medico_crm or None, "tipo": tipo_atestado,
            "obs": observacoes or None,
        })
        db.commit()
        return redirect_with_message("/folha/atestados", success="Atestado registrado.")


# ─── excluir ─────────────────────────────────────────────────────────────────

@router.post("/folha/atestados/{id}/excluir")
def atestados_excluir(id: int, db: Session = Depends(get_db)):
    db.execute(text("DELETE FROM atestados_medicos WHERE id=:id"), {"id": id})
    db.commit()
    return redirect_with_message("/folha/atestados", success="Atestado excluído.")


# ─── modelo Excel ─────────────────────────────────────────────────────────────

@router.get("/folha/atestados/modelo")
def atestados_modelo():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atestados"

    headers = [
        "Data do Atestado *", "Data de Entrega *", "Empresa *", "Departamento *",
        "Matrícula *", "Nome do Colaborador *", "Cargo",
        "Data Início Afastamento", "Data Fim Afastamento",
        "CID", "Nome do Médico / CRM", "Tipo de Atestado *", "Observações"
    ]

    header_fill = PatternFill("solid", fgColor="C00000")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # linha de exemplo
    exemplo = [
        "26/05/2026", "02/06/2026", "PIRES", "POWERTRAIN",
        "517", "LUCAS MATHEUS FERREIRA", "MECANICO COMPETIÇÃO JR",
        "", "", "", "Arthur Iera", "Odontológico",
        "Declaração de Horas 09:50 as 10:20"
    ]
    row_fill = PatternFill("solid", fgColor="FFF2CC")
    for col, val in enumerate(exemplo, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = row_fill
        cell.font = Font(name="Arial", size=9, italic=True)
        cell.alignment = Alignment(horizontal="center")

    # larguras
    widths = [15, 15, 12, 22, 10, 30, 22, 18, 16, 8, 22, 20, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo_atestados.xlsx"'},
    )


# ─── importar Excel ───────────────────────────────────────────────────────────

@router.post("/folha/atestados/importar")
async def atestados_importar(request: Request, arquivo: UploadFile = File(...), db: Session = Depends(get_db)):
    import openpyxl

    conteudo = await arquivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active

    # detecta linha de cabeçalho
    header_row = 1
    for r in ws.iter_rows(min_row=1, max_row=10):
        vals = [str(c.value or "").strip().lower() for c in r]
        if any("atestado" in v or "colaborador" in v or "entrega" in v for v in vals):
            header_row = r[0].row
            break

    headers = [str(c.value or "").strip().lower() for c in ws[header_row]]

    def col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return None

    idx = {
        "data_atestado":          col("atestado"),
        "data_entrega":           col("entrega"),
        "empresa":                col("empresa"),
        "departamento":           col("departamento", "depto"),
        "matricula":              col("matrícula", "matricula"),
        "nome_colaborador":       col("colaborador", "nome"),
        "cargo":                  col("cargo"),
        "data_inicio_afastamento":col("início afastamento", "inicio afastamento", "inicio afas"),
        "data_fim_afastamento":   col("fim afastamento", "data fim"),
        "cid":                    col("cid"),
        "medico_crm":             col("médico", "medico", "crm"),
        "tipo_atestado":          col("tipo"),
        "observacoes":            col("observa"),
    }

    CAMPOS_OBRIGATORIOS = ["data_atestado", "data_entrega", "empresa", "departamento", "nome_colaborador", "tipo_atestado"]
    TIPOS_VALIDOS = {t.lower() for t in TIPOS_ATESTADO}

    importados = 0
    duplicados = 0
    erros = []

    with engine.connect() as conn:
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if all(v is None for v in row):
                continue

            def cel(key):
                i = idx.get(key)
                if i is None or i >= len(row):
                    return ""
                v = row[i]
                if v is None:
                    return ""
                s = str(v).strip()
                return "" if s in ("-", "None", "none") else s

            data_atestado    = _normalizar_data(cel("data_atestado"))
            data_entrega     = _normalizar_data(cel("data_entrega"))
            empresa          = cel("empresa")
            departamento     = cel("departamento")
            matricula        = cel("matricula")
            nome_colaborador = cel("nome_colaborador")
            cargo            = cel("cargo")
            data_inicio      = _normalizar_data(cel("data_inicio_afastamento"))
            data_fim         = _normalizar_data(cel("data_fim_afastamento"))
            cid              = cel("cid")
            medico           = cel("medico_crm")
            tipo_atestado    = cel("tipo_atestado")
            obs              = cel("observacoes")

            # validação de obrigatórios
            lv = locals()
            faltando = [f for f in CAMPOS_OBRIGATORIOS if not lv.get(f, "").strip()]
            if faltando:
                erros.append(f"Linha {row_num}: campos faltando — {', '.join(faltando)}")
                continue

            # normaliza tipo
            tipo_norm = next((t for t in TIPOS_ATESTADO if t.lower() == tipo_atestado.lower()), tipo_atestado)

            # checagem de duplicidade
            if _is_duplicado(conn, data_atestado, nome_colaborador, tipo_norm):
                duplicados += 1
                continue

            conn.execute(text("""
                INSERT INTO atestados_medicos
                    (data_atestado, data_entrega, empresa, departamento, matricula,
                     nome_colaborador, cargo, data_inicio_afastamento, data_fim_afastamento,
                     cid, medico_crm, tipo_atestado, observacoes)
                VALUES (:da,:de,:emp,:dep,:mat,:nome,:cargo,:di,:df,:cid,:med,:tipo,:obs)
            """), {
                "da": data_atestado, "de": data_entrega, "emp": empresa, "dep": departamento,
                "mat": matricula or None, "nome": nome_colaborador, "cargo": cargo or None,
                "di": data_inicio or None, "df": data_fim or None,
                "cid": cid or None, "med": medico or None, "tipo": tipo_norm,
                "obs": obs or None,
            })
            importados += 1

        conn.commit()

    msg = f"{importados} atestado(s) importado(s)."
    if duplicados:
        msg += f" {duplicados} duplicado(s) ignorado(s)."
    if erros:
        msg += f" {len(erros)} linha(s) com erro: {erros[0]}"
        if len(erros) > 1:
            msg += f" (e mais {len(erros)-1})"

    return redirect_with_message("/folha/atestados", success=msg if importados or duplicados else None,
                                  error=msg if not importados and erros else None)
