from datetime import datetime
from pathlib import Path
import hashlib
import json
import re
import unicodedata

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import RedirectResponse
from openpyxl import load_workbook
from sqlalchemy import (
    Column,
    DateTime,
    Float,
    Integer,
    MetaData,
    String,
    Table,
    Text,
    and_,
    func,
    insert,
    select,
    text,
    update,
)
from sqlalchemy.orm import Session

from app.auth import is_admin as _is_admin
from app.database import BASE_DIR, engine, get_db
from app.models import DimAutonomo, DimEtapa, DimPiloto
from app.template_config import templates
from app.utils import flash_from_request, redirect_with_message

router = APIRouter(tags=["pesquisas"])

metadata_pesquisas = MetaData()

pesquisa_uploads = Table(
    "pesquisa_uploads",
    metadata_pesquisas,
    Column("id_upload", Integer, primary_key=True, autoincrement=True),
    Column("id_etapa", Integer, nullable=False),
    Column("tipo_pesquisa", String(80), nullable=False),
    Column("arquivo_nome", String(255), nullable=False),
    Column("abas", Text),
    Column("qtd_linhas", Integer, default=0),
    Column("qtd_respostas", Integer, default=0),
    Column("status", String(40), default="Importado"),
    Column("hash_conteudo", String(40)),
    Column("criado_em", DateTime, default=datetime.utcnow),
)

pesquisa_respostas = Table(
    "pesquisa_respostas",
    metadata_pesquisas,
    Column("id_resposta", Integer, primary_key=True, autoincrement=True),
    Column("id_upload", Integer, nullable=False),
    Column("id_etapa", Integer, nullable=False),
    Column("tipo_pesquisa", String(80), nullable=False),
    Column("aba", String(160)),
    Column("linha_excel", Integer),
    Column("carimbo_data_hora", String(80)),
    Column("email", String(180)),
    Column("respondente_nome", String(180)),
    Column("respondente_chave", String(220)),
    Column("alvo_nome", String(180)),
    Column("alvo_chave", String(220)),
    Column("tipo_alvo", String(60)),
    Column("id_autonomo", Integer),
    Column("id_piloto", Integer),
    Column("categoria_forms", String(180)),
    Column("funcao_forms", String(180)),
    Column("lider_forms", String(180)),
    Column("periodo_forms", String(120)),
    Column("grupo_pergunta", String(120)),
    Column("subgrupo_pergunta", String(160)),
    Column("pergunta_original", Text),
    Column("resposta_original", Text),
    Column("nota_num", Float),
    Column("resposta_padronizada", String(120)),
    Column("manter_trocar", String(80)),
    Column("comentario", Text),
    Column("status_mapeamento", String(40), default="Pendente"),
    Column("criado_em", DateTime, default=datetime.utcnow),
)

pesquisa_mapeamentos = Table(
    "pesquisa_mapeamentos",
    metadata_pesquisas,
    Column("id_mapeamento", Integer, primary_key=True, autoincrement=True),
    Column("tipo_alvo", String(60), nullable=False),
    Column("texto_origem", String(220), nullable=False),
    Column("texto_chave", String(220), nullable=False),
    Column("id_autonomo", Integer),
    Column("id_piloto", Integer),
    Column("criado_em", DateTime, default=datetime.utcnow),
)

metadata_pesquisas.create_all(engine)


def _migrar_pesquisas():
    """Colunas novas em bancos criados antes desta versão."""
    novas = {
        "pesquisa_respostas": {"subgrupo_pergunta": "VARCHAR(160)"},
        "pesquisa_uploads": {"hash_conteudo": "VARCHAR(40)"},
    }
    with engine.connect() as conn:
        dialect = conn.dialect.name
        for tabela, cols in novas.items():
            for col, tipo in cols.items():
                try:
                    if dialect == "postgresql":
                        conn.execute(text(
                            f"ALTER TABLE {tabela} ADD COLUMN IF NOT EXISTS {col} {tipo}"))
                        conn.commit()
                    else:
                        try:
                            conn.execute(text(f"SELECT {col} FROM {tabela} LIMIT 1"))
                        except Exception:
                            conn.rollback()
                            conn.execute(text(f"ALTER TABLE {tabela} ADD COLUMN {col} {tipo}"))
                            conn.commit()
                except Exception as exc:
                    print(f"AVISO - migração {tabela}.{col}: {exc}")


_migrar_pesquisas()

UPLOAD_DIR = BASE_DIR / "data" / "uploads_pesquisas"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

TIPOS_PESQUISA = [
    ("feedback_autonomo", "Avaliação do Autônomo (líder avalia)"),
    ("satisfacao_autonomo", "Satisfação do Autônomo (autônomo responde)"),
    ("feedback_equipe_tecnica", "Feedback Equipe Técnica (ocorrências)"),
    ("feedback_piloto", "Feedback do Piloto"),
]

# rótulo curto, para caber nas tabelas
TIPOS_CURTO = {
    "feedback_autonomo": "Avaliação Autônomo",
    "satisfacao_autonomo": "Satisfação Autônomo",
    "feedback_equipe_tecnica": "Equipe Técnica",
    "feedback_piloto": "Piloto",
}

# Escala qualitativa → nota, para permitir média junto com as escalas numéricas
ESCALA_QUALITATIVA = {
    "excelente": 5.0,
    "otimo": 5.0,
    "muito bom": 4.5,
    "bom": 4.0,
    "boa": 4.0,
    "satisfatorio": 3.5,
    "regular": 3.0,
    "ruim": 2.0,
    "pessimo": 1.0,
    "muito ruim": 1.0,
}


def detectar_tipo(nome_arquivo: str) -> str | None:
    """Descobre o tipo de pesquisa pelo nome do arquivo."""
    k = normalizar(nome_arquivo)
    if "freelancer" in k or "equipe tecnica" in k:
        return "feedback_equipe_tecnica"
    if "satisfacao" in k and "autonomo" in k:
        return "satisfacao_autonomo"
    if "avaliacao de autonomos" in k or ("avaliacao" in k and "autonomo" in k):
        return "feedback_autonomo"
    if "pos etapa" in k or "piloto" in k:
        return "feedback_piloto"
    return None


def detectar_tipo_por_conteudo(wb) -> str | None:
    """Descobre o tipo pelas colunas, quando o nome do arquivo não diz nada.

    O Windows manda o nome curto (PORSCH~1.XLS) quando o caminho é longo, e aí
    não há o que reconhecer no nome.
    """
    cabecalhos = []
    for ws in wb.worksheets:
        for linha in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            cabecalhos += [normalizar(c) for c in linha if c]
            break
    texto = " | ".join(cabecalhos)
    if not texto:
        return None

    if "tipo de feedback" in texto or "descreva o ocorrido" in texto:
        return "feedback_equipe_tecnica"
    if "manter ou trocar o seu engenheiro" in texto or "selecione sua categoria" in texto:
        return "feedback_piloto"
    if "engenheiro" in texto and "mecanico" in texto and "seu nome" in texto:
        return "feedback_piloto"
    if "funcao desempenhou nesta etapa" in texto or "avalie o evento nos seguintes aspectos" in texto:
        return "satisfacao_autonomo"
    if "nome do autonomo" in texto and "lider" in texto:
        return "feedback_autonomo"
    return None


def detectar_etapa_por_conteudo(wb, etapas: list):
    """Etapa citada no conteúdo da planilha (título ou coluna ETAPA)."""
    for ws in wb.worksheets:
        achou = detectar_etapas(ws.title, etapas)
        if achou:
            return achou
    return []


def detectar_etapas(nome_arquivo: str, etapas: list) -> list:
    """IDs de etapa citados no nome do arquivo.

    Aceita '26ET4' e 'Etapa 4'. Um arquivo pode cobrir mais de uma etapa
    (ex.: '26ET5_ 26ET6 - Portugal').
    """
    k = normalizar(nome_arquivo)
    numeros: list[int] = []

    for m in re.finditer(r"\b\d{2}et(\d{1,2})\b", k):
        numeros.append(int(m.group(1)))
    if not numeros:
        for m in re.finditer(r"\betapa (\d{1,2})\b", k):
            numeros.append(int(m.group(1)))

    achados = []
    for num in dict.fromkeys(numeros):
        for e in etapas:
            nome = normalizar(e.nome_etapa)
            if re.search(rf"\b\d{{2}}et0*{num}\b", nome) or re.search(rf"\betapa 0*{num}\b", nome):
                if e.id_etapa not in achados:
                    achados.append(e.id_etapa)
                break
    return achados


def normalizar(txt):
    if txt is None:
        return ""
    txt = str(txt).strip().lower()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = re.sub(r"[^a-z0-9]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    # "e-mail" vira "e mail" e nunca casava com o termo "email"
    return txt.replace("e mail", "email")


def is_empty(v):
    return v is None or str(v).strip() == ""


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def resposta_padrao(v):
    if v is None:
        return None

    s = str(v).strip()
    k = normalizar(s)

    if k in {"sim", "s"}:
        return "Sim"
    if k in {"nao", "n"}:
        return "Não"
    if "nao se aplica" in k or "n a" in k:
        return "Não se aplica"
    if "excelente" in k or k == "otimo":
        return "Excelente"
    if k in {"bom", "boa"} or "muito bom" in k:
        return "Bom"
    if "regular" in k or "satisfatorio" in k:
        return "Regular"
    if "pessimo" in k or "muito ruim" in k:
        return "Péssimo"
    if "ruim" in k:
        return "Ruim"
    if "manter" in k:
        return "Quero manter"
    if "trocar" in k:
        return "Quero trocar"

    return s


def nota_da_escala(v):
    """Converte a escala qualitativa em nota, para poder tirar média."""
    k = normalizar(v)
    if not k:
        return None
    if k in ESCALA_QUALITATIVA:
        return ESCALA_QUALITATIVA[k]
    for termo, nota in ESCALA_QUALITATIVA.items():
        if termo in k:
            return nota
    return None


def _pergunta_de_permanencia(coluna) -> bool:
    """A coluna pergunta se mantém ou troca a pessoa?"""
    k = normalizar(coluna)
    if "manter" in k or "trocar" in k:
        return True
    return "deseja" in k and "proxima" in k


def manter_trocar(coluna, v):
    """Só classifica quando a pergunta é sobre permanência.

    Antes olhava só a resposta, então qualquer 'SIM' do formulário da equipe
    técnica ("respeitou os horários?") virava 'Manter'.
    """
    if not _pergunta_de_permanencia(coluna):
        return None
    k = normalizar(v)
    if "manter" in k or k == "sim":
        return "Manter"
    if "trocar" in k or k == "nao":
        return "Trocar"
    return None


def grupo_pergunta(coluna):
    k = normalizar(coluna)

    # comentários primeiro: "comentários sobre o seu engenheiro" é comentário,
    # não avaliação do engenheiro
    if "coment" in k or "observ" in k or "sugest" in k or "critica" in k or "opiniao" in k:
        return "Comentário"
    if "descreva o ocorrido" in k:
        return "Comentário"

    # avaliação de pessoas
    if "engenheiro" in k:
        return "Engenheiro"
    if "mecanico" in k:
        return "Mecânico"
    if "coach" in k:
        return "Coach"
    if "analise de dados" in k or "analista" in k:
        return "Análise de Dados"

    # satisfação do autônomo
    if "prestado pelo rh" in k or k.startswith("avalie a atuacao e o suporte prestado pelo rh"):
        return "RH"
    if "apoiadores" in k or "supervisores" in k:
        return "Supervisão"
    if "avalie a equipe" in k:
        return "Equipe"

    # ocorrências da equipe técnica
    if "tipo de ocorrencia" in k or "tipo de feedback" in k:
        return "Ocorrência"
    if "horarios" in k or "asseio" in k or "uniforme" in k or "postura" in k:
        return "Conduta"
    if "qualidade esperada" in k:
        return "Entrega"
    if "ciente da ocorrencia" in k or "reacao do autonomo" in k:
        return "Ocorrência"
    if _pergunta_de_permanencia(coluna):
        return "Permanência"

    # estrutura do evento
    if "relacionamento com pilotos" in k:
        return "Relacionamento com Pilotos"
    if "diretor de provas" in k:
        return "Diretor de Provas"
    if "comissarios" in k or "cba" in k:
        return "Comissários"
    if "buffet" in k:
        return "Buffet"
    if "alimentacao" in k:
        # refeição da equipe, não é o buffet dos pilotos
        return "Alimentação"
    if "aplicativo" in k:
        return "Aplicativo"
    if "carro" in k:
        return "Carro"
    if "evento" in k or "credenciamento" in k or "estacionamento" in k or "desmontagem" in k:
        return "Evento"

    return "Geral"


def subgrupo_pergunta(coluna):
    """Item entre colchetes do Google Forms: 'Engenheiro [Comunicação]' → 'Comunicação'."""
    m = re.search(r"\[(.+?)\]", str(coluna or ""))
    if m:
        return m.group(1).strip().rstrip(";").strip() or None
    return None


def achar_coluna(colunas, termos):
    for c in colunas:
        k = normalizar(c)
        if all(t in k for t in termos):
            return c
    return None


def detectar_contexto(tipo_pesquisa, row):
    colunas = list(row.keys())

    col_email = achar_coluna(colunas, ["endereco", "email"]) or achar_coluna(colunas, ["email"])
    col_data = achar_coluna(colunas, ["carimbo"])

    email = row.get(col_email) if col_email else None
    carimbo = row.get(col_data) if col_data else None

    respondente_nome = None
    alvo_nome = None
    tipo_alvo = None

    categoria_forms = None
    funcao_forms = None
    lider_forms = None
    periodo_forms = None

    # "categoria" solto casava com "...supervisores da categoria...", que é uma
    # pergunta de nota; só vale como coluna própria de categoria.
    col_categoria = achar_coluna(colunas, ["selecione", "categoria"])
    if not col_categoria:
        col_categoria = next((c for c in colunas if normalizar(c) == "categoria"), None)
    col_nome_autonomo = achar_coluna(colunas, ["nome", "autonomo"])
    col_funcao = achar_coluna(colunas, ["funcao", "autonomo"])
    col_lider = achar_coluna(colunas, ["lider"])
    col_periodo = achar_coluna(colunas, ["periodo", "trabalhou"])
    col_respondente = achar_coluna(colunas, ["seu", "nome"]) or achar_coluna(colunas, ["nome"])

    if col_categoria:
        categoria_forms = row.get(col_categoria)

    if col_funcao:
        funcao_forms = row.get(col_funcao)

    if col_lider:
        lider_forms = row.get(col_lider)

    if col_periodo:
        periodo_forms = row.get(col_periodo)

    if tipo_pesquisa in ("feedback_autonomo", "feedback_equipe_tecnica"):
        # o líder responde sobre um autônomo
        alvo_nome = row.get(col_nome_autonomo) if col_nome_autonomo else None
        respondente_nome = row.get(col_lider) if col_lider else None
        tipo_alvo = "autonomo"

    elif tipo_pesquisa == "satisfacao_autonomo":
        # o próprio autônomo responde sobre o evento; pode se identificar ou não
        col_identifica = (achar_coluna(colunas, ["identificar"])
                          or achar_coluna(colunas, ["nome", "completo"]))
        respondente_nome = row.get(col_identifica) if col_identifica else None
        alvo_nome = respondente_nome
        tipo_alvo = "autonomo"

    else:  # feedback_piloto
        respondente_nome = row.get(col_respondente) if col_respondente else None
        alvo_nome = respondente_nome
        tipo_alvo = "piloto"

    return {
        "email": str(email).strip() if not is_empty(email) else None,
        "carimbo": str(carimbo).strip() if not is_empty(carimbo) else None,
        "respondente_nome": str(respondente_nome).strip() if not is_empty(respondente_nome) else None,
        "alvo_nome": str(alvo_nome).strip() if not is_empty(alvo_nome) else None,
        "tipo_alvo": tipo_alvo,
        "categoria_forms": str(categoria_forms).strip() if not is_empty(categoria_forms) else None,
        "funcao_forms": str(funcao_forms).strip() if not is_empty(funcao_forms) else None,
        "lider_forms": str(lider_forms).strip() if not is_empty(lider_forms) else None,
        "periodo_forms": str(periodo_forms).strip() if not is_empty(periodo_forms) else None,
    }


def colunas_metadata(colunas):
    ignorar = set()

    for termos in [
        ["carimbo"],
        ["endereco", "email"],
        ["email"],
        ["seu", "nome"],
        ["nome", "autonomo"],
        ["funcao", "autonomo"],
        ["lider"],
        ["periodo", "trabalhou"],
        ["selecione", "categoria"],
        # satisfação do autônomo
        ["identificar"],
        ["nome", "completo"],
        ["funcao", "desempenhou"],
        # equipe técnica
        ["piloto", "atendido"],
        ["data", "ocorrido"],
        ["id", "feedback"],
    ]:
        c = achar_coluna(colunas, termos)
        if c:
            ignorar.add(c)

    # "etapa" e "categoria" precisam ser a coluna inteira: por trecho pegavam
    # perguntas como "...supervisores da categoria..." e "Feedback Pós Etapa",
    # que eram descartadas em vez de importadas.
    for exato in ("etapa", "categoria"):
        ignorar.update(c for c in colunas if normalizar(c) == exato)

    return ignorar


def aplicar_mapeamento(db: Session, tipo_alvo, texto_origem):
    chave = normalizar(texto_origem)

    if not chave:
        return None, None, "Pendente"

    m = db.execute(
        select(pesquisa_mapeamentos).where(
            and_(
                pesquisa_mapeamentos.c.tipo_alvo == tipo_alvo,
                pesquisa_mapeamentos.c.texto_chave == chave,
            )
        )
    ).mappings().first()

    if m:
        return m.get("id_autonomo"), m.get("id_piloto"), "Mapeado"

    if tipo_alvo == "autonomo":
        candidatos = db.query(DimAutonomo).all()
        for a in candidatos:
            if normalizar(a.nome_autonomo) == chave:
                return a.id_autonomo, None, "Mapeado"

    if tipo_alvo == "piloto":
        candidatos = db.query(DimPiloto).all()
        for p in candidatos:
            if normalizar(p.nome_piloto) == chave:
                return None, p.id_piloto, "Mapeado"

    return None, None, "Pendente"


def options(db: Session):
    return {
        "etapas": db.query(DimEtapa).order_by(DimEtapa.temporada.desc(), DimEtapa.data_inicio.desc()).all(),
        "autonomos": db.query(DimAutonomo).order_by(DimAutonomo.nome_autonomo).all(),
        "pilotos": db.query(DimPiloto).order_by(DimPiloto.nome_piloto).all(),
        "tipos_pesquisa": TIPOS_PESQUISA,
        "tipos_label": dict(TIPOS_PESQUISA),
        "tipos_curto": TIPOS_CURTO,
    }


def _filtro_base(request):
    """WHERE + parâmetros da base consolidada, a partir da querystring."""
    f_etapa = request.query_params.get("f_etapa", "")
    f_tipo = request.query_params.get("f_tipo", "")
    f_grupo = request.query_params.get("f_grupo", "")
    busca = (request.query_params.get("busca", "") or "").strip()

    where, params = ["1=1"], {}
    if f_etapa:
        where.append("p.id_etapa = :et")
        params["et"] = int(f_etapa)
    if f_tipo:
        where.append("p.tipo_pesquisa = :tp")
        params["tp"] = f_tipo
    if f_grupo:
        where.append("p.grupo_pergunta = :gr")
        params["gr"] = f_grupo
    if busca:
        where.append(
            "(LOWER(COALESCE(p.respondente_nome,'')) LIKE :bu"
            " OR LOWER(COALESCE(p.alvo_nome,'')) LIKE :bu"
            " OR LOWER(COALESCE(p.subgrupo_pergunta,'')) LIKE :bu"
            " OR LOWER(COALESCE(p.pergunta_original,'')) LIKE :bu"
            " OR LOWER(COALESCE(p.resposta_original,'')) LIKE :bu)"
        )
        params["bu"] = f"%{busca.lower()}%"

    return " AND ".join(where), params, {
        "f_etapa": f_etapa, "f_tipo": f_tipo, "f_grupo": f_grupo, "busca": busca,
    }


@router.get("/pesquisas")
def pesquisas_home(request: Request, db: Session = Depends(get_db)):
    uploads = db.execute(
        select(pesquisa_uploads).order_by(pesquisa_uploads.c.id_upload.desc()).limit(50)
    ).mappings().all()

    # ── base consolidada de todas as pesquisas ──
    try:
        pagina = max(1, int(request.query_params.get("pagina", "1")))
    except ValueError:
        pagina = 1
    por_pagina = 100

    w, params, filtros = _filtro_base(request)

    total_base = db.execute(
        text(f"SELECT COUNT(*) FROM pesquisa_respostas p WHERE {w}"), params
    ).scalar() or 0

    base = db.execute(text(f"""
        SELECT COALESCE(e.nome_etapa, '—') AS etapa,
               p.tipo_pesquisa, p.respondente_nome, p.alvo_nome,
               p.categoria_forms, p.funcao_forms,
               p.grupo_pergunta, p.subgrupo_pergunta, p.pergunta_original,
               p.resposta_original, p.nota_num, p.resposta_padronizada,
               p.manter_trocar
        FROM pesquisa_respostas p
        LEFT JOIN dim_etapas e ON e.id_etapa = p.id_etapa
        WHERE {w}
        ORDER BY p.id_resposta
        LIMIT :lim OFFSET :off
    """), {**params, "lim": por_pagina, "off": (pagina - 1) * por_pagina}).mappings().all()

    grupos = [r[0] for r in db.execute(text(
        "SELECT DISTINCT grupo_pergunta FROM pesquisa_respostas "
        "WHERE grupo_pergunta IS NOT NULL ORDER BY 1"
    )).fetchall()]

    return templates.TemplateResponse(
        "pesquisas/index.html",
        {
            "request": request,
            "uploads": uploads,
            "base": base,
            "base_total": total_base,
            "base_grupos": grupos,
            "ultima": uploads[0] if uploads else None,
            "pagina": pagina,
            "por_pagina": por_pagina,
            **filtros,
            **options(db),
            **flash_from_request(request),
        },
    )


@router.get("/pesquisas/exportar")
def pesquisas_exportar(request: Request, db: Session = Depends(get_db)):
    """Exporta a base consolidada, respeitando os filtros da tela."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse

    w, params, _ = _filtro_base(request)

    linhas = db.execute(text(f"""
        SELECT COALESCE(e.nome_etapa, '') AS etapa,
               p.tipo_pesquisa, p.categoria_forms, p.funcao_forms,
               p.respondente_nome, p.alvo_nome,
               p.grupo_pergunta, p.subgrupo_pergunta, p.pergunta_original,
               p.resposta_original, p.resposta_padronizada, p.nota_num,
               p.manter_trocar, p.carimbo_data_hora
        FROM pesquisa_respostas p
        LEFT JOIN dim_etapas e ON e.id_etapa = p.id_etapa
        WHERE {w}
        ORDER BY p.id_resposta
    """), params).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pesquisas"

    cabecalhos = ["Etapa", "Pesquisa", "Categoria", "Função", "Respondente", "Avaliado",
                  "Grupo", "Item", "Pergunta original", "Resposta original",
                  "Resposta padronizada", "Nota", "Permanência", "Data/hora"]
    ws.append(cabecalhos)
    for c in range(1, len(cabecalhos) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="1E293B")

    for r in linhas:
        ws.append([
            r["etapa"], TIPOS_CURTO.get(r["tipo_pesquisa"], r["tipo_pesquisa"]),
            r["categoria_forms"], r["funcao_forms"], r["respondente_nome"], r["alvo_nome"],
            r["grupo_pergunta"], r["subgrupo_pergunta"], r["pergunta_original"],
            r["resposta_original"], r["resposta_padronizada"], r["nota_num"],
            r["manter_trocar"], r["carimbo_data_hora"],
        ])

    for i, largura in enumerate([22, 20, 16, 20, 26, 26, 18, 24, 46, 40, 22, 8, 14, 20], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = largura
    ws.freeze_panes = "A2"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"pesquisas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )


@router.post("/pesquisas/{id_upload}/etapa")
def pesquisas_definir_etapa(
    id_upload: int,
    request: Request,
    id_etapa: int = Form(...),
    db: Session = Depends(get_db),
):
    """Ajusta a etapa de um upload — e das respostas que vieram dele."""
    if not _is_admin(request):
        return RedirectResponse("/?sem_acesso=pesquisas", status_code=303)

    db.execute(
        update(pesquisa_uploads)
        .where(pesquisa_uploads.c.id_upload == id_upload)
        .values(id_etapa=id_etapa)
    )
    db.execute(
        update(pesquisa_respostas)
        .where(pesquisa_respostas.c.id_upload == id_upload)
        .values(id_etapa=id_etapa)
    )
    db.commit()
    return redirect_with_message("/pesquisas", success="Etapa atualizada.")


@router.post("/pesquisas/upload")
async def pesquisas_upload(
    request: Request,
    arquivos: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    """Importa uma ou várias planilhas, deduzindo etapa e tipo pelo nome."""
    if not _is_admin(request):
        return RedirectResponse("/?sem_acesso=pesquisas", status_code=303)

    arquivos = [a for a in (arquivos or []) if a and a.filename]
    if not arquivos:
        return redirect_with_message(
            "/pesquisas", error="Escolha ao menos uma planilha .xlsx antes de validar.")

    etapas = db.query(DimEtapa).order_by(DimEtapa.temporada.desc(), DimEtapa.data_inicio).all()

    analises = []
    vistos: dict = {}
    for arquivo in arquivos:
        conteudo = await arquivo.read()
        analises.append(analisar_arquivo(db, arquivo.filename, conteudo, etapas, vistos))

    resumo = {
        "arquivos": len(analises),
        "ok": sum(1 for a in analises if a["status"] == "OK"),
        "respostas": sum(a["respostas"] for a in analises),
        "linhas": sum(a["linhas"] for a in analises),
        "tipos": len({a["tipo"] for a in analises if a["tipo"]}),
        "sem_etapa": sum(1 for a in analises if a["status"] == "Etapa indefinida"),
        "duplicados": sum(1 for a in analises if a["status"] == "Duplicado"),
        "erros": sum(1 for a in analises if a["status"] == "Erro"),
        "avisos": sum(len(a["avisos"]) for a in analises),
    }

    return templates.TemplateResponse(
        "pesquisas/revisao.html",
        {
            "request": request,
            "analises": analises,
            "resumo": resumo,
            **options(db),
        },
    )


@router.post("/pesquisas/confirmar")
async def pesquisas_confirmar(request: Request, db: Session = Depends(get_db)):
    """Passo 4: grava o que foi revisado na prévia."""
    if not _is_admin(request):
        return RedirectResponse("/?sem_acesso=pesquisas", status_code=303)

    form = await request.form()
    etapas = db.query(DimEtapa).all()

    indices = sorted({k.split("_")[-1] for k in form.keys() if k.startswith("arq_")})
    importados, ignorados = [], []

    for i in indices:
        if not form.get(f"incluir_{i}"):
            continue
        nome = form.get(f"nome_{i}", "")
        disco = form.get(f"arq_{i}", "")
        tipo = form.get(f"tipo_{i}", "")
        try:
            id_etapa = int(form.get(f"etapa_{i}") or 0)
        except ValueError:
            id_etapa = 0

        ok, msg = gravar_arquivo(db, nome, disco, tipo, id_etapa, etapas)
        (importados if ok else ignorados).append(msg)

    db.commit()

    partes = []
    if importados:
        partes.append(f"{len(importados)} arquivo(s) importado(s): " + " | ".join(importados))
    if ignorados:
        partes.append("Não importados — " + " | ".join(ignorados))

    if importados:
        return redirect_with_message("/pesquisas", success=" || ".join(partes))
    return redirect_with_message(
        "/pesquisas",
        error=" || ".join(partes) or
        "Nenhum arquivo foi marcado na revisão — nada foi importado. "
        "Envie as planilhas de novo e confira as caixas de seleção.")


def _mapa_etapas(etapas):
    mapa = {}
    for e in etapas:
        n = normalizar(e.nome_etapa)
        m = re.search(r"\b\d{2}et(\d{1,2})\b", n) or re.search(r"\betapa (\d{1,2})\b", n)
        if m:
            mapa[int(m.group(1))] = e.id_etapa
    return mapa


def percorrer_planilha(wb, tipo_pesquisa, id_etapa_arquivo, etapas):
    """Percorre a planilha e devolve as respostas ja padronizadas.

    Usado tanto pela previa quanto pela importacao, para o que o usuario ve
    na revisao ser exatamente o que sera gravado.

    Devolve (respostas, total_linhas, avisos).
    """
    mapa = _mapa_etapas(etapas)

    def etapa_da_linha(valor):
        m = re.search(r"(\d{1,2})", str(valor or ""))
        return mapa.get(int(m.group(1))) if m else None

    id_etapa = id_etapa_arquivo or 0
    respostas: list[dict] = []
    total_linhas = 0
    avisos: list[str] = []
    abas_vistas: set = set()
    linhas_incompletas = 0

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        headers = [h if h else f"coluna_{i+1}" for i, h in enumerate(headers)]

        # O export do Forms traz abas derivadas ("Comentarios") com as mesmas
        # colunas da aba principal; importar as duas duplicava toda resposta.
        assinatura = tuple(normalizar(h) for h in headers)
        if assinatura in abas_vistas:
            avisos.append(f"aba '{ws.title}' ignorada: repete as colunas de outra aba")
            continue
        abas_vistas.add(assinatura)

        ignorar = colunas_metadata(headers)
        col_etapa_planilha = achar_coluna(headers, ["etapa"])

        for idx, values in enumerate(rows[1:], start=2):
            row = {headers[i]: values[i] if i < len(values) else None
                   for i in range(len(headers))}
            if all(is_empty(v) for v in row.values()):
                continue

            total_linhas += 1
            contexto = detectar_contexto(tipo_pesquisa, row)

            # No Forms dos pilotos a categoria da aba TROPHY esta no nome da
            # aba, nao em coluna; sem isso essas respostas ficavam sem categoria.
            if not contexto["categoria_forms"]:
                aba_k = normalizar(ws.title)
                if aba_k and "resposta" not in aba_k and "coment" not in aba_k:
                    contexto["categoria_forms"] = ws.title.strip().title()

            id_etapa_linha = id_etapa
            if col_etapa_planilha:
                da_linha = etapa_da_linha(row.get(col_etapa_planilha))
                if da_linha:
                    id_etapa_linha = da_linha

            if not id_etapa_linha:
                linhas_incompletas += 1

            respondidas = 0
            for col, val in row.items():
                if col in ignorar or is_empty(val):
                    continue

                resposta_texto = str(val).strip()
                grupo = grupo_pergunta(col)
                nota = to_float(val)
                if nota is None:
                    nota = nota_da_escala(val)

                respondidas += 1
                respostas.append({
                    "id_etapa": id_etapa_linha,
                    "tipo_pesquisa": tipo_pesquisa,
                    "aba": ws.title,
                    "linha_excel": idx,
                    "carimbo_data_hora": contexto["carimbo"],
                    "email": contexto["email"],
                    "respondente_nome": contexto["respondente_nome"],
                    "respondente_chave": normalizar(contexto["respondente_nome"]),
                    "alvo_nome": contexto["alvo_nome"],
                    "alvo_chave": normalizar(contexto["alvo_nome"]),
                    "tipo_alvo": contexto["tipo_alvo"],
                    "categoria_forms": contexto["categoria_forms"],
                    "funcao_forms": contexto["funcao_forms"],
                    "lider_forms": contexto["lider_forms"],
                    "periodo_forms": contexto["periodo_forms"],
                    "grupo_pergunta": grupo,
                    "subgrupo_pergunta": subgrupo_pergunta(col),
                    "pergunta_original": col,
                    "resposta_original": resposta_texto,
                    "nota_num": nota,
                    "resposta_padronizada": resposta_padrao(resposta_texto),
                    "manter_trocar": manter_trocar(col, resposta_texto),
                    "comentario": resposta_texto if grupo == "Comentario" or grupo == "Comentário" else None,
                })

            if not respondidas:
                linhas_incompletas += 1

    if linhas_incompletas:
        avisos.append(f"{linhas_incompletas} linha(s) sem etapa definida ou sem resposta")

    return respostas, total_linhas, avisos


def analisar_arquivo(db, nome_arquivo, conteudo, etapas, vistos=None):
    """Le o arquivo sem gravar nada e devolve o diagnostico da previa.

    `vistos` acumula os hashes do lote: o mesmo arquivo costuma vir repetido
    em varias pastas de etapa, e sem isso a previa prometia respostas que a
    gravacao ia recusar.
    """
    info = {
        "arquivo": nome_arquivo,
        "tipo": None,
        "tipo_label": None,
        "id_etapa": 0,
        "etapa_label": None,
        "linhas": 0,
        "respostas": 0,
        "avisos": [],
        "status": "OK",
        "digest": None,
        "arquivo_disco": None,
    }

    # O Windows manda o nome curto (PORSCH~1.XLS) quando o caminho é longo:
    # a extensão engana e o nome não identifica nada. O que vale é o conteúdo.
    digest = hashlib.sha1(conteudo).hexdigest()
    info["digest"] = digest

    if vistos is not None and digest in vistos:
        info["status"] = "Duplicado"
        info["avisos"].append(
            f"conteúdo idêntico a '{vistos[digest]}', que já está neste envio")
        return info

    ja = db.execute(
        select(pesquisa_uploads.c.id_upload)
        .where(pesquisa_uploads.c.hash_conteudo == digest)
    ).first()
    if ja:
        info["status"] = "Duplicado"
        info["avisos"].append(f"conteúdo idêntico ao upload #{ja[0]}, já importado")
        return info

    if vistos is not None:
        vistos[digest] = nome_arquivo

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    destino = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{digest[:8]}.xlsx"
    try:
        destino.write_bytes(conteudo)
    except OSError as exc:
        info["status"] = "Erro"
        info["avisos"].append(f"não consegui salvar o arquivo ({exc})")
        return info
    info["arquivo_disco"] = destino.name

    try:
        wb = load_workbook(destino, data_only=True)
    except Exception:
        info["status"] = "Erro"
        info["avisos"].append(
            "não consegui abrir como Excel. Se for um .xls antigo, "
            "abra no Excel e salve como .xlsx.")
        return info

    # tipo: primeiro pelo nome; se não der, pelas colunas da planilha
    tipo = detectar_tipo(nome_arquivo)
    if not tipo:
        tipo = detectar_tipo_por_conteudo(wb)
        if tipo:
            info["avisos"].append(
                "tipo identificado pelas colunas — o nome do arquivo não indicava")
    if not tipo:
        info["status"] = "Erro"
        info["avisos"].append(
            "não reconheci o tipo nem pelo nome nem pelas colunas da planilha")
        return info

    info["tipo"] = tipo
    info["tipo_label"] = dict(TIPOS_PESQUISA).get(tipo, tipo)

    ids_etapa = detectar_etapas(nome_arquivo, etapas) or detectar_etapa_por_conteudo(wb, etapas)
    id_etapa = ids_etapa[0] if ids_etapa else 0
    info["id_etapa"] = id_etapa

    respostas, linhas, avisos = percorrer_planilha(wb, tipo, id_etapa, etapas)
    info["linhas"] = linhas
    info["respostas"] = len(respostas)
    info["avisos"].extend(avisos)

    tem_etapa_na_planilha = any(r["id_etapa"] for r in respostas)
    if id_etapa:
        info["etapa_label"] = next(
            (e.nome_etapa for e in etapas if e.id_etapa == id_etapa), None)
    elif tem_etapa_na_planilha:
        info["etapa_label"] = "Pela coluna ETAPA"
    else:
        info["etapa_label"] = None
        info["status"] = "Etapa indefinida"
        info["avisos"].append("não identifiquei a etapa pelo nome nem por coluna")

    if not respostas and info["status"] == "OK":
        info["status"] = "Vazio"
        info["avisos"].append("nenhuma resposta encontrada")

    return info


def gravar_arquivo(db, nome_arquivo, caminho_disco, tipo_pesquisa, id_etapa, etapas):
    """Grava de fato as respostas de um arquivo ja analisado."""
    caminho = UPLOAD_DIR / caminho_disco
    if not caminho.exists():
        return False, f"{nome_arquivo}: arquivo temporário não encontrado, envie de novo"

    try:
        wb = load_workbook(caminho, data_only=True)
    except Exception as exc:
        return False, f"{nome_arquivo}: não consegui abrir o Excel ({exc})"

    digest = hashlib.sha1(caminho.read_bytes()).hexdigest()
    ja = db.execute(
        select(pesquisa_uploads.c.id_upload)
        .where(pesquisa_uploads.c.hash_conteudo == digest)
    ).first()
    if ja:
        return False, f"{nome_arquivo}: já importado (upload #{ja[0]})"

    respostas, total_linhas, _ = percorrer_planilha(wb, tipo_pesquisa, id_etapa, etapas)

    result = db.execute(
        insert(pesquisa_uploads).values(
            id_etapa=id_etapa or 0,
            tipo_pesquisa=tipo_pesquisa,
            arquivo_nome=nome_arquivo,
            abas=json.dumps(wb.sheetnames, ensure_ascii=False),
            qtd_linhas=total_linhas,
            qtd_respostas=len(respostas),
            status="Importado",
            hash_conteudo=digest,
            criado_em=datetime.utcnow(),
        )
    )
    id_upload = result.inserted_primary_key[0]

    for r in respostas:
        id_autonomo, id_piloto, status_map = aplicar_mapeamento(
            db, r["tipo_alvo"], r["alvo_nome"])
        db.execute(insert(pesquisa_respostas).values(
            id_upload=id_upload,
            id_autonomo=id_autonomo,
            id_piloto=id_piloto,
            status_mapeamento=status_map,
            criado_em=datetime.utcnow(),
            **r,
        ))

    rotulo = dict(TIPOS_PESQUISA).get(tipo_pesquisa, tipo_pesquisa)
    return True, f"{nome_arquivo}: {rotulo} — {total_linhas} linhas, {len(respostas)} respostas"


@router.get("/pesquisas/{id_upload}")
def pesquisas_detalhe(id_upload: int, request: Request, db: Session = Depends(get_db)):
    upload = db.execute(
        select(pesquisa_uploads).where(pesquisa_uploads.c.id_upload == id_upload)
    ).mappings().first()

    if not upload:
        return redirect_with_message("/pesquisas", error="Upload não encontrado.")

    resumo = db.execute(
        text("""
            SELECT
                grupo_pergunta,
                COUNT(*) AS qtd,
                AVG(nota_num) AS media_nota
            FROM pesquisa_respostas
            WHERE id_upload = :id_upload
            GROUP BY grupo_pergunta
            ORDER BY qtd DESC
        """),
        {"id_upload": id_upload},
    ).mappings().all()

    respostas = db.execute(
        text("""
            SELECT *
            FROM pesquisa_respostas
            WHERE id_upload = :id_upload
            ORDER BY status_mapeamento DESC, alvo_nome, linha_excel, id_resposta
            LIMIT 500
        """),
        {"id_upload": id_upload},
    ).mappings().all()

    pendencias = db.execute(
        text("""
            SELECT
                tipo_alvo,
                alvo_nome,
                alvo_chave,
                COUNT(*) AS qtd
            FROM pesquisa_respostas
            WHERE id_upload = :id_upload
              AND status_mapeamento = 'Pendente'
              AND alvo_chave IS NOT NULL
              AND alvo_chave <> ''
            GROUP BY tipo_alvo, alvo_nome, alvo_chave
            ORDER BY qtd DESC, alvo_nome
        """),
        {"id_upload": id_upload},
    ).mappings().all()

    return templates.TemplateResponse(
        "pesquisas/detalhe.html",
        {
            "request": request,
            "upload": upload,
            "resumo": resumo,
            "respostas": respostas,
            "pendencias": pendencias,
            **options(db),
            **flash_from_request(request),
        },
    )


@router.post("/pesquisas/mapear")
def pesquisas_mapear(
    texto_origem: str = Form(...),
    tipo_alvo: str = Form(...),
    id_autonomo: str = Form(""),
    id_piloto: str = Form(""),
    redirect_to: str = Form("/pesquisas"),
    db: Session = Depends(get_db),
):
    chave = normalizar(texto_origem)

    if not chave:
        return redirect_with_message(redirect_to, error="Texto de origem inválido.")

    id_autonomo_int = int(id_autonomo) if id_autonomo else None
    id_piloto_int = int(id_piloto) if id_piloto else None

    existe = db.execute(
        select(pesquisa_mapeamentos).where(
            and_(
                pesquisa_mapeamentos.c.tipo_alvo == tipo_alvo,
                pesquisa_mapeamentos.c.texto_chave == chave,
            )
        )
    ).mappings().first()

    if existe:
        db.execute(
            update(pesquisa_mapeamentos)
            .where(pesquisa_mapeamentos.c.id_mapeamento == existe["id_mapeamento"])
            .values(id_autonomo=id_autonomo_int, id_piloto=id_piloto_int)
        )
    else:
        db.execute(
            insert(pesquisa_mapeamentos).values(
                tipo_alvo=tipo_alvo,
                texto_origem=texto_origem,
                texto_chave=chave,
                id_autonomo=id_autonomo_int,
                id_piloto=id_piloto_int,
                criado_em=datetime.utcnow(),
            )
        )

    valores_update = {
        "status_mapeamento": "Mapeado",
        "id_autonomo": id_autonomo_int,
        "id_piloto": id_piloto_int,
    }

    db.execute(
        update(pesquisa_respostas)
        .where(
            and_(
                pesquisa_respostas.c.tipo_alvo == tipo_alvo,
                pesquisa_respostas.c.alvo_chave == chave,
            )
        )
        .values(**valores_update)
    )

    db.commit()

    return redirect_with_message(redirect_to, success="Mapeamento salvo e aplicado aos registros existentes.")


@router.post("/pesquisas/{id_upload}/reprocessar-mapeamentos")
def reprocessar_mapeamentos(request: Request, id_upload: int, db: Session = Depends(get_db)):
    if not _is_admin(request):
        return RedirectResponse("/?sem_acesso=pesquisas", status_code=303)
    respostas = db.execute(
        select(pesquisa_respostas).where(pesquisa_respostas.c.id_upload == id_upload)
    ).mappings().all()

    atualizados = 0

    for r in respostas:
        id_autonomo, id_piloto, status = aplicar_mapeamento(db, r["tipo_alvo"], r["alvo_nome"])

        if status == "Mapeado":
            db.execute(
                update(pesquisa_respostas)
                .where(pesquisa_respostas.c.id_resposta == r["id_resposta"])
                .values(
                    id_autonomo=id_autonomo,
                    id_piloto=id_piloto,
                    status_mapeamento="Mapeado",
                )
            )
            atualizados += 1

    db.commit()

    return redirect_with_message(
        f"/pesquisas/{id_upload}",
        success=f"Mapeamentos reaplicados. Registros atualizados: {atualizados}.",
    )
